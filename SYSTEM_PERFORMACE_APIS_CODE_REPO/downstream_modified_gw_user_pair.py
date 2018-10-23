#!/usr/bin/python
"""Script for Layer 1 automation"""

import sys
import os
import subprocess
import argparse
import json
import paramiko
import socket
import ipywidgets as widgets
from collections import OrderedDict
#from test_python_script_network import get_max_bw_channel_and_beam_id

#matplotlib notebook
import pandas as pd
#import matplotlib.pyplot as plt
from ipywidgets import *
from IPython.display import display
#from notebook import display
from IPython.html import widgets
from openpyxl import Workbook
try:
    import openpyxl, platform
except:
    os.system("pip install platform")
    #if platform.system() == "Linux":
    #    os.system("pip install openpyxl")

try:
    import time
except:
    import platform
    platform_test = platform.system()
    #if platform_test == "Linux":
    #    os.system("pip install time")

try:
    from requests.auth import HTTPBasicAuth
except:
    os.system("pip install requests")
try:
    from ssh import Ssh_Console
except ImportError:
    import platform
    platform_test = platform.system()
    print "I am here", platform_test
    #if platform_test == "Windows":
    #    os.system("pip install ssh")
try:
    from ssh import Ssh_Console_Client
except ImportError:
    import platform
    platform_test_build = platform.system()        
    #if platform_test_build == "Windows":
    #    os.system("pip install ssh")

import xlwt
from xlwt import Workbook
import test_python_script_network
import numpy as np


def write_access_log(ip, command, cmd_ouput):
    import platform
    if platform.system() == "Windows": 
        with open("C:/Windows/Temp/layer1_access.log", "a") as f:
            header = "------ Node: " + ip + "------- Command: " + command + " -----------\n"
            f.write(header)
            f.write(cmd_ouput)
    elif platform.system() == "Linux":
        with open("/tmp/layer1_access.log", "a") as f:
            header = "------ Node: " + ip + "------- Command:" + command + "-------------\n"
            f.write(header)
            f.write(cmd_output)

def ssh_session(ip, USERNAME, PASSWORD):
   #import paramiko
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip, username=USERNAME, password=PASSWORD, timeout=20)
        return ssh
    except (paramiko.BadHostKeyException, paramiko.AuthenticationException, \
            paramiko.SSHException, socket.error) as e:
        print "SSHerror reaching Terminal[{0}]: {1}".format(ip, e)
        # add error log which says skipping the particular node as it is unreachable

def send_string_and_wait(shell, command, wait_time, should_print):
    # Send the su command
    shell.send(command)

    # Wait a bit, if necessary
    time.sleep(wait_time)

    # Flush the receive buffer
    receive_buffer = shell.recv(1024)

    # Print the receive buffer, if necessary
    if should_print:
        print receive_buffer

def send_string_and_wait_for_string(shell, command, wait_string, should_print):
    # Send the su command
    shell.send(command)

    time.sleep(1)

    # Create a new receive buffer
    receive_buffer = ""

    while not wait_string in receive_buffer:
        # Flush the receive buffer
        receive_buffer += shell.recv(1024)
        print receive_buffer

    # Print the receive buffer, if necessary
    if should_print:
        print receive_buffer

def stdin_write(stdin, command):
    cmd_stdin = command + "\n"
    stdin.flush()
    time.sleep(0.5)
    stdin.write(cmd_stdin)

def telnet_run_command(ip, username, password, port, command, remote=None):
    before_xoff = []
    test_b = ""
    try:
        ssh = ssh_session(ip, username, password)
    
        #print "The object type is",    type(ssh)

        import paramiko
    
    except:
        #os.system("pip install paramiko")
        ssh =ssh_session(ip, username, password)

    if ssh is not None:
        cmd_telnet = "telnet 0 " + port
        stdin, stdout, stderr = ssh.exec_command(cmd_telnet)
        if(remote):
            cmd_admin_login = ["admin", "iDirect", "xoff", "rmt " + remote, command, "exit"]

        cmd_admin_login = ["admin", "iDirect", "xoff", command, "exit"]
        for cmd in cmd_admin_login:
            # if(cmd == command):
                # time.sleep(3)
            stdin_write(stdin, cmd)
        print("I am here before ssh")
        ssh.close()
        print("I am here")
        return stdout.read()
    
    return "Error reaching node: " + str(ip)

def find_command_output_body(cmd_ouput, command):
    cmd_ouput_body = []
    flag = 0
    for line in cmd_ouput.split("\n"):
        if "xoff" in line:
            flag += 1
        if command in line:
            flag += 1

        if flag == 2:
            cmd_ouput_body.append(line)

    return "\n".join(cmd_ouput_body)

def get_terminal_did_list(ip, username, password, tpa_port):
    command = "rmt list"
    cmd_output = telnet_run_command(ip, username, password, tpa_port, command)
    cmd_ouput_body = find_command_output_body(cmd_output, command)
    print("<<<THE COMMAND OUTPUT BODY>>>", cmd_ouput_body)
    
    cmd_ouput_body_split = cmd_ouput_body.splitlines()
    print("The command output>>>>", cmd_ouput_body_split)

    term_list = []
    
    for line in cmd_ouput_body_split: 
        if('CX' in line):
            term_list.append(line.split('(')[1].split(')')[0])
    
    #term_list_in = cmd_ouput_body_split[2]
    #term_list_in = term_list_in[term_list_in.find("("): term_list_in.rfind(")")]

    #term_list_fi = term_list_in[6:]
    #print "THE TERMINAL LIST", term_list_in
    #print "THE TERMINAL LIST", term_list_fi
    
    #term_list.append(term_list_fi)

    #print("THE TERM LIST>>>", term_list)
    
    return term_list


#def excel_sheet_generation_for_crc()
row1=1
def excel_sheet_generation(dictionary_set, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, active_object_state, work_s2, SYMBOLS):
    #row = 1
    global row1
    try:
        work_b = openpyxl.Workbook()
        work_s = work_b.active
        #work_s1 = work_b.create_sheet()
        
#        if column
        for i in zip(value):
            #work_s['A%s' %(i)] = cmd_input_for_excel 
            work_s.append(i)
         #   work_s1.append(j)
            #work_s['Command Output'] = cmd_output
            test_work = work_b.save("xlwt c:\\Users\\akarthik\\Downloads\\Text1.xls")
    except: 
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.chart import (LineChart, ScatterChart, Reference)
        from openpyxl.chart.axis import DateAxis
        from openpyxl.chart.layout import Layout, ManualLayout
        import copy

    finally:
        #print("The value of row", row)
        print("The set of keys are given by", dictionary_set.keys())
        print("The set of values are given by", dictionary_set.values())

        print("<<<<<THE MODCOD LIST LENGTH IS GIVEN>>>>", modcod_list_length) 
        work_s = active_object_state.active
        
        #work_s2 = active_object_state.get_sheet_by_name("Chart")

        #work_s1 = work_b.create_sheet(
        cell1 = work_s.cell(row=1, column=1)
       # wb = copy(rb) # a writable copy (I can't read values out of this, only write to it)
       # w_sheet = wb.get_sheet(0)
        #i = 0
        #j = 0 
        
        dictionary_list_values = [] 

        for i in range(0, len(dictionary_set.values())-1, 1):
            dictionary_values_set = ''.join(dictionary_set.values()[i])
            dictionary_list_values.append(dictionary_values_set)
        
        print (dictionary_list_values)
       # print("COUNTER", counter)

        for n, i in enumerate(dictionary_list_values):
            if i == "":
                dictionary_list_values[n] = "_"
        for item in range(0, len(dictionary_list_values)-1, 1):
            if dictionary_list_values[item] != "_":
                dictionary_list_values[item] = int(dictionary_list_values[item])
                print(type(dictionary_list_values[item]))            
        
        print (dictionary_list_values) 

        if not dictionary_list_combined:
            print(type(dictionary_list_combined))
            print("The contents of the str object", dictionary_list_combined) 
            dictionary_list_combined.append(dictionary_set.keys())
            dictionary_list_combined.append(dictionary_list_values)

        else:
            dictionary_list_combined.append(dictionary_list_values)


        print ("The dictionary list of keys and values", dictionary_list_combined)
        
        if cell1.value == None:
            i=0
            print("I am inside the for loop for the first row execution")
            #raw_input("BEFORE THE FOR LOOP FIRST ROW EXECUTION....")
            for key in range(1, 34):
                #work_s['A%s' %(i)] = cmd_input_for_excel 
                cell_ref = work_s.cell(row = row1, column = key)
                cell_ref.value = dictionary_set.keys()[i]
                i+=1

        if not cell1.value == None:
            print("I am here when cell value is not none")
            """
            for n, i in enumerate(dictionary_list_values):
            if i == "":
                dictionary_list_values[n] = "_"
            for item in range(0, len(dictionary_list_values)-1, 1):
            if dictionary_list_values[item] != "_":
                dictionary_list_values[item] = int(dictionary_list_values[item])
                print(type(dictionary_list_values[item]))            
        
            print (dictionary_list_values) 
            """
            
            print("j when 0 : ", dictionary_list_values[0])
            print("j when checked for last element : ", dictionary_list_values[-1])
            j=0
        #for row_counter in range(0, len(dictionary_list_values)-1, 1):
            if dictionary_list_values.count("_") == 33:
                del dictionary_list_values
     
            else:
                print("row value before the for loop", row1)
                #raw_input("IN ELSE BEFORE FOR....")
                row1+=1
                for col in range(1, 34):
                    cell_ref_col = work_s.cell(row = row1, column = col)
                    cell_ref_col.value = dictionary_list_values[j]
                    j+=1
            #row1+=1
                print("row value after the for loop for col is :", row1)
                #raw_input("ROW IN ELSE.....")
                
            #   work_s1.append(j)
                #work_s['Command Output'] = cmd_output
                 
        """This is the code section for the line chart"""

            
        


            #for row in dictionary_list_combined:
            #    work_s.append(row)
        elim_zero_value_list = []
        ByteTotal_Mode_qpsk_1_4 = []
        ByteTotal_Mode_qpsk_1_3 = []
        ByteTotal_Mode_qpsk_2_5 = []
        ByteTotal_Mode_qpsk_1_2 = []
        ByteTotal_Mode_qpsk_3_5 = []
        ByteTotal_Mode_qpsk_2_3 = []
        ByteTotal_Mode_qpsk_3_4 = []
        ByteTotal_Mode_qpsk_4_5 = []
        ByteTotal_Mode_qpsk_5_6 = []
        ByteTotal_Mode_qpsk_8_9 = []
        ByteTotal_Mode_qpsk_9_10= []
        ByteTotal_Mode_8psk_3_5 = []
        ByteTotal_Mode_8psk_2_3 = []
        ByteTotal_Mode_8psk_3_4 = []
        ByteTotal_Mode_8psk_5_6 = []
        ByteTotal_Mode_8psk_8_9 = []
        ByteTotal_Mode_8psk_9_10 = []
        ByteTotal_Mode_16apsk_2_3 = []
        ByteTotal_Mode_16apsk_3_4 = []
        ByteTotal_Mode_16apsk_4_5 = []
        ByteTotal_Mode_16apsk_5_6 = []
        ByteTotal_Mode_16apsk_8_9= []
        ByteTotal_Mode_16apsk_9_10 = []
        ByteTotal_Mode_32apsk_3_4 = []
        ByteTotal_Mode_32apsk_4_5 = []
        ByteTotal_Mode_32apsk_5_6 = []
        ByteTotal_Mode_32apsk_8_9 = []
        ByteTotal_Mode_32apsk_9_10 = []
        bBFrameCount = []
        legsTotalCount = []
        legsPacketCount = []
        fastFadeCount = []
        modCodChangeCount = []
        #final_value_list = []

        #dictionary_chart = {"chart1" :1 , "chart2" : 2, "chart3" :3, "chart4" : 4, "chart5" : 5, "chart6" : 6, "chart7" : 7, "chart8" : 8, "chart9" : 9, "chart10" : 10, "chart11" : 11, "chart12"}
        for chart in range(1, 34):
            
            
            # Style the lines
            #s1 = c.series[0]
            #s1.marker.symbol = "triangle"
            #s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            #s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            #s1.graphicalProperties.line.noFill = True

            """
            s2 = c.series[1]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FF9900" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FF9900" # Marker outline
                #s2.graphicalProperties.line.width = 100050 # width in EMUs

            s2 = c.series[2]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[3]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FF6A00" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FF6A00" # Marker outline

            s2 = c.series[4]
            s2.graphicalProperties.line.solidFill = "FFB200"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[5]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FFE100" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FFE100" # Marker outline

            s2 = c.series[6]
            s2.graphicalProperties.line.solidFill = "E9FF00"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[7]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "8CFF00" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "8CFF00" # Marker outline

            s2 = c.series[8]
            s2.graphicalProperties.line.solidFill = "00FF61"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[9]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "00FFA5" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "00FFA5" # Marker outline

            s2 = c.series[10]
            s2.graphicalProperties.line.solidFill = "00FFF2"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[11]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "00D0FF" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "00D0FF" # Marker outline

            s2 = c.series[12]
            s2.graphicalProperties.line.solidFill = "003FFF"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[13]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "5D00FF" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "5D00FF" # Marker outline

            s2 = c.series[14]
            s2.graphicalProperties.line.solidFill = "AA00FF"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[15]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FF00A1" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FF00A1" # Marker outline

            s2 = c.series[16]
            s2.graphicalProperties.line.solidFill = "FF6A00"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[17]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "6E00A9" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "6E00A9" # Marker outline

            s2 = c.series[18]
            s2.graphicalProperties.line.solidFill = "0079A9"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[19]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "00A902" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "00A902" # Marker outline

            s2 = c.series[20]
            s2.graphicalProperties.line.solidFill = "0093A9"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[21]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "0082A9" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "0082A9" # Marker outline

            s2 = c.series[22]
            s2.graphicalProperties.line.solidFill = "0074A9"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[23]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "0054A9" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "0054A9" # Marker outline

            s2 = c.series[24]
            s2.graphicalProperties.line.solidFill = "0038A9"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[25]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "000EA9" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "000EA9" # Marker outline

            s2 = c.series[26]
            s2.graphicalProperties.line.solidFill = "6000A9"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[27]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "9300A9" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "9300A9" # Marker outline

            s2 = c.series[28]
            s2.graphicalProperties.line.solidFill = "A90093"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[29]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "A90057" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "A90057" # Marker outline

            s2 = c.series[30]
            s2.graphicalProperties.line.solidFill = "A9003B"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[31]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "A90022" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "A90022" # Marker outline

            s2 = c.series[32]
            s2.graphicalProperties.line.solidFill = "A90000"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c.series[33]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "A9007F" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "A9007F" # Marker outline

            """ 
        # s = "s" 
        # for k in range(3, 34):
        #     var = s +"%s" %(k)
        #     var = c.series[k]
        #     var.graphicalProperties.line.solidFill = "00AAAA"
        #     var.graphicalProperties.line.dashStyle = "sysDot"  
                
                        
                #work_s2 = copy.deepcopy(work_s) 
            #c.layout=Layout(
            #    manualLayout=ManualLayout(
            #        x=0.25, y=0.25,
            #        h=0.5, w=0.5,
            #    )
            #)
            row_counter_data = 1
            row_counter_squared = 1
            for row in work_s.get_squared_range(min_col=chart, min_row=1, max_col=chart, max_row=modcod_list_length+1):
                print("ROW SQUARED DATA>>>>", row_counter_squared)
                for cell in row:
                    #print "The element given is", row[cell]
                    #row[cell] = row[cell].encode("utf-8")
                    print("THE ROW COUNTER DATA", row_counter_data)
                    row_counter_data+=1
                    elim_zero_value_list.append(cell.value)
                    print("THE ELIM ZERO VALUE LIST", elim_zero_value_list)

            

        #elim_value_list = elim_zero_value_list
                if row_counter_squared == 11:
                     final_value_list = elim_zero_value_list 
                row_counter_squared+=1
            
            """
            if chart == 1:
                work_s2.add_chart(c, "A2")
            if chart == 2:# and ByteTotal_Mode_qpsk_1_3.count("_") !=5: 
                work_s2.add_chart(c, "J2")
            if chart == 3:# and ByteTotal_Mode_qpsk_2_5.count("_") !=5: 
                work_s2.add_chart(c, "S2")
            if chart == 4:# and ByteTotal_Mode_qpsk_1_2.count("_") != 5: 
                work_s2.add_chart(c, "AB2")
            if chart == 5:#and ByteTotal_Mode_qpsk_3_5.count("_") != 5: 
                work_s2.add_chart(c, "A20")
            if chart == 6:# and ByteTotal_Mode_qpsk_2_3.count("_") != 5: 
                work_s2.add_chart(c, "J20")
            if chart == 7:#and ByteTotal_Mode_qpsk_3_4.count("_") != 5: 
                work_s2.add_chart(c, "S20")
            if chart == 8:#and ByteTotal_Mode_qpsk_4_5.count("_") != 5: 
                work_s2.add_chart(c, "AB20")
            if chart == 9:#and ByteTotal_Mode_qpsk_5_6.count("_") != 5: 
                work_s2.add_chart(c, "A40")
            if chart == 10:# and ByteTotal_Mode_8psk_8_9.count("_") != 5 : 
                work_s2.add_chart(c, "J40")
            if chart == 11:# and ByteTotal_Mode_qpsk_9_10.count("_") != 5: 
                work_s2.add_chart(c, "S40")
            if chart == 12:# and ByteTotal_Mode_8psk_3_5.count("_") !=5: 
                work_s2.add_chart(c, "AB40")
            if chart == 13:# and ByteTotal_Mode_8psk_2_3.count("_")!=5: 
                work_s2.add_chart(c, "A60")
            if chart == 14:# and ByteTotal_Mode_8psk_3_4.count("_")!=5: 
                work_s2.add_chart(c, "J60")
            if chart == 15:# and ByteTotal_Mode_8psk_5_6.count("_")!=5: 
                work_s2.add_chart(c, "S60")
            if chart == 16:# and ByteTotal_Mode_8psk_8_9.count("_")!=5: 
                work_s2.add_chart(c, "AB60")
            if chart == 17:# and ByteTotal_Mode_8psk_9_10.count("_")!=5: 
                work_s2.add_chart(c, "A80")
            if chart == 18:# and ByteTotal_Mode_16apsk_2_3.count("_")!=5: 
                work_s2.add_chart(c, "J80")
            if chart == 19:# and ByteTotal_Mode_16apsk_3_4.count("_")!=5: 
                work_s2.add_chart(c, "S80")
            if chart == 20:#and ByteTotal_Mode_16apsk_4_5.count("_")!=5: 
                work_s2.add_chart(c, "AB80")
            if chart == 21:# and ByteTotal_Mode_16apsk_5_6.count("_")!=5: 
                work_s2.add_chart(c, "A100")
            if chart == 22:# and ByteTotal_Mode_16apsk_8_9.count("_")!=5 : 
                work_s2.add_chart(c, "J100")
            if chart == 23:# and ByteTotal_Mode_16apsk_9_10.count("_")!=5: 
                work_s2.add_chart(c, "S100")
            if chart == 24:# and ByteTotal_Mode_32apsk_3_4.count("_")!=5: 
                work_s2.add_chart(c, "AB100")
            if chart == 25:# and ByteTotal_Mode_32apsk_4_5.count("_")!=5: 
                work_s2.add_chart(c, "A120")
            if chart == 26:# and ByteTotal_Mode_32apsk_5_6.count("_")!=5: 
                work_s2.add_chart(c, "J120")
            if chart == 27:# and ByteTotal_Mode_32apsk_8_9.count("_")!=5: 
                work_s2.add_chart(c, "S120")
            if chart == 28:# and ByteTotal_Mode_32apsk_9_10.count("_")!=5: 
                work_s2.add_chart(c, "AB120")
            if chart == 29:# and bBFrameCount.count("_")!=5: 
                work_s2.add_chart(c, "A140")
            if chart == 30:# and legsTotalCount.count("_")!=5: 
                work_s2.add_chart(c, "J140")
            if chart == 31:# and legsPacketCount.count("_")!=5: 
                work_s2.add_chart(c, "S140")
                work_s2.
            if chart == 32:# and fastFadeCount.count("_")!=5: 
                work_s2.add_chart(c, "AB140")
            if chart == 33:# and modCodChangeCount.count("_")!=5: 
                work_s2.add_chart(c, "A160")
            """
        #new_elim_list = np.array(final_value_list)
        #print "This is the elimated zero value list elim list", new_elim_list
            ##raw_input("THE ARRAY IS??>>>>>..")
            #new_elim_list_test = new_elim_list.reshape(33, 11)
        #print("THE LIST WITH 2D to 1D is", new_elim_list)
            
           
         
            
        



        #print("THE LIST AFTER RESHAPE", new_elim_list_test)  
        print("THE LIST BEFORE TH SYMBOL LIST", final_value_list)
        new_elim_list1 = np.array(final_value_list)
        #print "This is the elimated zero value list elim list", new_elim_list
        ##raw_input("THE ARRAY IS??>>>>>..")
        new_elim_list_test = new_elim_list1.reshape(-1, 11)
        #new_elim_list_test = final_value_list.reshape(33, 11)

        ByteTotal_Mode_qpsk_1_4.append(new_elim_list_test[0])
        #ByteTotal_Mode_8psk_2_3.append(new_elim_list_test[12])

        #print("I AM HERE ELIM>>>", new_elim_list[0][1]) 
        #print("I AM HERE ELIM>>>", new_elim_list[0][2]) 
        #print("I AM HERE ELIM>>>", new_elim_list[0][3]) 
        
        ByteTotal_Mode_qpsk_1_3.append(new_elim_list_test[1])
        ByteTotal_Mode_qpsk_2_5.append(new_elim_list_test[2])
        ByteTotal_Mode_qpsk_1_2.append(new_elim_list_test[3])
        ByteTotal_Mode_qpsk_3_5.append(new_elim_list_test[4])
        ByteTotal_Mode_qpsk_2_3.append(new_elim_list_test[5])
        ByteTotal_Mode_qpsk_3_4.append(new_elim_list_test[6])
        ByteTotal_Mode_qpsk_4_5.append(new_elim_list_test[7])
        ByteTotal_Mode_qpsk_5_6.append(new_elim_list_test[8])
        ByteTotal_Mode_qpsk_8_9.append(new_elim_list_test[9])
        ByteTotal_Mode_qpsk_9_10.append(new_elim_list_test[10])
        
        ByteTotal_Mode_8psk_3_5.append(new_elim_list_test[11])
        ByteTotal_Mode_8psk_2_3.append(new_elim_list_test[12]) 
        ByteTotal_Mode_8psk_3_4.append(new_elim_list_test[13])
        ByteTotal_Mode_8psk_5_6.append(new_elim_list_test[14])
        ByteTotal_Mode_8psk_8_9.append(new_elim_list_test[15])
        ByteTotal_Mode_8psk_9_10.append(new_elim_list_test[16])
        
        ByteTotal_Mode_16apsk_2_3.append(new_elim_list_test[17])
        ByteTotal_Mode_16apsk_3_4.append(new_elim_list_test[18])
        ByteTotal_Mode_16apsk_4_5.append(new_elim_list_test[19])
        ByteTotal_Mode_16apsk_5_6.append(new_elim_list_test[20])
        ByteTotal_Mode_16apsk_8_9.append(new_elim_list_test[21])
        ByteTotal_Mode_16apsk_9_10.append(new_elim_list_test[22])
        
        ByteTotal_Mode_32apsk_3_4.append(new_elim_list_test[23])
        ByteTotal_Mode_32apsk_4_5.append(new_elim_list_test[24])
        ByteTotal_Mode_32apsk_5_6.append(new_elim_list_test[25])
        ByteTotal_Mode_32apsk_8_9.append(new_elim_list_test[26])
        ByteTotal_Mode_32apsk_9_10.append(new_elim_list_test[27])
        bBFrameCount.append(new_elim_list_test[28])
        legsTotalCount.append(new_elim_list_test[29])
        legsPacketCount.append(new_elim_list_test[30])
        fastFadeCount.append(new_elim_list_test[31])
        modCodChangeCount.append(new_elim_list_test[32])
        
        ByteTotal_Mode_qpsk_1_4 = np.array(ByteTotal_Mode_qpsk_1_4)
        ByteTotal_Mode_qpsk_1_4 = ByteTotal_Mode_qpsk_1_4.ravel()
        ByteTotal_Mode_qpsk_1_4 = ByteTotal_Mode_qpsk_1_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_1_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_1_3 = np.array(ByteTotal_Mode_qpsk_1_3)
        ByteTotal_Mode_qpsk_1_3 = ByteTotal_Mode_qpsk_1_3.ravel()
        ByteTotal_Mode_qpsk_1_3 = ByteTotal_Mode_qpsk_1_3.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_1_3)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_2_5 = np.array(ByteTotal_Mode_qpsk_2_5)
        ByteTotal_Mode_qpsk_2_5 = ByteTotal_Mode_qpsk_2_5.ravel()
        ByteTotal_Mode_qpsk_2_5 = ByteTotal_Mode_qpsk_2_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_2_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_1_2 = np.array(ByteTotal_Mode_qpsk_1_2)
        ByteTotal_Mode_qpsk_1_2 = ByteTotal_Mode_qpsk_1_2.ravel()
        ByteTotal_Mode_qpsk_1_2 = ByteTotal_Mode_qpsk_1_2.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_1_2)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_3_5 = np.array(ByteTotal_Mode_qpsk_3_5)
        ByteTotal_Mode_qpsk_3_5 = ByteTotal_Mode_qpsk_3_5.ravel()
        ByteTotal_Mode_qpsk_3_5 = ByteTotal_Mode_qpsk_3_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_3_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_2_3 = np.array(ByteTotal_Mode_qpsk_2_3)
        ByteTotal_Mode_qpsk_2_3 = ByteTotal_Mode_qpsk_2_3.ravel()
        ByteTotal_Mode_qpsk_2_3 = ByteTotal_Mode_qpsk_2_3.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_2_3)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_3_4 = np.array(ByteTotal_Mode_qpsk_3_4)
        ByteTotal_Mode_qpsk_3_4 = ByteTotal_Mode_qpsk_3_4.ravel()
        ByteTotal_Mode_qpsk_3_4 = ByteTotal_Mode_qpsk_3_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_3_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_4_5 = np.array(ByteTotal_Mode_qpsk_4_5)
        ByteTotal_Mode_qpsk_4_5 = ByteTotal_Mode_qpsk_4_5.ravel()
        ByteTotal_Mode_qpsk_4_5 = ByteTotal_Mode_qpsk_4_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_4_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_5_6 = np.array(ByteTotal_Mode_qpsk_5_6)
        ByteTotal_Mode_qpsk_5_6 = ByteTotal_Mode_qpsk_5_6.ravel()
        ByteTotal_Mode_qpsk_5_6 = ByteTotal_Mode_qpsk_5_6.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_5_6)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_8_9 = np.array(ByteTotal_Mode_qpsk_8_9)
        ByteTotal_Mode_qpsk_8_9 = ByteTotal_Mode_qpsk_8_9.ravel()
        ByteTotal_Mode_qpsk_8_9 = ByteTotal_Mode_qpsk_8_9.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_8_9)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_1_4 = np.array(ByteTotal_Mode_qpsk_1_4)
        ByteTotal_Mode_qpsk_1_4 = ByteTotal_Mode_qpsk_1_4.ravel()
        ByteTotal_Mode_qpsk_1_4 = ByteTotal_Mode_qpsk_1_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_1_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_qpsk_9_10 = np.array(ByteTotal_Mode_qpsk_9_10)
        ByteTotal_Mode_qpsk_9_10 = ByteTotal_Mode_qpsk_9_10.ravel()
        ByteTotal_Mode_qpsk_9_10 = ByteTotal_Mode_qpsk_9_10.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_qpsk_9_10)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_3_5 = np.array(ByteTotal_Mode_8psk_3_5)
        ByteTotal_Mode_8psk_3_5 = ByteTotal_Mode_8psk_3_5.ravel()
        ByteTotal_Mode_8psk_3_5 = ByteTotal_Mode_8psk_3_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_3_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_2_3 = np.array(ByteTotal_Mode_8psk_2_3)
        ByteTotal_Mode_8psk_2_3 = ByteTotal_Mode_8psk_2_3.ravel()
        ByteTotal_Mode_8psk_2_3 = ByteTotal_Mode_8psk_2_3.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_2_3)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_3_4 = np.array(ByteTotal_Mode_8psk_3_4)
        ByteTotal_Mode_8psk_3_4 = ByteTotal_Mode_8psk_3_4.ravel()
        ByteTotal_Mode_8psk_3_4 = ByteTotal_Mode_8psk_3_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_3_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_5_6 = np.array(ByteTotal_Mode_8psk_5_6)
        ByteTotal_Mode_8psk_5_6 = ByteTotal_Mode_8psk_5_6.ravel()
        ByteTotal_Mode_8psk_5_6 = ByteTotal_Mode_8psk_5_6.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_5_6)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_8_9 = np.array(ByteTotal_Mode_8psk_8_9)
        ByteTotal_Mode_8psk_8_9 = ByteTotal_Mode_8psk_8_9.ravel()
        ByteTotal_Mode_8psk_8_9 = ByteTotal_Mode_8psk_8_9.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_8_9)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_8psk_9_10 = np.array(ByteTotal_Mode_8psk_9_10)
        ByteTotal_Mode_8psk_9_10 = ByteTotal_Mode_8psk_9_10.ravel()
        ByteTotal_Mode_8psk_9_10 = ByteTotal_Mode_8psk_9_10.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_8psk_9_10)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_2_3 = np.array(ByteTotal_Mode_16apsk_2_3)
        ByteTotal_Mode_16apsk_2_3 = ByteTotal_Mode_16apsk_2_3.ravel()
        ByteTotal_Mode_16apsk_2_3 = ByteTotal_Mode_16apsk_2_3.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_2_3)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_3_4 = np.array(ByteTotal_Mode_16apsk_3_4)
        ByteTotal_Mode_16apsk_3_4 = ByteTotal_Mode_16apsk_3_4.ravel()
        ByteTotal_Mode_16apsk_3_4 = ByteTotal_Mode_16apsk_3_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_3_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_4_5 = np.array(ByteTotal_Mode_16apsk_4_5)
        ByteTotal_Mode_16apsk_4_5 = ByteTotal_Mode_16apsk_4_5.ravel()
        ByteTotal_Mode_16apsk_4_5 = ByteTotal_Mode_16apsk_4_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_4_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_5_6 = np.array(ByteTotal_Mode_16apsk_5_6)
        ByteTotal_Mode_16apsk_5_6 = ByteTotal_Mode_16apsk_5_6.ravel()
        ByteTotal_Mode_16apsk_5_6 = ByteTotal_Mode_16apsk_5_6.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_5_6)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_8_9 = np.array(ByteTotal_Mode_16apsk_8_9)
        ByteTotal_Mode_16apsk_8_9 = ByteTotal_Mode_16apsk_8_9.ravel()
        ByteTotal_Mode_16apsk_8_9 = ByteTotal_Mode_16apsk_8_9.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_8_9)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_16apsk_9_10 = np.array(ByteTotal_Mode_16apsk_9_10)
        ByteTotal_Mode_16apsk_9_10 = ByteTotal_Mode_16apsk_9_10.ravel()
        ByteTotal_Mode_16apsk_9_10 = ByteTotal_Mode_16apsk_9_10.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_16apsk_9_10)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_32apsk_3_4 = np.array(ByteTotal_Mode_32apsk_3_4)
        ByteTotal_Mode_32apsk_3_4 = ByteTotal_Mode_32apsk_3_4.ravel()
        ByteTotal_Mode_32apsk_3_4 = ByteTotal_Mode_32apsk_3_4.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_32apsk_3_4)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_32apsk_4_5 = np.array(ByteTotal_Mode_32apsk_4_5)
        ByteTotal_Mode_32apsk_4_5 = ByteTotal_Mode_32apsk_4_5.ravel()
        ByteTotal_Mode_32apsk_4_5 = ByteTotal_Mode_32apsk_4_5.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_32apsk_4_5)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_32apsk_5_6 = np.array(ByteTotal_Mode_32apsk_5_6)
        ByteTotal_Mode_32apsk_5_6 = ByteTotal_Mode_32apsk_5_6.ravel()
        ByteTotal_Mode_32apsk_5_6 = ByteTotal_Mode_32apsk_5_6.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_32apsk_5_6)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_32apsk_8_9 = np.array(ByteTotal_Mode_32apsk_8_9)
        ByteTotal_Mode_32apsk_8_9 = ByteTotal_Mode_32apsk_8_9.ravel()
        ByteTotal_Mode_32apsk_8_9 = ByteTotal_Mode_32apsk_8_9.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_32apsk_8_9)
        #raw_input("THE BYTEMODEQPSK...")

        ByteTotal_Mode_32apsk_9_10 = np.array(ByteTotal_Mode_32apsk_9_10)
        ByteTotal_Mode_32apsk_9_10 = ByteTotal_Mode_32apsk_9_10.ravel()
        ByteTotal_Mode_32apsk_9_10 = ByteTotal_Mode_32apsk_9_10.tolist()
        print("The array in 1 and 3", ByteTotal_Mode_32apsk_9_10)
        #raw_input("THE BYTEMODEQPSK...")

        bBFrameCount = np.array(bBFrameCount)
        bBFrameCount = bBFrameCount.ravel()
        bBFrameCount = bBFrameCount.tolist()
        print("The array in 1 and 3", bBFrameCount)
        #raw_input("THE BYTEMODEQPSK...")

        legsTotalCount = np.array(legsTotalCount)
        legsTotalCount = legsTotalCount.ravel()
        legsTotalCount = legsTotalCount.tolist()
        print("The array in 1 and 3", legsTotalCount)
        #raw_input("THE BYTEMODEQPSK...")

        legsPacketCount = np.array(legsPacketCount)
        legsPacketCount = legsPacketCount.ravel()
        legsPacketCount = legsPacketCount.tolist()
        print("The array in 1 and 3", legsPacketCount)
        #raw_input("THE BYTEMODEQPSK...")

        fastFadeCount = np.array(fastFadeCount)
        fastFadeCount = fastFadeCount.ravel()
        fastFadeCount = fastFadeCount.tolist()
        print("The array in 1 and 3", fastFadeCount)
        #raw_input("THE BYTEMODEQPSK...")
        
        modCodChangeCount = np.array(modCodChangeCount)
        modCodChangeCount = modCodChangeCount.ravel()
        modCodChangeCount = modCodChangeCount.tolist()
        print("The array in 1 and 3", modCodChangeCount)
        #raw_input("THE BYTEMODEQPSK...")


        print("The array in 1 and 3", ByteTotal_Mode_qpsk_1_3)
        print("The array in 8 for 2 and 3", ByteTotal_Mode_8psk_2_3)
        ##raw_input("BYTE 1 3...")
        
        for chrt in range(1, 34):
            c = LineChart()
            c.title = "Gateway Frequency = %s User Frequency = %s" %(gw, u_freq)
            c.style = 13
            c.y_axis.title = 'Value in Bytes'
            c.x_axis.title = 'Samples Interval'
            c.x_axis.scaling.min = 0
            #c.y_axis.scaling.min = 0
            c.x_axis.scaling.max = modcod_list_length
            c.x_axis.unit = 1
            #c.y_axis.scaling.max = 70000
            data = Reference(work_s, min_col=chrt, min_row=1, max_col=chrt, max_row=modcod_list_length+1)
            print ("THE DATA IS>>>>", data)
            c.add_data(data, titles_from_data=True)
            if chrt == 1 and ByteTotal_Mode_qpsk_1_4.count("_")!=row1-1:
                work_s2.add_chart(c, "A2")
            if chrt ==2 and ByteTotal_Mode_qpsk_1_3.count("_") !=row1-1: 
                work_s2.add_chart(c, "J2")
            if chrt ==3 and ByteTotal_Mode_qpsk_2_5.count("_") !=row1-1: 
                work_s2.add_chart(c, "S2")
            if chrt ==4 and ByteTotal_Mode_qpsk_1_2.count("_") != row1-1: 
                work_s2.add_chart(c, "AB2")
            if chrt ==5 and ByteTotal_Mode_qpsk_3_5.count("_") != row1-1: 
                work_s2.add_chart(c, "A20")
            if chrt ==6 and ByteTotal_Mode_qpsk_2_3.count("_") != row1-1: 
                work_s2.add_chart(c, "J20")
            if chrt ==7 and ByteTotal_Mode_qpsk_3_4.count("_") != row1-1: 
                work_s2.add_chart(c, "S20")
            if chrt ==8 and ByteTotal_Mode_qpsk_4_5.count("_") != row1-1: 
                work_s2.add_chart(c, "AB20")
            if chrt ==9 and ByteTotal_Mode_qpsk_5_6.count("_") != row1-1: 
                work_s2.add_chart(c, "A40")
            if chrt ==10 and  ByteTotal_Mode_qpsk_8_9.count("_") != row1-1 : 
                work_s2.add_chart(c, "J40")
            if chrt ==11 and ByteTotal_Mode_qpsk_9_10.count("_") != row1-1: 
                work_s2.add_chart(c, "S40")
            if chrt ==12 and ByteTotal_Mode_8psk_3_5.count("_") !=row1-1: 
                work_s2.add_chart(c, "AB40")
            if chrt ==13 and ByteTotal_Mode_8psk_2_3.count("_")!=row1-1: 
                work_s2.add_chart(c, "A60")
            if chrt ==14 and ByteTotal_Mode_8psk_3_4.count("_")!=row1-1: 
                work_s2.add_chart(c, "J60")
            if chrt ==15 and ByteTotal_Mode_8psk_5_6.count("_")!=row1-1: 
                work_s2.add_chart(c, "S60")
            if chrt ==16 and ByteTotal_Mode_8psk_8_9.count("_")!=row1-1: 
                work_s2.add_chart(c, "AB60")
            if chrt ==17 and ByteTotal_Mode_8psk_9_10.count("_")!=row1-1: 
                work_s2.add_chart(c, "A80")
            if chrt ==18 and ByteTotal_Mode_16apsk_2_3.count("_")!=row1-1: 
                work_s2.add_chart(c, "J80")
            if chrt ==19 and ByteTotal_Mode_16apsk_3_4.count("_")!=row1-1: 
                work_s2.add_chart(c, "S80")
            if chrt ==20 and ByteTotal_Mode_16apsk_4_5.count("_")!=row1-1: 
                work_s2.add_chart(c, "AB80")
            if chrt ==21 and ByteTotal_Mode_16apsk_5_6.count("_")!=row1-1: 
                work_s2.add_chart(c, "A100")
            if chrt ==22 and ByteTotal_Mode_16apsk_8_9.count("_")!=row1-1:
                work_s2.add_chart(c, "J100")
            if chrt ==23 and ByteTotal_Mode_16apsk_9_10.count("_")!=row1-1: 
                work_s2.add_chart(c, "S100")
            if chrt ==24 and ByteTotal_Mode_32apsk_3_4.count("_")!=row1-1: 
                work_s2.add_chart(c, "AB100")
            if chrt == 25 and ByteTotal_Mode_32apsk_4_5.count("_")!=row1-1: 
                work_s2.add_chart(c, "A120")
            if chrt ==26 and ByteTotal_Mode_32apsk_5_6.count("_")!=row1-1: 
                work_s2.add_chart(c, "J120")
            if chrt ==27 and  ByteTotal_Mode_32apsk_8_9.count("_")!=row1-1: 
                work_s2.add_chart(c, "S120")
            if chrt ==28 and ByteTotal_Mode_32apsk_9_10.count("_")!=row1-1: 
                work_s2.add_chart(c, "AB120")
            if chrt == 29 and bBFrameCount.count("_")!=row1-1: 
                work_s2.add_chart(c, "A140")
            if chrt ==30 and legsTotalCount.count("_")!=row1-1: 
                work_s2.add_chart(c, "J140")
            if chrt ==31 and legsPacketCount.count("_")!=row1-1: 
                work_s2.add_chart(c, "S140")
            if chrt ==32 and fastFadeCount.count("_")!=row1-1: 
                work_s2.add_chart(c, "AB140")
            if chrt ==33 and modCodChangeCount.count("_")!=row1-1: 
                work_s2.add_chart(c, "A160")

        print("AFTER RESHAPING", new_elim_list_test)
        Symbol_string = ','.join(str(x) for x in SYMBOLS)
        work_s2["A1"] = "The Symbols which are being iterated through are = %s" %(Symbol_string)
        
#        elif row1 == 40:
#           work_s2.add_chart(c, "A2") 
                    
            
        print("The row is given by", row1)
        test_work = os.path.expanduser('~\\Downstream_Excel.xlsx')
        
        Number_counter+=1
    return work_s

def excel_sheet_updation(test_bytes, test_values, workbook_object):
    import openpyxl

    work_book_name = "c:\\Users\\akarthik\\Downloads\\Downstream_Excel.xlsx"
    workbook_object1 = openpyxl.load_workbook(work_book_name)

    workbook_object12 = workbook_object1.active
    work_sheet2 = workbook_object1.create_sheet('Sheet2')
    for j in zip(test_bytes, test_values):
        work_sheet2.append(j)
        test_worksheet2 = workbook_object1.save(work_book_name)
    
        return test_worksheet2
global counter

def func_contain_excel_generation(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, modcod, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols, direction="downstream"):
    work_book_name = "c:\\Users\\akarthik\\Downloads\\Upstream_Excel_File.xlsx"
    if os.path.exists(work_book_name):
        print "The workbook is given by"
        print("The dictionary list combined is given by", dictionary_list_combined)
        
        if direction == "upstream":
            print(" I AM IN EXCEL SHEET FOR UPSTREAM>>>>")
            

            test_python_script_network.excel_sheet_generation(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, modcod, enumerate_list, work_book_name, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols) 

        else:
            excel_sheet_generation(dictionary_test, dictionary_list_combined, Number_counter, carrier, active_object_state, work_s2, gw, u_freq, modcod_list_length)

def func_contain_excel_generation_downstream(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, active_object_state, work_s2, SYMBOLS):
    work_book_name = os.path.expanduser('~\\Downstream_Excel.xlsx')
    if os.path.exists(work_book_name):
        print "The workbook is given by"
        print("The dictionary list combined is given by", dictionary_list_combined)
        
        excel_sheet_generation(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, active_object_state, work_s2, SYMBOLS)
         

def run_pp_commands(ip, username, password, tpa_port, work_b, dictionary_test, dictionary_list_combined, carrier, work_s2, active_object_state, gw, u_freq, modcod_list_length, SYMBOLS, flag=False):
    terminal_list = get_terminal_did_list(ip, username, password, tpa_port)

    cmd_output_array = []
    list_seperation_for_elements = []
    list_to_have_seperated_values = []
    list_bytes = []
    list_values = []
    list_bytes_test = []
    list_values_test = []

    dictionary_test_final  = {}

    list_bytes_test_new = []
     
    list_values_test_new = []

    dictionary_test["ByteTotal_Mode_qpsk_1_4"] = [] 
    dictionary_test["ByteTotal_Mode_qpsk_1_3"] = []
    dictionary_test["ByteTotal_Mode_qpsk_2_5"] = []
    dictionary_test["ByteTotal_Mode_qpsk_1_2"] = []
    dictionary_test["ByteTotal_Mode_qpsk_3_5"] = []
    dictionary_test["ByteTotal_Mode_qpsk_2_3"] = []
    dictionary_test["ByteTotal_Mode_qpsk_3_4"] = []
    dictionary_test["ByteTotal_Mode_qpsk_4_5"] = []
    dictionary_test["ByteTotal_Mode_qpsk_5_6"] = []
    dictionary_test["ByteTotal_Mode_qpsk_8_9"] = []
    dictionary_test["ByteTotal_Mode_qpsk_9_10"] = []
    dictionary_test["ByteTotal_Mode_8psk_3_5"] = []
    dictionary_test["ByteTotal_Mode_8psk_2_3"] = []
    dictionary_test["ByteTotal_Mode_8psk_3_4"] = []
    dictionary_test["ByteTotal_Mode_8psk_5_6"] = []
    dictionary_test["ByteTotal_Mode_8psk_8_9"] = []
    dictionary_test["ByteTotal_Mode_8psk_9_10"] = []
    dictionary_test["ByteTotal_Mode_16apsk_2_3"] = []
    dictionary_test["ByteTotal_Mode_16apsk_3_4"] = []
    dictionary_test["ByteTotal_Mode_16apsk_4_5"] = []
    dictionary_test["ByteTotal_Mode_16apsk_5_6"] = []
    dictionary_test["ByteTotal_Mode_16apsk_8_9"] = []
    dictionary_test["ByteTotal_Mode_16apsk_9_10"] = []
    dictionary_test ["ByteTotal_Mode_32apsk_3_4"] = []
    dictionary_test ["ByteTotal_Mode_32apsk_4_5"] = []
    dictionary_test["ByteTotal_Mode_32apsk_5_6"] = []
    dictionary_test["ByteTotal_Mode_32apsk_8_9"] = []
    dictionary_test["ByteTotal_Mode_32apsk_9_10"] = []
    dictionary_test["bBFrameCount"] = []
    dictionary_test["legsTotalCount"] = []
    dictionary_test["legsPacketCount"] = []
    dictionary_test["fastFadeCount"] = []
    dictionary_test["modCodChangeCount"] = []
    dictionary_test["invalidMessageCount"] = []

    Number_counter = 1     
    for term in terminal_list:
        print("<<<INSIDE TERMINAL LIST FOR LOOP>>")
        cmd_output = telnet_run_command(ip, username, password, tpa_port, "modcod", term)
        # print find_command_output_body(cmd_output, "modcod") + "\n"
        print cmd_output + "\n"
        # write_access_log(ip, "modcod", cmd_ouput)

        cmd_output = telnet_run_command(ip, username, password, tpa_port, "dvbs2rmt snr", term)
        print find_command_output_body(cmd_output, "dvbs2rmt snr") + "\n"
        # write_access_log(ip, "dvbs2rmt snr", cmd_ouput)

        cmd_output = telnet_run_command(ip, username, password, tpa_port, "dvbs2rmt stats", term)
        print "The command output for the dvbs2rmt stats \n", cmd_output
        print type(cmd_output)
        sys.stdout.flush()

        cmd_output_section = cmd_output[cmd_output.find("ByteTotal_Mode_qpsk_1_4"): cmd_output.rfind("admin@telnet") + 1]
        print "Command output section \n", cmd_output_section
        
        
        print("cmd after flush", cmd_output_section)

       
        cmd_output_array.append(cmd_output_section)
        print "The command output array is ", cmd_output_array

        if not cmd_output_array[-1]:
            cmd_output_array.pop(-1)

        #cmd_output_array = list(cmd_output_array)
        
        #print "list is given by", cmd_output_array

        cmd_output_array_Text1 = cmd_output_array[-1].split("\r\n")

        print "The cmd_output after split", cmd_output_array_Text1

        cmd_output_array_Text1.pop(-1)
        if not cmd_output_array_Text1[-1]:
            cmd_output_array_Text1.pop(-1)
        
        print "The array after pop", cmd_output_array_Text1
        
        for elements_sep in range(0, len(cmd_output_array_Text1), 1):
            cmd_output_array_Text1_after_split = cmd_output_array_Text1[elements_sep].split(" = ")
            print "output array", cmd_output_array_Text1_after_split
            list_to_have_seperated_values.append(cmd_output_array_Text1_after_split)
        
        print list_to_have_seperated_values
        
        for i in range(0, len(list_to_have_seperated_values)-1, 1):
            dictionary_test_final[list_to_have_seperated_values[i][0]] = list_to_have_seperated_values[i][1]
        
        print "The dictionary which contains the keys and values", dictionary_test

        for key in list(dictionary_test_final.keys()):
            if dictionary_test_final[key] == "0":
                #dictionary_test.pop(value)
                del dictionary_test_final[key]

        print "The dictionary after the key value has been popped", dictionary_test
        print " I am here after the dict values have been popped"
        #print "The list of bytes present", list_bytes
        #print "The list of values present", list_values
        """
        for i in range(0, len(list_values), 1):
            if list_values[i] == "0":
                list_values.remove(i)
        """     
        #for count in dictionary_test:
        #    list_bytes_test = list(dictionary_test.keys())
        #    list_values_test = list(dictionary_test.values())
    
        if "ByteTotal_Mode_qpsk_1_4" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_1_4"].append(dictionary_test_final["ByteTotal_Mode_qpsk_1_4"])
            print("The dictionary_test1", dictionary_test)
            #work_sheet_object_excel1 = func_contain_excel_generation(dictionary_test["ByteTotal_Mode_qpsk_1_4"])
        
        if "ByteTotal_Mode_qpsk_1_3" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_1_3"].append(dictionary_test_final["ByteTotal_Mode_qpsk_1_3"])
            print("The dictionary_test2", dictionary_test)
            #work_sheet_object_excel2 = func_contain_excel_generation(dictionary_test["ByteTotal_Mode_qpsk_1_3"])

        if "ByteTotal_Mode_qpsk_2_5" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_2_5"].append(dictionary_test_final["ByteTotal_Mode_qpsk_2_5"])
            print("The dictionary_test3", dictionary_test)
        
        if "ByteTotal_Mode_qpsk_1_2" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_1_2"].append(dictionary_test_final["ByteTotal_Mode_qpsk_1_2"])
            print("The dictionary_test4", dictionary_test)

        if "ByteTotal_Mode_qpsk_3_5" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_3_5"].append(dictionary_test_final["ByteTotal_Mode_qpsk_3_5"])
            print("The dictionary_test5", dictionary_test)
        
        if "ByteTotal_Mode_qpsk_2_3" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_2_3"].append(dictionary_test_final["ByteTotal_Mode_qpsk_2_3"])
            print("The dictionary_test6", dictionary_test)

        if "ByteTotal_Mode_qpsk_3_4" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_3_4"].append(dictionary_test_final["ByteTotal_Mode_qpsk_3_4"])
            print("The dictionary_test7", dictionary_test)

        if "ByteTotal_Mode_qpsk_4_5" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_4_5"].append(dictionary_test_final["ByteTotal_Mode_qpsk_4_5"])
            print("The dictionary_test8", dictionary_test)

        if "ByteTotal_Mode_qpsk_5_6" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_5_6"].append(dictionary_test_final["ByteTotal_Mode_qpsk_5_6"])

        if "ByteTotal_Mode_8psk_8_9" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_8psk_8_9"].append(dictionary_test_final["ByteTotal_Mode_8psk_8_9"])

        if "ByteTotal_Mode_qpsk_9_10" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_qpsk_9_10"].append(dictionary_test_final["ByteTotal_Mode_qpsk_9_10"])

        if "ByteTotal_Mode_16apsk_2_3" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_2_3"].append(dictionary_test_final["ByteTotal_Mode_16apsk_2_3"])

        if "ByteTotal_Mode_16apsk_3_4" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_3_4"].append(dictionary_test_final["ByteTotal_Mode_16apsk_3_4"])

        if "ByteTotal_Mode_16apsk_4_5" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_4_5"].append(dictionary_test_final["ByteTotal_Mode_16apsk_4_5"])

        if "ByteTotal_Mode_16apsk_5_6" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_5_6"].append(dictionary_test_final["ByteTotal_Mode_16apsk_5_6"])

        if "ByteTotal_Mode_16apsk_8_9" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_8_9"].append(dictionary_test_final["ByteTotal_Mode_16apsk_8_9"])

        if "ByteTotal_Mode_16apsk_9_10" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_16apsk_9_10"].append(dictionary_test_final["ByteTotal_Mode_16apsk_9_10"])

        if "ByteTotal_Mode_32apsk_3_4" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_32apsk_3_4"].append(dictionary_test_final["ByteTotal_Mode_32apsk_3_4"])

        if "ByteTotal_Mode_32apsk_4_5" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_32apsk_4_5"].append(dictionary_test_final["ByteTotal_Mode_32apsk_4_5"])

        if "ByteTotal_Mode_32apsk_5_6" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_32apsk_5_6"].append(dictionary_test_final["ByteTotal_Mode_32apsk_5_6"])

        if "ByteTotal_Mode_32apsk_8_9" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_32apsk_8_9"].append(dictionary_test_final["ByteTotal_Mode_32apsk_8_9"])

        if "ByteTotal_Mode_32apsk_9_10" in dictionary_test_final.keys():
            dictionary_test["ByteTotal_Mode_32apsk_9_10"].append(dictionary_test_final["ByteTotal_Mode_32apsk_9_10"])

        if "bBFrameCount" in dictionary_test_final.keys():
            dictionary_test["bBFrameCount"].append(dictionary_test_final["bBFrameCount"])

        if "legsTotalCount" in dictionary_test_final.keys():
            dictionary_test["legsTotalCount"].append(dictionary_test_final["legsTotalCount"])

        if "legsPacketCount" in dictionary_test_final.keys():
            dictionary_test["legsPacketCount"].append(dictionary_test_final["legsPacketCount"])

        if "fastFadeCount" in dictionary_test_final.keys():
            dictionary_test["fastFadeCount"].append(dictionary_test_final["fastFadeCount"])

        if "modCodChangeCount" in dictionary_test_final.keys():
            dictionary_test["modCodChangeCount"].append(dictionary_test_final["modCodChangeCount"])

        if "invalidMessageCount" in dictionary_test_final.keys():
            dictionary_test["invalidMessageCount"].append(dictionary_test_final["invalidMessageCount"]) 
    

        list_bytes_test_new = dictionary_test.keys()
        list_values_test_new = dictionary_test.values()
            
        
            #list_values_test.append(list_values)

        print dictionary_test.keys()
        print dictionary_test.values() 
        print "Bytes list sent", list_bytes_test_new
        print "Bytes values sent", list_values_test_new
        print "The keys in the list are given by", list_bytes
        print "The list After the 0 values are removed", list_values

        """
        work_book_name = "c:\\Users\\akarthik\\Downloads\\Downstream_Excel.xlsx"
        if os.path.exists(work_book_name):
            print "The workbook is given by"
            active_object_state = openpyxl.load_workbook(work_book_name)
           # for key,value in kwargs.iteritems():
            excel_sheet_generation(dictionary_test, active_object_state, work_book_name)
        """

        print("the dictionary is given as follows", dictionary_test)

        func_contain_excel_generation_downstream(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, active_object_state, work_s2, SYMBOLS) 

        


            #if " = " in cmd_output_array_Text1[elements_sep]:
            #    cmd_output_array_Text1[elements_sep].strip(" = ")
            #    list_seperation_for_elements.append(cmd_output_array_Text1[elements_sep])

        #print "The list after striping the required pattern", list_seperation_for_elements 



        """ 
        for count in range(cmd_output_array_Text1):
            if "\n" in cmd_output_array_Text1[count]:
                cmd_output_array_Text1[count].strip("\n")
        """
        print cmd_output_array_Text1

        print type(list_bytes)
        print type(list_values)

        #print ("The TEST_OBJ RETURN VALUE", test_obj)
        print find_command_output_body(cmd_output, "dvbs2rmt stats") + "\n"
        # write_access_log(ip, "dvbs2rmt stats", cmd_ouput)
    
    return terminal_list, dictionary_test



def reset_pp_stats(ip, username, password, tpa_port):
    terminal_list = get_terminal_did_list(ip, username, password, tpa_port)
    for term in terminal_list:
        telnet_run_command(ip, username, password, tpa_port, "dvbs2rmt stats reset", term)

def reset_terminal_stats(ip, username, password):
    telnet_run_command(ip, username, password, "", "rx griffin reset")

def get_terminal_stats(ip, username, password):  
    cmd_output = telnet_run_command(ip, username, password, "", "rx griffin stats")
    print cmd_output
    # print find_command_output_body(cmd_output, "rx griffin stats") + "\n"

"""Set of functions which execute the change the user frequency, gw frequency, symbol rate, maxmodcod"""
def exec_func_requests_maxmod(api_uri, headers, ip, username, password, maxmodcod_value, carrier_id):
    import requests 
    maxmodcod_change = requests.patch(api_uri, data=json.dumps({"maxmodcod": maxmodcod_value}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print "<<<<THE MAXMODCODCHANGE>>>", maxmodcod_change

#    print "Modcod Changed [IP: " + str(ip) + "| TerminalType: " + str(carrier_id) + "]: " + str(maxmodcod_change.json()['data']['obj_attributes']['maxmodcod'])

    return maxmodcod_change

def exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri):
    import requests 
    user_freq_change = requests.patch(api_uri, data=json.dumps({"fwdbandwidth": "54" , "fwduserfrequency": user_frequency}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
#    print "User Frequency Changed [IP: " + str(ip) + "| Channel: " + str(channel_id) + "]: " + str(user_freq_change.json()['data']['obj_attributes']['fwduserfrequency'])

    return user_freq_change

def exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri):
    import requests 
    gw_freq_change = requests.patch(api_uri, data=json.dumps({"fwdbandwidth": "54", "fwdgatewayfrequency": gw_frequency}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print "THE JSON OUTPUT", gw_freq_change.json()
#    print "GW Frequency Changed [IP: " + str(ip) + "| Channel: " + str(channel_id) + "]: " + str(gw_freq_change.json()['data']['obj_attributes']['fwdgatewayfrequency']) 

def exec_change_symbol_rate(sym_rate, headers, ip, carrier_id, username, password, api_uri):
    import requests 
    symrate_change = requests.patch(api_uri, data=json.dumps({"symbolrate": sym_rate}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
 #   print "SymbolRate Changed [IP: " + str(ip) + "| Carrier: " + str(carrier_id) + "]: " + str(symrate_change.json()['data']['obj_attributes']['symbolrate'])



def change_maxmodcod(ip, username, password, carrier_id, maxmodcod_value):
    API = "api/1.0/config/element/terminaltype"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=carrier_id)
    print "The contents of", api_uri

    exec_func_requests_maxmod(api_uri, headers, ip, username, password, maxmodcod_value, carrier_id)     
    
            

    # write_access_log(ip, "NMS : change modcod ", str(maxmodcod_change.content))

def change_user_frequency(ip, username, password, channel_id, user_frequency):
    API = "api/1.0/config/element/channel"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=channel_id)
   # fwdbandwidth = "54"
    try:
        exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri)     
    except:
        exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri)     

    # print user_freq_change.content
    # write_access_log(ip, "NMS : change modcod ", str(maxmodcod_change.content))

def change_gw_frequency(ip, username, password, channel_id, gw_frequency):
    API = "api/1.0/config/element/channel"
    API_test = ""
  #  headers = requests.
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=channel_id)
    print "THE API URI IS", api_uri
    try:
        exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri)
    except:
        exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri)

    
    # print gw_freq_change.content
    # write_access_log(ip, "NMS : change modcod ", str(maxmodcod_change.content))


def change_symbol_rate(ip, username, password, carrier_id, sym_rate):
    API = "api/1.0/config/element/downstreamcarrier"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=carrier_id)
   
    exec_change_symbol_rate(sym_rate, headers, ip, carrier_id,  username, password, api_uri)
    

    # print symrate_change.content
    # write_access_log(ip, "NMS : change modcod ", str(maxmodcod_change.content))

def apply_changes(ip, username, password, object_id):
    API = "api/1.0/config/element"
    print ("<<<<<THE OBJECT ID>>>", object_id)
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}/apply_changes'.format(api_endpoint=ip, nms_api=API, obj_id=object_id)
    try:
        apply_change = requests.post(api_uri, data=json.dumps({}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "The change for the json", apply_change
    except:
        #os.system("pip install requests")
        import requests
    finally:
        apply_change = requests.post(api_uri, data=json.dumps({}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "The change for the json", apply_change
    #except:
    #    import platform
    #    platform_apply_changes = platform.system()
    #    if (platform_apply_changes == "Linux"):
    #        apply_change_test = requests.post(api_uri, data=json.dumps({}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))

    # print "Config Status: " + str(apply_change.json()['code']) 
    # print apply_change.content

def default_maxmodcod(ip, username, password):
    default_termtype = {}
    API = "api/1.0/config/element/downstreamcarrier"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    print "API URI", api_uri
    try:
        termtype_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for dtype in termtype_info.json()['data']:
            if len(str(dtype['obj_attributes']['minmodcod'])) > 1:
                default_termtype[dtype['obj_id']] = dtype['obj_attributes']['minmodcod']
           
    except:
        #os.system("pip install requests")
        import requests
    finally:
        termtype_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "I AM IN TERMTYPE INFO..", termtype_info
        for dtype in termtype_info.json()['data']:
            if len(str(dtype['obj_attributes']['minmodcod'])) > 1:
                default_termtype[dtype['obj_id']] = dtype['obj_attributes']['minmodcod'] 
        
    return default_termtype
            


def get_terminal_list(ip, username, password):
    term_did_list = {}
    API = "api/1.0/config/element/satelliterouter"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    try:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_did_list[term['obj_attributes']['did']] = term['obj_id'] 
    except:
        import requests
    finally:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_did_list[term['obj_attributes']['did']] = term['obj_id']
        print ("The dictionary for the did list", term_did_list) 

    return term_did_list

def get_terminal_info(ip, username, password):
    term_ip_list = {}
    term_termtype_list = []
    API = "api/1.0/config/element/terminal"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    
    try:     
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_ip_list[term['obj_attributes']['coremodule_id']] = term['obj_attributes']['mgmtipaddress']
            term_termtype_list.append(term['obj_attributes']['terminaltype_id'])
    #        break
    except:
        #os.system("pip install requests, json")
        import requests, json
    #    brea
    finally:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "I AM IN TERMINAL INFO...", terminal_info
        for term in terminal_info.json()['data']:
            term_ip_list[term['obj_attributes']['coremodule_id']] = term['obj_attributes']['mgmtipaddress']
            term_termtype_list.append(term['obj_attributes']['terminaltype_id'])

    return term_ip_list, term_termtype_list

def get_channel_info(ip, username, password):

    try:
        channel_list = []
        channel_revert_list = []
        API = "api/1.0/config/element/channel"
        headers =  {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
        channel_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for channel in channel_info.json()['data']:
            if channel['obj_attributes']['state'] == "Up":
                channel_list.append(channel['obj_id'])
                r_channel = {}
                r_channel["obj_id"] = channel['obj_id']
                r_channel["fwduserfrequency"] = channel['obj_attributes']['fwduserfrequency']
                r_channel["fwdgatewayfrequency"] = channel['obj_attributes']['fwdgatewayfrequency']
                r_channel["fwdbandwidth"] = channel['obj_attributes']['fwdbandwidth']
                channel_revert_list.append(r_channel)
    except:
        #os.system("pip install requests")
        import requests, json
    finally:
#        channel_list = get_max_bw_channel_and_beam_id(ip, username, password, "api/1.0/config/element/channel", "downstream")[2]
        API = "api/1.0/config/element/channel"
        headers =  {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
        channel_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        #print get_info_request_channel_info 
    
        get_info_request_in_json = channel_info.json()['data']
    
        #get_info_request_in_json = json.dumps(get_info_request_in_json, indent=4, sort_keys=True)
        #print get_info_request_in_json

        rtn_bw_list = []
        channel_list = []
        channel_revert_list = []

        dictionary_for_channel_bw_pair = {}
    
        for bw in get_info_request_in_json:
            fw_bw_value = bw["obj_attributes"]["fwdbandwidth"]
            fw_bw_obj_id = bw["obj_id"]
            
            print "The forward Bandwidth is", fw_bw_value
            channel_list.append(fw_bw_value)

        channel_value = max(channel_list) 

        for channel_orig in get_info_request_in_json:
            if channel_orig["obj_attributes"]["fwdbandwidth"] == channel_value:
                fw_gw_freq = bw['obj_attributes']['fwdgatewayfrequency']
                fw_user_freq = bw['obj_attributes']['fwduserfrequency']        
                fw_gw_user_tuple_list = zip(fw_gw_freq, fw_user_freq) 
        """
        channel_list = []
        channel_revert_list = []
        API = "api/1.0/config/element/channel"
        headers =  {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
        channel_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for channel in channel_info.json()['data']:
            if channel['obj_attributes']['state'] == "Up":
                channel_list.append(channel['obj_id'])
                r_channel = {}
                r_channel["obj_id"] = channel['obj_id']
                r_channel["fwduserfrequency"] = channel['obj_attributes']['fwduserfrequency']
                r_channel["fwdgatewayfrequency"] = channel['obj_attributes']['fwdgatewayfrequency']
                r_channel["fwdbandwidth"] = channel['obj_attributes']['fwdbandwidth']
                channel_revert_list.append(r_channel)
        print("THE CHANNEL LIST IS GIVEN BY>>>>>>", channel_list) 
        """
    return channel_value, fw_gw_user_tuple_list, fw_bw_obj_id, fw_gw_freq, fw_user_freq   

def get_downstream_info(ip, username, password):
    try:
        down_list = []
        down_revert_list = {}
        API = "api/1.0/config/element/downstreamcarrier"
        headers =  {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
        channel_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for channel in channel_info.json()['data']:
            down_list.append(channel['obj_id'])
            down_revert_list[channel['obj_id']] = channel['obj_attributes']['symbolrate']
    except:
        #os.system("pip install requests")
        import requests, json
    finally:
        down_list = []
        down_revert_list = {}
        API = "api/1.0/config/element/downstreamcarrier"
        headers =  {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
        channel_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for channel in channel_info.json()['data']:
            down_list.append(channel['obj_id'])
            down_revert_list[channel['obj_id']] = channel['obj_attributes']['symbolrate']

    return down_list, down_revert_list

def process_cluster_output(cmd_ouput, command):
    cmd_ouput_body = find_command_output_body(cmd_ouput, command)
    flag = 0
    json_out = []
    for line in cmd_ouput_body.split("\n"):
        if "From cluster" in line:
            flag = 1
            continue

        if "ConfigConfirmed" in line:
            flag = 0

        if flag == 1:
            json_out.append(line.replace('\r', ''))

    return ''.join(json_out)

def get_pp_info(ip, username, password):
    cmd_ouput = telnet_run_command(ip, username, password, "13256", "cluster status")
    
    write_access_log(ip, "cluster status", cmd_ouput)
    
    test = process_cluster_output(cmd_ouput, "cluster status")
    print ("THE TEST", test)
    json_cmd = json.loads(test)

    ip_list = []
    for item in json_cmd["NODES"]:
        ip_list.append(str(item["ip"].split(";")[1]))
    
    print "The IP LIST is : ", ip_list
    
    return ip_list

def func_include_filename(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list):
    # Read and assign config variables
    """
    data = json.load(open(filename))
    NMS_USER= data['nms_username']
    NMS_PASSWORD= data['nms_password']
    PP_USER= data['pp_username']
    PP_PASSWORD= data['pp_password']
    PP_TPA = "10007"
    NMS_NODEIP = data['nms_ip']
    PP_NODEIP = data['pp_ip']
    DURATION = data['duration']
    SYMBOLS = data['symbols']
    """
    NMS_USER=nms_user 
    NMS_PASSWORD=nms_pass 
    PP_USER=pp_user
    PP_PASSWORD= pp_pass
    PP_TPA = "10005"
    NMS_NODEIP =nms_ip 
    PP_NODEIP = pp_ip  
    SYMBOLS = symbol_list 
    return (NMS_USER, NMS_PASSWORD, PP_USER, PP_PASSWORD, PP_TPA, NMS_NODEIP, PP_NODEIP, SYMBOLS)

def test_sat0_reach(ip, user, password, modcod_list_length, work_s2, active_object_state, work_b):
    
    
    
    test_ssh_op = test_python_script_network.Ssh_Console(ip, user, password)

    print "cmd_output", test_ssh_op
    send_string = test_ssh_op.send_string_and_wait("ssh -o StrictHostKeyChecking=no root@172.17.224.152", exit_on_error= False)
    print "The send_string", send_string

    send_string_ssh = test_ssh_op.send_string_and_wait("iDirect\n", exit_on_error= False)
    print ("The output", send_string_ssh)

    telnet_run_command = test_ssh_op.telnet_run_command("", "rx griffin stats")
    telnet_command_output_reset = test_ssh_op.send_the_command_in_telnet("rx griffin stats reset")
    telnet_command_output = test_ssh_op.send_the_command_in_telnet("rx griffin stats")
    print ("THE RX GRIFFIN STATS", telnet_command_output)
    
    telnet_command_output = telnet_command_output[telnet_command_output.find("Receiver ID 1"): telnet_command_output.rfind("[GRIFFIN] ") + 1]
    print("THE GRIFFIN STATS", telnet_command_output)

    telnet_command_output_split = telnet_command_output.split("\r\n")
    del telnet_command_output_split[0]
    del telnet_command_output_split[0]
    
    telnet_append_list = []
    
    for element in telnet_command_output_split:
        item_element = element[18:]
        telnet_append_list.append(item_element)
    print "The split list", telnet_command_output_split
    print "The split list append", telnet_append_list

    telnet_append_list_extend = []
    for items in telnet_append_list:
        telnet_append_list_extend.append(items.split())

    print("THE LIST ON SPLIT IS>>>", telnet_append_list_extend)

    del telnet_append_list_extend[-1]
    del telnet_append_list_extend[-1]

    modcod_list = ["MC"]
    crc8_list = ["CRC8"]
    crc32_list = ["CRC32"]

    for element in range(1, len(telnet_append_list_extend)):
        modcod_list.append(telnet_append_list_extend[element][0])
        crc8_list.append(telnet_append_list_extend[element][4])
        crc32_list.append(telnet_append_list_extend[element][5])
    
    print "MODCOD", modcod_list
    print "CRC8", crc8_list
    print "CRC32", crc32_list
    
    modcod_list_length = 10 
    func_contain_excel_generation_downstream_crc(modcod_list, crc8_list, crc32_list,  modcod_list_length, active_object_state, work_s2, work_b)

    #telnet_append_list_str = "".join(telnet_append_list)
    
def func_contain_excel_generation_downstream_crc(modcod_list, crc8_list, crc32_list,  modcod_list_length, active_object_state, work_s2, work_b):
    work_book_name = os.path.expanduser("~\\Downstream_Excel.xlsx")
    if os.path.exists(work_book_name):
        print "The workbook is given by"
        excel_sheet_generation_crc(modcod_list, crc8_list, crc32_list, modcod_list_length, active_object_state, work_s2, work_b)

    #file_text = "C:\\Users\\akarthik\\Downloads\\griffin.txt"
    #f = open(file_text, "w")
    #f.write(telnet_command_output)
    #f.close() 
def excel_sheet_generation_crc(modcod_list, crc8_list, crc32_list, modcod_list_length, active_object_state, work_s2, work_b):
    #row = 1
    global row2
    try:
        work_b = openpyxl.Workbook()
        work_s = work_b.active
        #work_s1 = work_b.create_sheet()
        
#        if column
        for i in zip(value):
            #work_s['A%s' %(i)] = cmd_input_for_excel 
            work_s.append(i)
         #   work_s1.append(j)
            #work_s['Command Output'] = cmd_output
            test_work = work_b.save("xlwt c:\\Users\\akarthik\\Downloads\\Text1.xls")
    except: 
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.chart import (LineChart, ScatterChart, Reference)
        from openpyxl.chart.axis import DateAxis
        from openpyxl.chart.layout import Layout, ManualLayout
        import copy

    finally:
        #print("The value of row", row)
        modcod_list_length_crc = 10
        print("<<<<<THE MODCOD LIST LENGTH IS GIVEN>>>>", modcod_list_length_crc) 
        work_s = active_object_state.active
        
        
        #work_s2 = active_object_state.get_sheet_by_name("Chart")

        #work_s1 = work_b.create_sheet(
        """
        cell1 = work_s.cell(row=1, column=1)
       # wb = copy(rb) # a writable copy (I can't read values out of this, only write to it)
       # w_sheet = wb.get_sheet(0)
        #i = 0
        #j = 0 
        
        dictionary_list_values = [] 

        for i in range(0, len(dictionary_set.values())-1, 1):
            dictionary_values_set = ''.join(dictionary_set.values()[i])
            dictionary_list_values.append(dictionary_values_set)
        
        print (dictionary_list_values)
       # print("COUNTER", counter)
       """
        #items = [i for i in modcod_list]
        i = 0
        j = 0
        k = 1
        l = 0
        m = 1
        print("THE LENGTH OF THE MODCOD>>", len(modcod_list))
        for i in range(0, len(modcod_list)-1):
            print "MODCOD LIST", modcod_list[i]
            print "MODCOD LIST", type(modcod_list[i])
            i+=1
        row_counter= modcod_list_length_crc+1 
        cell_ref = work_s.cell(row = row_counter, column=2)
        cell_ref.value=crc8_list[0] 
        for row1 in range(1, len(modcod_list)+1):
            cell_reference = work_s.cell(row=row_counter, column=1)
            cell_reference.value = modcod_list[j]
            cell_reference12 = work_s.cell(row=row_counter, column=4)
            cell_reference12.value = modcod_list[l]
            
            j+=1
            l+=1
            row_counter+=1
        
        row_counter12= modcod_list_length_crc+1 
        row_counter_crc = modcod_list_length_crc+2
        
        cell_ref1 = work_s.cell(row=row_counter12, column=5)
        cell_ref1.value=crc32_list[0]
        
        for row2 in range(2, len(crc8_list)+1):
            cell_reference1 = work_s.cell(row=row_counter_crc, column=2)
            crc8_int = int(crc8_list[m]) 
            cell_reference1.value = crc8_int
            
            cell_reference21 = work_s.cell(row=row_counter_crc, column=5)
            crc32_int = int(crc32_list[k])
            cell_reference21.value = crc32_int
            row_counter_crc+=1
            m+=1
            k+=1
        row_counter_crc_data = modcod_list_length_crc+1
        for chart in modcod_list:
            c = LineChart()
            c.style = 13
            c.y_axis.title = 'Value in Bytes'
            c.x_axis.title = 'Samples Interval'
            c.x_axis.scaling.min = 0
            #c.y_axis.scaling.min = 0
            c.x_axis.scaling.max = len(crc8_list)
            c.x_axis.unit = 1
            #c.y_axis.scaling.max = 70000
            c1 = LineChart()
            c1.style = 13
            c1.y_axis.title = 'Value in Bytes'
            c1.x_axis.title = 'Samples Interval'
            c1.x_axis.scaling.min = 0
            #c.y_axis.scaling.min = 0
            c1.x_axis.scaling.max = len(crc32_list)
            c1.x_axis.unit = 1
            
            data8 = Reference(work_s, min_col=1, min_row=row_counter_crc_data, max_col=2, max_row=row_counter_crc_data)
            data32 = Reference(work_s, min_col=4, min_row=row_counter_crc_data, max_col=5, max_row=row_counter_crc_data)
            print ("THE DATA IS>>>>", data8)
            c.add_data(data8, titles_from_data=True)
            c1.add_data(data32, titles_from_data=True)

            if chart =="QPSK-1/4":
                c.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=2).value, chart)
                c1.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=5).value, chart)
                work_s2.add_chart(c, "A180")
                work_s2.add_chart(c1, "A196")
            if chart == "QPSK-1/3":
                c.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=2).value, chart)
                c1.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=5).value, chart)
                work_s2.add_chart(c, "K180")
                work_s2.add_chart(c1, "K196")
            if chart == "QPSK-2/5":
                c.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=2).value, chart)
                c1.title = "This is %s and modcod %s " %(work_s.cell(row=modcod_list_length_crc+1, column=5).value, chart)
                work_s2.add_chart(c, "A212")
                work_s2.add_chart(c1, "K212")
            if chart == "QPSK-1/2":
                c.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=2).value, chart)
                c1.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=5).value, chart)
                work_s2.add_chart(c, "A228")
                work_s2.add_chart(c1, "K228")
            if chart == "QPSK-3/5":
                c.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=2).value, chart)
                c1.title = "This is %s and modcod %s" %(work_s.cell(row=modcod_list_length_crc+1, column=5).value, chart)
                work_s2.add_chart(c, "A244")
                work_s2.add_chart(c1, "K244")
            
            row_counter_crc_data+=1
            
        active_object_state.save(os.path.expanduser("~\\Downstream_Excel.xlsx"))

def main(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list):
    filename = "c:\\Users\\akarthik\\Downloads\\New_folder\\config.json"
    NMS_USER, NMS_PASSWORD, PP_USER, PP_PASSWORD, PP_TPA, NMS_NODEIP, PP_NODEIP, SYMBOLS_1 = func_include_filename(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list)
    
    # Get downstream carrier info
    down_list, down_rev_list = get_downstream_info(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
    
    # Get Channel Info
    channel, c_rev_list, fw_bw_obj_id, fw_gw_freq, fw_user_freq = get_channel_info(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
    print "THE CREVLIST", c_rev_list
#    (gw_orig, user_orig) = c_rev_list

    try:
        ACCESS_LOG="/tmp/layer1_access.log"
        DEBUG=False
        #modcod_list = ["QPSK-1/4", "QPSK-1/3", "QPSK-2/5", "QPSK-1/2", "QPSK-3/5", \
        #               "QPSK-2/3", "QPSK-3/4", "QPSK-4/5", "QPSK-5/6", "QPSK-8/9", \
        #               "8PSK-3/5", "8PSK-2/3", "8PSK-3/4", "8PSK-5/6", "8PSK-8/9", \
        #              "16APSK-2/3", "16APSK-3/4", "16APSK-4/5", "16APSK-5/6", "16APSK-8/9"]
        modcod_list = ["QPSK-1/4", "QPSK-1/3", "QPSK-2/5","QPSK-1/2", "QPSK-3/5"]

        
        

        #data_test = data['symbols']   
        
        #test_symbols = "symbols" in data
        #print "test_symbols", test_symbols

        #elements_list = []
        #for elements in data["symbols"]:
        #    print elements
        #    elements_list.append(elements)

        #print "The list is as follows", elements_list
            
        #print "The symbol list is given by", elements

        # Get the list of nodes in pp cluster    
        pp_ip_list = get_pp_info(PP_NODEIP, PP_USER, PP_PASSWORD)
        # Get the present value maxmodcod for Terminaltype, to recover the state after test
        def_maxmodcod = default_maxmodcod(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
        # Get list of Terminal DIDs fromlen() NMS
        term_did_list = get_terminal_list(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
        # Get the managemnet IPs of terminals and the list of their terminaltype
        term_ip_list, term_termtype_list = get_terminal_info(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
        # Get Terminal DID list in Network from PP
        pp_did_list = []
        for ip in pp_ip_list:
            pp_did_list = pp_did_list + get_terminal_did_list(ip, PP_USER, PP_PASSWORD, PP_TPA)

        print ("The DID LIST", pp_did_list)
        

        if DEBUG:
            print "PP IP List:" + str(pp_ip_list)
            print "Terminaltype default maxmodcod List:\n" + str(def_maxmodcod)
            print "Terminal DID List:\n" + str(term_did_list)
            print "Terminal IP List:\n" + str(term_ip_list)
            print "Terminal Type List:\n" + str(term_termtype_list)
            print "Terminal DID (PP) List:\n" + str(pp_did_list)
            print "Channel List:\n" + str(channel)
            print "Downstream Carrier List:\n" + str(down_list)
        
        SYMBOLS = [1000, 5000, 10000, 20000, 30000, 45000]
        #SYMBOLS = [1000]
        #gw_freq_list = [1.0, 1.05, 1.10, 1.15, 1.20, 1.25, 1.30, 1.35, 1.40, 1.45, 1.50, 1.55, 1.60, 1.65, 1.70, 1.75, 1.80, 1.85, 1.90, 1.95]
        gw_freq_list = [1.0, 1.05]
        #user_freq_list = [19.3, 19.35, 19.40, 19.45, 19.50, 19.55, 19.60, 19.65, 19.70, 19.75, 19.80, 19.85, 19.90, 19.95, 20.0, 20.05, 20.10, 20.15, 20.20, 20.25]
        user_freq_list = [19.3, 19.35]

        gw_user_pair = zip(gw_freq_list, user_freq_list)

        print ("THE COMINATION PAIR OF GW AND USER", gw_user_pair)

        modcod_list_length = 10

        print("<<THE MODCOD LIST LENGTH>>", modcod_list_length)
        #SYMBOLS_from_file = 
        #test_symbols = SYMBOLS

        if len(SYMBOLS) == 6:
        # SYMBOLS = SYMBOLS_from_file
        

            print "I am in if condition"
            print "The symbol list is given by", SYMBOLS

        print "I am here"
        #counter += 1
        #print counter
        
        '''The workbook object created to open the workbook'''
        import openpyxl
        work_b = openpyxl.Workbook()
        work_book_name = os.path.expanduser('~\\Downstream_Excel.xlsx')
        
        work_b.save(work_book_name) 
            # print "the obj declaration", test_work
        
        #print "The work_book", type(work_b)   
        #excel_sheet_generation.counter+=1

        dictionary_test = OrderedDict()
        dictionary_list_values = []
        dictionary_list_combined = []
        #i1 = 1 
        work_sheet_rename = work_b.get_sheet_by_name("Sheet")
        work_sheet_rename.title = "Raw Data"
        for carrier in down_list:
            print("------------------------------------")
            print("<<<OBJ ID IS GIVEN BY>>>>", carrier)
            
            for gw, u_freq in gw_user_pair:
                active_object_state = openpyxl.load_workbook(work_book_name)
                work_s2 = active_object_state.create_sheet("Chart%s,%s"%(gw, u_freq))
                 
                print("<<<GW FREQUENCY IS>>", gw)
                change_gw_frequency(NMS_NODEIP, NMS_USER, NMS_PASSWORD, fw_bw_obj_id, str(gw))
                apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, fw_bw_obj_id)
                print("<<<<USER_FREQUENCY IS>>>>", u_freq)
                change_user_frequency(NMS_NODEIP, NMS_USER, NMS_PASSWORD, fw_bw_obj_id, str(u_freq))
                apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, fw_bw_obj_id)
                for symbol in SYMBOLS:
                    print("<<<<<<<<SYMBOL LIST HIT>>>>>>>>>", symbol)
                    # for channel in c_list:
                    #     print("<<<<<<<<<<<THE CARRIER FREQUNCY>>>>>>")
                    #     print("-------------------------------------")
                    #     print("THE CARRIER FREQUENCY IS GIVEN BY.....", channel)
                    change_symbol_rate(NMS_NODEIP, NMS_USER, NMS_PASSWORD, str(carrier), symbol)
                    apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, str(carrier))
                    
                    # Change maxmodcod
                    for modcod in modcod_list:
                        print("<<<MOD COD >>>", modcod)
                        for term_type in set(term_termtype_list):
                            print "<<<TERMTYPE>>>", term_type
                            print "<<<TERMTYPE LIST>>>", term_termtype_list
                            change_maxmodcod(NMS_NODEIP, NMS_USER, NMS_PASSWORD, term_type, modcod)
                            apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, term_type)
                        
                        print "Sleep, to let the configuration propogate......"
                        time.sleep(60)

                        #Reset PP stats
                        for ip in pp_ip_list:
                            reset_pp_stats(ip, PP_USER, PP_PASSWORD, PP_TPA)
                    #Ping terminals
                        for did in pp_did_list:
                            terminal_list = term_did_list[did]
                            print("The terminal list is GIVEN BY>>>", terminal_list)
                            terminal_ip = term_ip_list[str(term_did_list[did])]
                            reset_terminal_stats(terminal_ip, "root", "iDirect")
                            if (platform.system == 'Windows'):
                                cmd_ping = ['ping', '-n', '10', terminal_ip, '>', '/dev/null', '&']
                                subprocess.call(cmd_ping)
                            elif(platform.system == 'Linux'):
                                cmd_ping = ['ping', '-c', '10', terminal_ip, '>', '/dev/null', '&']
                                subprocess.call(cmd_ping) 
                        # Run commands on PP
                        for ip in pp_ip_list:
                            print("EXCEL SHEET HIT....")
                            print("Before running the PP Commands")
                            run_pp_commands(ip, PP_USER, PP_PASSWORD, PP_TPA, work_b, dictionary_test, dictionary_list_combined, carrier, work_s2, active_object_state, gw, u_freq, modcod_list_length, SYMBOLS, flag=True)

                    # Run commands on carrierRemote
                        for did in pp_did_list:
                            test_sat0_reach(NMS_NODEIP, "idirect", "iDirect", modcod_list_length, work_s2, active_object_state, work_b)
                            print("In the for loop for the pp did check")
                            #get_terminal_stats(term_ip_list[str(term_did_list[did])], "root", "iDirect")
                            
                            if DEBUG:
                                print "DID: " + str(did)
                                print "Terminal ID: " + str(term_did_list[did])
                                print "Terminal IP: " + str(term_ip_list[str(term_did_list[did])])

    except KeyboardInterrupt:
        import requests
        print "Change modcod to previous values:\n"
        filename = "C:\\Users\\akarthik\\Downloads\\New_folder\\config.json"
        NMS_USER, NMS_PASSWORD, PP_USER, PP_PASSWORD, PP_TPA, NMS_NODEIP, PP_NODEIP, SYMBOLS = func_include_filename(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list)
        def_maxmodcod = default_maxmodcod(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
        print ("THE DEFAULT MODCOD>>>", def_maxmodcod) 
        for term_type in def_maxmodcod:
            for key,value in def_maxmodcod.items():
                maxmodcod_value = def_maxmodcod.values()
                maxmodcod_value = "".join(maxmodcod_value)
                maxmodcod_value_str = maxmodcod_value.encode("utf-8") 
            change_maxmodcod(NMS_NODEIP, NMS_USER, NMS_PASSWORD, term_type, maxmodcod_value_str)
            apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, term_type)
        
        #down_list, down_rev_list = get_downstream_info(NMS_NODEIP, NMS_USER, NMS_PASSWORD)

        for carrier, symbol in down_rev_list.iteritems():
            change_symbol_rate(NMS_NODEIP, NMS_USER, NMS_PASSWORD, str(carrier), str(symbol))
            apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD, str(carrier))

        #channel, c_rev_list, fw_bw_obj_id, fw_gw_freq, fw_user_freq = get_channel_info(NMS_NODEIP, NMS_USER, NMS_PASSWORD)
        for json_c in channel:
            change_gw_frequency(NMS_NODEIP, NMS_USER, NMS_PASSWORD,  fw_bw_obj_id, fw_gw_freq)
            change_user_frequency(NMS_NODEIP, NMS_USER, NMS_PASSWORD,  fw_bw_obj_id, fw_user_freq)
            apply_changes(NMS_NODEIP, NMS_USER, NMS_PASSWORD,  fw_bw_obj_id)   
        print("PLEASE CLOSE THE CONSOLE TO RETURN TO THE CMD PROMPT....")
    
    # print "\nACCESS LOGS: /tmp/layer1_access.log\n"

if __name__ == '__main__':
    #filename = sys.argv[1]
    #print ("The filename is", filename)
    main(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list)