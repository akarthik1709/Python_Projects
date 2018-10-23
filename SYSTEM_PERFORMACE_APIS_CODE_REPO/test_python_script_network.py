import json
from requests.auth import HTTPBasicAuth
import requests
import sys
import argparse
import os
#import downstream
#path = "C:\\Users\\akarthik\\Downloads\\pp_tools\\toli\\ssh\\__init__.py"
#print (path)
#sys.path.append(os.path.join(path))
import ssh
try:
    import np
except:
    import platform
    #if platform.system() == "Windows":
        #os.system("pip install np")
try:
    import paramiko
except:
    import platform
    #if platform.system() == "Windows":
        #os.system("pip install paramiko")

from random import randint
import itertools
import time
import re
import downstream_modified_gw_user_pair
import wxpython
from collections import OrderedDict
import socket
import datetime



# Class which is used for post query to the NMS
class Post_Request_Reverse_Carrier:
    def __init__(self, ip, username, password,obj_id):
        """
        :param: ip: IP Address of the NMS
        :param: api_test_query: API Query given to the NMS
        :param: obj_id: The Object ID which is got from the get response
        :param: obj_parentid: The Object parent ID which is got as a response from the NMS
        """
        self.obj_id = obj_id
        self.ip = ip
        self.username = username
        self.password = password

    
    def request_for_upstream(self, body, api_test_query):
        """
        :param: body: The JSON body which contains the post request done to the NMS
        """
        headers = {"Content-Type":"application/json;charset=utf-8"} 
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api = api_test_query)
        print ("upstream ip", api_uri)
        
        post_info_request_carrier = requests.post(api_uri, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        print (post_info_request_carrier)
        post_request_in_json = post_info_request_carrier.json()['data']
        
        print (post_request_in_json)

        test_json = json.dumps(post_request_in_json, indent=4, sort_keys=True)
        print (test_json)

        obj_id_upstream_carrier = post_request_in_json['obj_id']
        print(type(obj_id_upstream_carrier))

        return  (post_request_in_json, obj_id_upstream_carrier)

    
    def request_for_composition(self, body, api_test_query):
        """
        :param: body: The JSON body which contains the post request done to the NMS
        """
        headers = {"Content-Type":"application/json;charset=utf-8"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api = api_test_query)
        print ("composition ip", api_uri)
        
        post_info_request_composition = requests.post(api_uri, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        print ("I am here", post_info_request_composition)

        post_info_request_comp = post_info_request_composition.json()['data']
        post_info_request_comp_id = post_info_request_comp["obj_id"]
        print ("The object ID", post_info_request_comp_id)

        test_json_print = json.dumps(post_info_request_comp, indent=4, sort_keys=True)
        print (test_json_print)

        return post_info_request_comp_id

class Delete_Request_Reverse_Carrier:
    def __init__(self, ip, username, password,obj_id, obj_parentid):
        self.ip = ip
        self.username=username
        self.password=password
        self.obj_id=obj_id
        self.obj_parent_id=obj_parentid

    
    def delete_inroutegroup_composition(self, api_test_query):
        """
        :param: The parameters which is used to check the key values for the deletion criteria
        """
        headers = {"Content-Type":"application/json;charset=utf-8"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api = api_test_query)
        print ("composition ip", api_uri)
        post_info_request_composition = requests.delete(api_uri, auth=HTTPBasicAuth(username, password), headers=headers)

        return post_info_request_composition

    
    def delete_upstreamcarrier(self, api_test_query):
        """
        :param: api_test_query:The parameters which are used to check the  
        """
        headers = {"Content-Type":"application/json;charset=utf-8"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api = api_test_query)
        print ("composition ip"), api_uri
        post_info_request_composition = requests.delete(api_uri, auth=HTTPBasicAuth(username, password), headers=headers)

        return post_info_request_composition

def get_network_name_and_id(ip, username, password):
        api_test_query = "api/1.0/config/element/network"
        headers = {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_query)
        #post_body_required = """ """
        channel_info = requests.get(api_uri,  headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print("channel Info", channel_info)
        json_data = channel_info.json()
        #json_data = str(json_data)

        
        test_json = json.dumps(json_data, indent=4, sort_keys=True)
        print (test_json)

        network_name = []
        network_obj_id = [] 
        for test in json_data['data']:
            #print(test['data']['obj_attributes']['service_mode'])
#            print (test['obj_attributes']['gscenabled'])
            network_name.append(test['obj_name'])
            network_obj_id.append(test['obj_id'])
            #test_attributes['obj_id']
        
        dict_network_name = {}
        dict_network_id = {} 
        i=1
        for elements in network_name:
            dict_network_name["Network%s"%(i)] = elements
            i+=1
        print ("The dictionary is given by", dict_network_name)

        j=1
        for elements_id in network_obj_id:
           dict_network_id["Network%s"%(j)] = elements_id 
           j+=1

        print ("The dictionary is given by", dict_network_id)

        print ("The network_name list", network_name)
        print ("The network obj id", network_obj_id)

        #print(test_json["data"]["obj_name"])
        #file_1 = open("c:\\Users\\akarthik\\Downloads\\file.txt", "w")
        #file_1.write(test_json)
        #print("The json data", json_data)
        #for test in channel_info.json()['data']:
        #    print (test['obj_name'])

def get_inroutegroup_profile_details(ip, username, password, igp_name):
        api_test_query = "api/1.0/config/element/inroutegroupprofile/obj_name=%s"%(igp_name)
        headers = {"Content-Type":"application/json"}
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_query)
        print("API_URI", api_uri)
        #post_body_required = """ """
        channel_info = requests.get(api_uri,  headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print("channel Info", channel_info)
        json_data = channel_info.json()
        json_data_1 = channel_info.json()['data']
        json_data_str = str(json_data)

        json_data = json_data_1["obj_attributes"]["skipinterval"]
        obj_id = json_data_1["obj_id"]
        obj_parentid = json_data_1["obj_parentid"]

        print ("test", json_data)
        print ("obj_id", obj_id)
        print ("obj_parent_id", obj_parentid)

    
        #test_json = json.dumps(json_data, indent=4, sort_keys=True)
        #print (test_json)

        #print ("1----", json_data_1)

        #json_data_1 = str(json_data_1)

        #test_json1 = json.loads(json_data_1, indent=4, sort_keys=True)
        #print (test_json1)

        #dict_1 = {}
        #for elements in test_json1:
            #print elements
        #    dict_1["child"] = elements["obj_attributes"]["framelength"]
        #print (dict_1) 
        #print ("a", test_json1[3])


    #    igp_section_json =  test_json[test_json.find("data"): test_json.rfind("}") + 1]
    #    print ("The json section of igp", igp_section_json)

   #     igp_section_json = igp_section_json.splitlines()
   
        """
        dictionary_beam_list = {}
        for test in json_data['data']:
            #print ("The parent object id", test['obj_parentid'])
            print (test['obj_child_num'])
            print ("The object id", test['obj_id'])
            dictionary_beam_list['Beam List'] = test['obj_attributes']['framelength']  
        
        print (dictionary_beam_list)
        print (json_data['obj_id'])
        print (json_data['obj_parent_id'])
        """
        return (obj_id, obj_parentid)

class Ssh_Console:

    def __init__(self, ipaddress, username, password, root_password = "iDirect", timeout = None):
        """
        Initialize the console
        :param ipaddress: The IP Address of the host
        :param username: The username of the host
        :param password: The password of the host
        :param root_password: The root password of the host
        :return: None
        """

        self.ssh_sftp_client_list = []        
        self.__receive_buffer_size  = 65536
  
        self.ssh_client = paramiko.SSHClient()
        paramiko.common.logging.basicConfig(level=paramiko.common.DEBUG)
        self.ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        socket.getaddrinfo('localhost', '8080')
        self.ssh_client.connect(ipaddress, 22, username, password, timeout = 30)

        self.root_user = self.ssh_client.get_transport().open_session()
        self.root_user.get_pty(term='vt100', width=80, height=20000)
        self.root_user.invoke_shell()
        self.root_user.settimeout(timeout)
        while not self.root_user.send_ready():
            continue
 
        print("Before changing to root user.")
 
        self.root_user.send("su - \n")
        print("Next wait for password request")
 
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
            print ("The output received after the command: su - : ", data_decoded)
            if "Password:" in data_decoded:
                break

        print("We are done executing the while loop which expects the password prompt")
 
        if self.root_user.recv_stderr_ready():
           data_std_err = self.root_user.recv_stderr(self.__std_data_recv_prompts)
           print ("The error received after the command: su - : ", data_std_err)
 
        self.root_user.send(root_password + "\n")
        
        print("We just sent the root password and next we are waiting for the root console prompt")
       
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
                print ("The output received after the the password was sent: ", data_decoded)
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded:
                break
 
        if self.root_user.recv_stderr_ready():
            data_std_err = self.root_user.recv_stderr(self.__std_data_recv_prompts)
            print ("The error received after the password was sent: ", data_std_err)

    
    def send_string_and_wait(self, command, exit_on_error=True):
        """Send the command onto the console
        :param command: Command parameter to input the command
        :return: console output
        """

        rand_str =  str(randint(1, 1234))
        self.std_error_string = "std_error_session_" + rand_str  + ".log"

        print "THE STDERROR STRING>>>",self.std_error_string 
       
        ''' empty the file ''' 
        the_command = "truncate -s 0  " +  self.std_error_string  + " \n" 
        self.root_user.send(the_command)
        ### Next we get stdout string ###
        stdout_str = ""
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
         
                print("FETCHING DATA after truncating error log file -------")
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
                stdout_str += data_decoded
                print ("data read decoded : ", data_decoded)
                print ("std_out string: : '%r'" % (stdout_str))
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded or "#" in data_decoded or "password:" in data_decoded:
                break


        ''' send the command '''
        the_command = command + "  2>  " + self.std_error_string  + " \n"
        self.root_user.send(the_command)
        print ("after sending the command: ", command)
        
        ### Next we get stdout string ###
        stdout_str = ""
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                print("FETCHING DATA after sending the actual command -------")
                data = self.root_user.recv(self.__receive_buffer_size)
                #data_decoded = bytes.decode(data)
                
                data_decoded = data
                stdout_str += data_decoded
                print ("data read decoded : ", data_decoded)
                print ("std_out string: : '%r'" % (stdout_str))
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded or "#" in data_decoded or "password:" in data_decoded:
                break

        print("<<<< done with log")

        
        ## Next we get the std error ##
        the_command = "cat  " + self.std_error_string  + " \n"
        self.root_user.send(the_command)
        stderr_str = ""
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                print("FETCHING DATA cat (ing) the error log file -------")
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
                stderr_str += data_decoded
                print ("error data read decoded : ", data_decoded)
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded or "#" in data_decoded or "password:" in data_decoded:
                break
  
        ## The next processing done for the standard error string ## 
        print("The stderr string:\n" + stderr_str + "\n<<<end of error string")
        stderr_str = stderr_str.splitlines()
        print ("The stderr after splitlines is: ", stderr_str)

        for i in range(0, len(stderr_str)):
            if "std_error_session" in stderr_str[i]:
                stderr_str.pop(i)
                break

        for i in range(0, len(stderr_str)):
            if "You have new mail in /var/spool/mail/root" in stderr_str[i]:
                stderr_str.pop(i)
                break

        for i in range(0, len(stderr_str)):
            if "#" in stderr_str[i]:
                stderr_str.pop(i)
                break
 
        
        ## print the std error string## 
        print("The stderr string after removing not error related strings:\n")
        print(stderr_str)
        print("\n<<<end of error string")
        print("The stderr string length is: ", len(stderr_str))

        ''' remmove the error file '''
        the_command = "rm -f  " +  self.std_error_string  + " \n" 
        self.root_user.send(the_command)
        

        ### Next we get stdout string ###
        stdout_str2 = ""
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                print("FETCHING DATA removing the error log file -------")
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
                stdout_str2 += data_decoded
#                print ("data read decoded : ", data_decoded)
#                print ("std_out string2: : '%r'" % (stdout_str2))
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded or "#" in data_decoded or "password:" in data_decoded:
                break


        ''' handle and print the standard error'''
        if len(stderr_str) > 0 :
            print("\n\n ******* ERROR ********")
            print("\n\nThe Error String:\n", stderr_str)
            print("\n\n ******* ERROR ********")
            sys.stdout.flush()
            if exit_on_error == True:
                raise RuntimeError(stderr_str)
            else:
                print("there is an error, but we continue, because exit_on_error is set to False")


        sys.stdout.flush()
        return (stdout_str, stderr_str)

    def send_the_command_in_telnet(self, command):
        the_command = command + "\n"
        self.root_user.send(the_command)
        print ("after sending the command: ", command)
        
        ### Next we get stdout string ###
        stdout_str = ""
        while not self.root_user.exit_status_ready():
            data_decoded = ''
            while self.root_user.recv_ready():
                print("FETCHING DATA after sending the actual command -------")
                data = self.root_user.recv(self.__receive_buffer_size)
                data_decoded = data
                stdout_str += data_decoded
#                print ("data read decoded : ", data_decoded)
#                print ("std_out string: : '%r'" % (stdout_str))
            if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded:
                break
                
        return stdout_str

    def get_sftp_client(ipaddress, username, password):
        """
        This function is used to get the SFTP client object
        :param: self
        :type: self.sftp_client_web
        :return: self.sftp_client_web
        """
        ssh_ftp_client = paramiko.SSHClient()
        paramiko.common.logging.basicConfig(level=paramiko.common.DEBUG)
        ssh_ftp_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh_ftp_client.connect(ipaddress, 22 , username, password, timeout = 30)

        sftp_client = ssh_ftp_client.open_sftp()
        sftp_client.sshclient = ssh_ftp_client

        ''' maybe this ensure that these are not closed while the FTP client is being used'''
#       self.ssh_sftp_client_list.append(ssh_ftp_client)
        
        return sftp_client
    
    def get_ipaddress_of_host(self):
        
        """
        This function is used to get the ip address of the remote host
        :return: self.ipaddress_obj
        """
        self.ipaddress_obj = self.root_user.get_transport().sock.getpeername()[0]
        print ("The IP Address of the remote host is: ", self.ipaddress_obj)
        
        return self.ipaddress_obj   

    
    def telnet_run_command(self, port, command, remote=None):
        before_xoff = []
        test_b = ""
        
        __receive_buffer_size = 65536
        if ssh is not None:
            cmd_telnet = "telnet 0 " + port
            print "THE CMD TELNET IS>>>", cmd_telnet
            #raw_input("ENTER CONTINUE IN TELNET RUN COMMAND.....")
            self.root_user.send(cmd_telnet + "\n")
            while not self.root_user.exit_status_ready():
                data_decoded = ''
                while self.root_user.recv_ready():
                    data = self.root_user.recv(self.__receive_buffer_size)
                    data_decoded = data
                print ("The output received after the command: su - : ", data_decoded)
                if "Username:" in data_decoded:
                    break

            print("We are done executing the while loop which expects the password prompt")
    
            if self.root_user.recv_stderr_ready():
                data_std_err = self.root_user.recv_stderr(self.__std_data_recv_prompts)
            
                print ("The error received after the command: su - : ", data_std_err)
    
            self.root_user.send("admin" + "\n")
            
            print("We just sent the root password and next we are waiting for the root console prompt")
        

            while not self.root_user.exit_status_ready():
                data_decoded = ''
                while self.root_user.recv_ready():
                    data = self.root_user.recv(self.__receive_buffer_size)
                    data_decoded = data
                print ("The output received after the command: su - : ", data_decoded)
                if "Password:" in data_decoded:
                    break

            print("We are done executing the while loop which expects the password prompt")
    
            if self.root_user.recv_stderr_ready():
                data_std_err = self.root_user.recv_stderr(self.__std_data_recv_prompts)
            
                print ("The error received after the command: su - : ", data_std_err)
    
            self.root_user.send("iDirect" + "\n")
            
            print("We just sent the root password and next we are waiting for the root console prompt")
        
            while not self.root_user.exit_status_ready():
                data_decoded = ''
                while self.root_user.recv_ready():
                    data = self.root_user.recv(self.__receive_buffer_size)
                    data_decoded = data
                    print ("The output received after the the password was sent: ", data_decoded)
                if "]#" in data_decoded or "~#" in data_decoded or "~]#" in data_decoded or ">" in data_decoded:
                    break
    
            if self.root_user.recv_stderr_ready():
                data_std_err = self.root_user.recv_stderr(self.__std_data_recv_prompts)
                print ("The error received after the password was sent: ", data_std_err)
            
            #if(remote):
            #    cmd_admin_login = ["admin", "iDirect", "xoff", "rmt " + remote, command, "exit"]

            #cmd_admin_login = ["admin", "iDirect", "xoff", command, "exit"]
            #for cmd in cmd_admin_login:
                # if(cmd == command):
                    # time.sleep(3)
            #    self.root_user(cmd)
            #print("I am here before ssh")
            #print("I am here")
  
        
        return self.root_user
        
        #return "Error reaching node: " + str(ip)

    def stdin_write(stdin, command):
        cmd_stdin = command + "\n"
        stdin.flush()
        time.sleep(0.5)
        stdin.write(cmd_stdin)

    def __del__(self):
        """
        Destructs the ssh client object
        """ 
        self.ssh_client.close()

def request_body_insertion(object_id_returned, return_post):
        # Body which gets injected into the post request 
        body1 = """{
                    "obj_parentid": %s,    
                    "error_correction": "",
                    "framelengthms": "125",
                    "acqguardinterval": "252",
                    "guardband": "252",
                    "carrierindex": "1",
                    "acqmethod": "Superburst",
                    "symbolrate": "5000",
                    "modulation": "",
                    "acqenable": "true",
                    "acqslottype": "",
                    "rxhlcid": "1",
                    "pilotinit": "6239",
                    "pilotpolynomial": "49153",
                    "linecardfrequency": "0",
                    "relativecentrefrequency": "0",
                    "acqslotnum": "1"
                    }
                """ % (object_id_returned)
        body2 = """{
                    "obj_parentid": %d,
                    "compositionindex": "",
                    "composition_list": "{\\"(%s,BPSK,1/2,\\\\\\"No Spreading\\\\\\")\\"}",
                    "defaultigc": "true",
                    "maxallowabledlfade": "8",
                    "adpfixedigc": "false"
                }
                """ % (object_id_returned, str(return_post))
        
        print("The body printed for the post query", body1)
        print (body2)

        return (body1, body2)

def func_include_filename(nms_Ip, nms_user, nms_pass, pp_Ip, pp_User, pp_pass, symbol_list, timer):
#def func_include_filename(filename):
    """
    # Read and assign config variables
    data = json.load(open(filename))
    nms_ip=data['nms_ip']
    nms_username= data['nms_username']
    nms_password= data['nms_password']
    igp_name=data['igp_name']
    pp_ip = data['pp_ip']
    pp_user = data['pp_user']
    pp_password = data['pp_password']
    symbols_custom = data['symbols']
    """
    nms_ip = nms_Ip
    nms_username = nms_user
    nms_password = nms_pass
    pp_ip = pp_Ip
    pp_user = pp_User
    pp_password = pp_pass
    symbols_custom = symbol_list
    time = timer
    
    return (nms_ip, nms_username, nms_password, pp_ip, pp_user, pp_password, symbols_custom, time)     
"""
def get_max_bw_channel_and_beam_id(ip, username, password, api_test_query):
    headers = {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_query) 
    get_info_request_channel_info = requests.get(api_uri, auth=HTTPBasicAuth(username, password), headers=headers)
    
    print get_info_request_channel_info 
    
    get_info_request_in_json = get_info_request_channel_info.json()['data']
    rtn_bw_list = []
    for bw in get_info_request_in_json:
        return_bw_list = rtn_bw_list.append(bw["rtnbandwidth"])
"""                
def get_igps_for_beam(ip, username, password, api_test_query, beam_id):
    headers = {"Content-Type":"application/json"}
    api_uri_igp_query = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_query)
    
    get_info_request_igp_query = requests.get(api_uri_igp_query, auth=HTTPBasicAuth(username, password), headers=headers) 

    get_info_request_igp_query_json = get_info_request_igp_query.json()["data"]

    print get_info_request_igp_query_json

    get_info_request_igp_query_json_pretty = json.dumps(get_info_request_igp_query_json, indent=4, sort_keys=True)

    print get_info_request_igp_query_json_pretty 

    beam_id = beam_id.encode("utf-8")
    beam_id = int(beam_id)
    print ("the type of beam", type(beam_id))


    
    
    dictionary_list = {}
    dictionary_list_ids = {}
    list_of_beams = []
    list_of_beams_values = []
    for beams in get_info_request_igp_query_json:
        beam_list = beams["obj_attributes"]["beamid_list"].encode("utf-8")
        print type(beam_list)
        if beam_list[0] != "":
            print "The first", beam_list[0]
            print "The last element", beam_list[-1]

            beam_list_stripped = beam_list[1:len(beam_list)-1]


            print "test", beam_list_stripped

            beam_list_split = beam_list_stripped.split(",")

            if not beam_list_split:
                print("The Beam has already been deleted, PLEASE CHECK THE IGP FROM PULSE UI....")
                sys.exit(0)

            print ("split beam", beam_list_split)

            if "" in  beam_list_split:
                dictionary_list["IGP_None"] = None
                dictionary_list_ids["IGP_None"] = None
            else:
                for i in range(0, len(beam_list_split), 1):
                    beam_list_split_int = int(beam_list_split[i])
                    print "The beam integer", beam_list_split_int
                    print "The beam integer", type(beam_list_split_int)
                    if beam_list_split_int == beam_id:
                        dictionary_list[beams["obj_name"]] = beam_list_split
                        dictionary_list_ids[beams["obj_id"]] = beam_list_split
    
        print dictionary_list

        print "The dictionary", dictionary_list_ids

        return (beam_id, dictionary_list, dictionary_list_ids)

class Patch_Post_Apply_Changes_IGP:
    def __init__(self, ip, username, password):
        self.ip = ip
        self.username = username
        self.password  = password
    
    
    def patch_query_to_IGP_for_beam_id_deletion(self, dictionary):
        try:
            headers = {"Content-Type":"application/json"}
        
            print "DICTIONARY INSIDE POST ..........", dictionary

            dictionary_list_keys = dictionary.keys()
            dictionary_list_values = dictionary.values()

            print "keys in dict", dictionary_list_keys
            print "values in dict", dictionary_list_values
        

            for count, value in dictionary.items():
                array_keys = np.array(dictionary.values())
                array_keys_1d = array_keys.ravel()
                print "1d array", array_keys_1d
                print "1d array", type(array_keys_1d)
                array_keys_1d = list(array_keys_1d)

                #array_keys_1d = str(array_keys_1d)

            # if not array_keys_1d:
            #     array_keys_1d = " ".join(array_keys_1d)
                print "array_keys_1d", array_keys_1d
                
                body = """{"beamid_list": []}""" 

                print "THE KEYS INSIDE DICTION", count.encode("utf-8")
                print ("The body is", body)

                #body_indent = json.dumps(body, indent=4, sort_keys=True)
                #print ("The body is", body_indent)

                api_test_query_patch = "api/1.0/config/element/inroutegroupprofile/%s?limit=0" %(count.encode("utf-8"))

                api_uri_test = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_patch)

                print api_uri_test
                
                patch_info_request_igp = requests.patch(api_uri_test, data=body, auth=HTTPBasicAuth(self.username, self.password), headers = headers)

                print "<<<REPONSE FOR PATCHING THE EMPTY BEAM LIST>>>", patch_info_request_igp
                print "<<<REPONSE FOR PATCHING THE EMPTY BEAM LIST IN JSON>>>", patch_info_request_igp.json()

                obj_id_list, obj_parent_id = get_inroutegroup_profile_details(self.ip, self.username, self.password, count.encode("utf-8")) 
                print "The object id list is: ", obj_id_list

                if obj_id_list == None or obj_parent_id == None:
                    print("The Beam has already been deleted, PLEASE CHECK THE IGP FROM PULSE UI....")
                    sys.exit(-1)
                return (obj_id_list, obj_parent_id)

        except TypeError:
            print("The Beam has already been deleted, PLEASE CHECK THE IGP FROM PULSE UI....")
            sys.exit(-1)

        
    def apply_changes_IGP(self, obj_id_list, dictionary_igp_beam_id_list):
        
        headers = {"Content-Type":"application/json"}
        
        for obj_id in dictionary_igp_beam_id_list.keys():
            body = """{"id": "%s"}"""%(obj_id)
            print "THE BODY", body
            api_test_query_apply_changes = "api/1.0/config/element/%s/apply_changes" %(obj_id)
            api_uri_apply = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_apply_changes)

            print "APPLY CHANGES RESPONSE", api_uri_apply 
            get_info_request_channel_info = requests.post(api_uri_apply, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers) 

            print "The JSON response for POST", get_info_request_channel_info.json()       

    def post_body_for_new_igp(self, obj_parent_id, Igp_Name,beam_id_retruned_from_channel):
        body =              """
                            {
                            "obj_parentid": %d,
                            "obj_child_num": 0,
                            "obj_name": "%s",
                            "obj_attributes": {
                                "adpupdateinterval": "30",
                                "guardinterval": 252,
                                "adphysteresismargin": "1",
                                "ucplevel2poweradjustvalue": "2",
                                "adpalgorithminterval": "2000",
                                "enablehscomt": "false",
                                "skipinterval": "",
                                "adpfadeslopemargin": "0.5",
                                "ucplevel2lowersnrthreshold": "-3",
                                "adpacquisitionmargin": "",
                                "beamid_list": "[%s]",
                                "ucplevel2uppersnrthreshold": "3",
                                "upstreambandwidth": 12,
                                "adpalloweddropoutfraction": "0.5",
                                "adplogoffinterval": "3",
                                "ucplowersnrthreshold": "",
                                "ucpmaxpowerlevel": "",
                                "framelength": "125",
                                "apertureacq": "46288",
                                "adpcontentionweight": "2",
                                "payload_size": "170",
                                "guardintervaloverwrite": "false",
                                "ucpuppersnrthreshold": "",
                                "adpmeasurementspacing": "1000",
                                "freeslotallocationenabled": "true",
                                "superburstm3": "3"
                                            }   
                            } 
                            """ %(obj_parent_id, Igp_Name, beam_id_retruned_from_channel)
        return body

    def post_IGP_value(self, dictionary_igp_beam_id_list, obj_id_list, obj_parent_id, beam_id_retruned_from_channel):
        headers = {"Content-Type":"application/json"}

        api_test_query = "api/1.0/config/element/inroutegroupprofile"

        print "IGP BEAM LIST IS :", dictionary_igp_beam_id_list
        print "OBJ_ID LIST IS: ", obj_id_list
        print "OBJ PARENT ID IS :",obj_parent_id
        print "OBJECT PARENT ID FROM CHANNEL :", beam_id_retruned_from_channel

        dictionary_list_keys = dictionary_igp_beam_id_list.keys()
        dictionary_list_values = dictionary_igp_beam_id_list.values()

        print "keys in dict igp key list", dictionary_list_keys
        print "values in dict igp value list", dictionary_list_values

        i = 1

        obj_name_list = []
        for count_key_id in dictionary_igp_beam_id_list.keys():
            api_uri_post_query = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query)
            print api_uri_post_query

            get_request_sent_IGP_profile =  requests.get(api_uri_post_query, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
            print get_request_sent_IGP_profile

            get_request_sent_IGP_profile_json = get_request_sent_IGP_profile.json()['data']
            print get_request_sent_IGP_profile_json

            print "The key list is: ", dictionary_igp_beam_id_list.keys()

            
            for igp_name in get_request_sent_IGP_profile_json:
                
                obj_igp_name = igp_name["obj_name"]
                print "I am inside the for loop", obj_igp_name
                 
                if obj_igp_name == "Test_IGP_Test1":
                    i+=1
                    Igp_Name = "Test_IGP_Test%d" %(i) 
                    print "Inside if"
                    body = self.post_body_for_new_igp(obj_parent_id, Igp_Name, beam_id_retruned_from_channel) 
                    print "The body in if is", body
                    get_info_request_channel_info = requests.post(api_uri_post_query, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
            
                    print "JSON format", get_info_request_channel_info.json()
                    print "JSON format data", get_info_request_channel_info.json()['data']

                    get_info_request_igp_created_id = requests.get(api_uri_post_query, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
                    
                    print "1", get_info_request_igp_created_id 

                    get_info_request_igp_created_id_json = get_info_request_igp_created_id.json()['data']
                    
                    print "The get info for igp created is", get_info_request_igp_created_id_json
            
                    for element_in_get in get_info_request_igp_created_id_json:
                        obj_id_igp_created = element_in_get["obj_id"]

                        print "The object ID for the IGP", obj_id_igp_created

                        return (Igp_Name, obj_id_igp_created)

                else:

                    Igp_Name = "Test_IGP_Test1"
                    body = self.post_body_for_new_igp(obj_parent_id, Igp_Name, beam_id_retruned_from_channel) 
                    print "The body in else is", body

                    get_info_request_channel_info = requests.post(api_uri_post_query, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
            
                    print "JSON format", get_info_request_channel_info.json()
                    print "JSON format data", get_info_request_channel_info.json()['data']

                    get_info_request_igp_created_id = requests.get(api_uri_post_query, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
                    
                    print "1", get_info_request_igp_created_id 

                    get_info_request_igp_created_id_json = get_info_request_igp_created_id.json()['data']
                    
                    print "The get info for igp created is", get_info_request_igp_created_id_json
            
                    for element_in_get in get_info_request_igp_created_id_json:
                        obj_id_igp_created = element_in_get["obj_id"]

                        print "The object ID for the IGP", obj_id_igp_created

                        return (Igp_Name, obj_id_igp_created)
            
    def patch_query_to_IGP_for_beam_addition(self, obj_id, beam_id):
        headers = {"Content-Type":"application/json"}

        api_test_get_query = "api/1.0/config/element/inroutegroupprofile/?obj_name=%s" %(obj_id)

        print "GET RESPONSE of the IGP CREATED", api_test_get_query 
        api_uri_get = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_get_query)
        print api_uri_get  

        get_info_request_created_igp = requests.get(api_uri_get, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        print "GET REQUEST INFO", get_info_request_created_igp

        get_info_request_created_igp_json = get_info_request_created_igp.json()['data']
        for data_element in get_info_request_created_igp_json:    
            get_info_request_created_igp_ID = data_element["obj_id"]

        print "The DATA IGP", get_info_request_created_igp_ID


        print "The object ID is", obj_id

        api_test_query = "api/1.0/config/element/inroutegroupprofile/%s" %(get_info_request_created_igp_ID) 
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query)
        print api_uri 

        body = """{"beamid_list": [%s],
                   "upstreambandwidth": "12",
                   "guardinterval": "252"}""" %(beam_id)

        print "The PATCH QUERY BODY IN IGP", body 

        get_info_request_channel_info = requests.patch(api_uri, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        print "This is the patch request given to the NMS", get_info_request_channel_info

        #get_info_request_channel_info_json = get_info_request_channel_info.json()['data']

        #print "The JSON version of the channel", get_info_request_channel_info_json
      
        return (get_info_request_created_igp_ID,beam_id) 

    def get_inet_info_from_beam(self):
        headers = {"Content-Type":"application/json"}
        api_test_query = "api/1.0/config/element/inet"
        api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query)
        print api_uri  

        get_info_request_beam_info = requests.get(api_uri, auth=HTTPBasicAuth(self.username, self.password), headers=headers)

        print("<<<<<THE GET REQUEST IS>>>>", get_info_request_beam_info)

        get_info_request_beam_info_json = get_info_request_beam_info.json()['data']

        print ("The json fomat for the request beam info", get_info_request_beam_info_json)

        body_indent = json.dumps(get_info_request_beam_info_json, indent=4, sort_keys=True)

        print("<<THE BODY INTENT IS>>", body_indent)

        beam_list_obj_name = []
        beam_list_obj_id = []   
        list_beam_list_id = []
        list_inet_id = []
        for data in get_info_request_beam_info_json:
            dictionary_inet_beam_id = {}
            dictionary_inet_id_beam_id = {}
            dictionary_inet_beam_id[data["obj_name"]] = data["obj_attributes"]["beam_id"].encode("utf-8") 
            dictionary_inet_id_beam_id[data["obj_id"]] = data["obj_attributes"]["beam_id"].encode("utf-8") 
            beam_list_obj_name.append(dictionary_inet_beam_id)
            beam_list_obj_id.append(dictionary_inet_id_beam_id)
            list_beam_list_id.append(data["obj_attributes"]["beam_id"].encode("utf-8"))
            list_inet_id.append(data["obj_id"])

        print ("The beam list ID is :", list_beam_list_id )
        print ("The INET ID LIST", list_inet_id)

        diction_d = zip(list_beam_list_id, list_inet_id)

        dictionary_d = dict(diction_d)

        print("THE LISTS COMBINED", diction_d)


        print ("The dictionary is given :", dictionary_d)
        

        print "The dictionary is : ", dictionary_inet_beam_id
        print "The dictionary list is : ", beam_list_obj_name

        print("the dictionary for the inet and beam id ", dictionary_inet_id_beam_id)
        print("the dictionary list for the same ", beam_list_obj_id)




        return (beam_list_obj_name, beam_list_obj_id, dictionary_d)

    def get_IGP_details_created_from_post(self, obj_id):
        headers = {"Content-Type":"application/json"}
        api_test_query_inrtgrpcmp = "api/1.0/config/element/inroutegroupcomposition/?obj_parentid=%d"%(obj_id)
        api_test_query_upstremcarrier = "api/1.0/config/element/upstreamcarrier/?obj_parentid=%d"%(obj_id)

        api_uri_test_upstream_carrier = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_upstremcarrier)
        api_uri_test_inrtgrpcmp = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_inrtgrpcmp) 
        print api_uri_test_upstream_carrier  
        print api_uri_test_inrtgrpcmp  

        get_info_request_upstream_carrier = requests.get(api_uri_test_upstream_carrier, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        get_info_request_upstream_carrier_json = get_info_request_upstream_carrier.json()['data']

        print "JSON", get_info_request_upstream_carrier_json

        #for data in get_info_request_upstream_carrier_json: 
        for nms_iter_obj in get_info_request_upstream_carrier_json: 
            upstream_carrier_list = nms_iter_obj["obj_attributes"]["symbolrate"].encode("utf-8") + "000"
            upstream_carrier_obj_id = nms_iter_obj["obj_id"] 
        
        print "THE UPSTREAM CARRIER  LIST>>>", upstream_carrier_list 
        print "THE UPSTREAM CARRIER  ID>>>", upstream_carrier_obj_id 

        get_info_request_inrtgrpcmp = requests.get(api_uri_test_inrtgrpcmp, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        get_info_request_inrtgrpcmp_json = get_info_request_inrtgrpcmp.json()['data']
        
        for inrtgroup_obj in get_info_request_inrtgrpcmp_json: 
            get_info_request_inrtgrpcmp_json_list = inrtgroup_obj["obj_attributes"]["composition_list"]
        
        print "JSON INROUTEGROUPCOMPOSITION>>", get_info_request_inrtgrpcmp_json_list

        get_info_request_inrtgrpcmp_json_list_split = get_info_request_inrtgrpcmp_json_list.split(",")
        print "SPLIT LIST", get_info_request_inrtgrpcmp_json_list_split

        get_info_request_inrtgrpcmp_json_list_modcod = get_info_request_inrtgrpcmp_json_list_split[-2]
        get_info_request_inrtgrpcmp_json_list_fec = get_info_request_inrtgrpcmp_json_list_split[-3]

        return (upstream_carrier_list, get_info_request_inrtgrpcmp_json_list_modcod, get_info_request_inrtgrpcmp_json_list_fec, upstream_carrier_obj_id)

def exec_change_symbol_rate(sym_rate, headers, ip, carrier_id, username, password, api_uri):
    import requests
    print ("API REQUEST URI IS ", api_uri)
    body = """{"symbolrate": %s}""" %(sym_rate)
    symrate_object = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print("THE GET RESPONSE IS>>>", symrate_object)
    symrate_object_json = symrate_object.json()
    for data_obj in symrate_object_json["data"]: 
        symrate_object_json_obj_id = data_obj["obj_id"]
    API_upstream = "api/1.0/config/element/upstreamcarrier/%s?limit=0" %(symrate_object_json_obj_id) 
    api_uri_upstream = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API_upstream) 
    symrate_change = requests.patch(api_uri_upstream, data=body, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
     
 #   print "SymbolRate Changed [IP: " + str(ip) + "| Carrier: " + str(carrier_id) + "]: " + str(symrate_change.json()['data']['obj_attributes']['symbolrate'])
    return symrate_object_json_obj_id

def change_symbol_rate(ip, username, password, obj_parentid, sym_rate):
    API = "api/1.0/config/element/upstreamcarrier/?obj_parentid=%s" %(obj_parentid)
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
   
    exec_change_symbol_rate(sym_rate, headers, ip, obj_parentid,  username, password, api_uri)

def change_maxmodcod(ip, username, password, obj_parentid, obj_id_ingrpprofile, modcod, fec_rate):
    API = "api/1.0/config/element/inroutegroupcomposition/%s?limit=0" %(obj_id_ingrpprofile + 1)
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    print "The contents of", api_uri

    maxmodcod_change, modcod, fec_rate = exec_func_requests_maxmod(api_uri, headers, ip, username, password, obj_parentid, obj_id_ingrpprofile, modcod, fec_rate)     

    return (maxmodcod_change, modcod, fec_rate)
 
def exec_func_requests_maxmod(api_uri, headers, ip, username, password, obj_parentid, obj_id_ingrpprofile, modcod, fec_rate):
    import requests
    body = """{"composition_list":"{\\"(\\\\\\"%s\\\\\\",\\\\\\"%s\\\\\\",\\\\\\"%s\\\\\\",\\\\\\"No Spreading\\\\\\")\\"}"}""" %(obj_id_ingrpprofile, modcod, fec_rate)
    print ("THE BODY BEFORE TH PATCH REQUEST>>>", body)
    maxmodcod_change = requests.patch(api_uri, data=body, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print "<<<<THE MAXMODCODCHANGE>>>", maxmodcod_change

#    print "Modcod Changed [IP: " + str(ip) + "| TerminalType: " + str(carrier_id) + "]: " + str(maxmodcod_change.json()['data']['obj_attributes']['maxmodcod'])

    return (maxmodcod_change, modcod, fec_rate)

def change_user_frequency(ip, username, password, channel_id, user_frequency):
    API = "api/1.0/config/element/channel"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=channel_id)
   # fwdbandwidth = "54"
    try:
        exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri)     
    finally:
        exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri)     

    # print user_freq_change.content
    # write_access_log(ip, "NMS : change modcod ", str(maxmodcod_change.content))

def exec_change_user_frequency(user_frequency, headers, ip, username, password, channel_id, api_uri):
    import requests
    user_freq_change = requests.patch(api_uri, data=json.dumps({"rtnbandwidth": "35" , "rtnuserfrequency": user_frequency}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
#    print "User Frequency Changed [IP: " + str(ip) + "| Channel: " + str(channel_id) + "]: " + str(user_freq_change.json()['data']['obj_attributes']['fwduserfrequency'])

    return user_freq_change

def change_gw_frequency(ip, username, password, channel_id, gw_frequency):
    API = "api/1.0/config/element/channel"
    API_test = ""
  #  headers = requests.
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=channel_id)
    print "THE API URI IS", api_uri
    try:
        exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri)
    finally:
        exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri)

def exec_change_gw_frequency(gw_frequency, headers, ip, username, password, channel_id, api_uri):
    import requests 
    gw_freq_change = requests.patch(api_uri, data=json.dumps({"rtnbandwidth": "35", "rtngatewayfrequency": gw_frequency}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print "THE JSON OUTPUT", gw_freq_change.json()

def apply_changes(ip, username, password, object_id):
    API = "api/1.0/config/element"
    print ("<<<<<THE OBJECT ID>>>", object_id)
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}/apply_changes'.format(api_endpoint=ip, nms_api=API, obj_id=object_id)
    try:
        apply_change = requests.post(api_uri, data=json.dumps({"id":"%s" %(object_id)}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "The change for the json", apply_change
    finally:
        apply_change = requests.post(api_uri, data=json.dumps({"id":"%s" %(object_id)}), headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "The change for the json", apply_change

def get_user_gw_frequency_of_downstream_for_skipping(ip, username, password, channel_id):
    API = "api/1.0/config/element/channel/%s?limit=0" %(channel_id)
  #  headers = requests.
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}/{obj_id}'.format(api_endpoint=ip, nms_api=API, obj_id=channel_id)
    print "THE API URI IS", api_uri
    try:
        fwd_user_freq, fwd_gw_freq  = exec_get_user_gw_frequency_of_downstream_for_skipping(headers, ip, username, password, channel_id, api_uri)
    finally:
        fwd_user_freq, fwd_gw_freq = exec_get_user_gw_frequency_of_downstream_for_skipping(headers, ip, username, password, channel_id, api_uri)

    return (fwd_user_freq, fwd_gw_freq)

def exec_get_user_gw_frequency_of_downstream_for_skipping(headers, ip, username, password, channel_id, api_uri):
    import requests 
    fwd_user_freq = []
    fwd_gw_freq = []
    gw_freq_change = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
    print "THE JSON OUTPUT", gw_freq_change.json()
    gw_freq_change_json  = gw_freq_change.json() 
    for data in gw_freq_change_json:
        fwd_user_freq.append(data["obj_attributes"]["fwduserfrequency"])
        fwd_gw_freq.append(data["obj_attributes"]["fwdgatewayfrequency"])
    
    return (fwd_user_freq, fwd_gw_freq)

def default_maxmodcod(ip, username, password):
    default_termtype = {}
    API = "api/1.0/config/element/terminaltype"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    print "API URI", api_uri
    try:
        termtype_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for dtype in termtype_info.json()['data']:
            if len(str(dtype['obj_attributes']['maxmodcod'])) > 1:
                default_termtype[dtype['obj_id']] = dtype['obj_attributes']['maxmodcod']
           
    except:
        #os.system("pip install requests")
        import requests
    finally:
        termtype_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "I AM IN TERMTYPE INFO..", termtype_info
        for dtype in termtype_info.json()['data']:
            if len(str(dtype['obj_attributes']['maxmodcod'])) > 1:
                default_termtype[dtype['obj_id']] = dtype['obj_attributes']['maxmodcod'] 
        
    return default_termtype


def get_max_bw_channel_and_beam_id(ip, username, password, api_test_query, direction):
    headers = {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_query)

    print api_uri 
    get_info_request_channel_info = requests.get(api_uri, auth=HTTPBasicAuth(username, password), headers=headers)
    
    print get_info_request_channel_info 
    
    get_info_request_in_json = get_info_request_channel_info.json()['data']
    
    #get_info_request_in_json = json.dumps(get_info_request_in_json, indent=4, sort_keys=True)
    print get_info_request_in_json

    rtn_bw_list = []
    channel_list = []
    dictionary_for_channel_bw_pair = {}
    dictionary_for_channel_id_bw_pair = {}
    
    for bw in get_info_request_in_json:
        if direction == "downstream":
            fw_bw_value = bw["obj_attributes"]["fwdbandwidth"]
            
        rtn_bw_value = rtn_bw_list.append(bw["obj_attributes"]["rtnbandwidth"])
        channel_name = channel_list.append(bw["obj_name"])
        dictionary_for_channel_bw_pair[bw["obj_name"]] = bw["obj_attributes"]["rtnbandwidth"] 
        dictionary_for_channel_id_bw_pair[bw["obj_id"]] = bw["obj_attributes"]["rtnbandwidth"]

    print channel_list
    print rtn_bw_list
    print dictionary_for_channel_bw_pair
    print dictionary_for_channel_id_bw_pair
    
    #for key, value in dictionary_for_channel_bw_pair.iteritems():
        #max_value_in_dict = max(dictionary_for_channel_bw_pair[key])
    
    key_max = max(dictionary_for_channel_bw_pair.keys(), key=(lambda k: dictionary_for_channel_bw_pair[k]))
    key_max_id = max(dictionary_for_channel_id_bw_pair.keys(), key=(lambda k: dictionary_for_channel_id_bw_pair[k]))
    
    print("The dictionary ", dictionary_for_channel_bw_pair[key_max])
    print("The dictionary with id ", dictionary_for_channel_id_bw_pair[key_max_id])
    print("The key maximum", key_max)
    print("The key maximum", key_max_id)

    #print max_value_in_dict

    api_test_beam_query = "api/1.0/config/element/channel/obj_name=%s" %(key_max)
    api_test_beam_query_id = "api/1.0/config/element/channel/obj_id=%s" %(key_max_id)
    api_uri_beam_query = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_beam_query) 
    api_uri_beam_query_id = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=api_test_beam_query_id) 

    print api_uri_beam_query
    print api_uri_beam_query_id

    get_info_request_beam_info = requests.get(api_uri_beam_query, auth=HTTPBasicAuth(username, password), headers=headers)
    get_info_request_beam_info_id = requests.get(api_uri_beam_query_id, auth=HTTPBasicAuth(username, password), headers=headers)
    get_info_request_beam_info_json = get_info_request_beam_info.json()['data']
    get_info_request_beam_info_json_id = get_info_request_beam_info_id.json()['data']
    
    beam_id_value = get_info_request_beam_info_json["obj_attributes"]["beam_id"] 
    print beam_id_value

    api_get_igp_query = "api/1.0/config/element/inroutegroupprofile"

    beam_id_retrurned, dictionary_present, dictionary_list_igp_id = get_igps_for_beam(ip, username, password, api_get_igp_query, beam_id_value) 
    
    print("The type of beam id", type(beam_id_retrurned))
    print("The type of beam id name", beam_id_retrurned)
    

    print "DICTIONARY OF ITEMS BEFORE FILTER :///",dictionary_present

    for key, value in dictionary_present.items():
        print dictionary_present[key]
        print key
    
    list_for_IGP_keys = dictionary_present.keys() 
    list_for_IGP_values = dictionary_present.values()


    print "The list of values", list_for_IGP_values

    list_for_IGP_values_without_none = filter(None, list_for_IGP_values)

    print "The list without the None values", list_for_IGP_values_without_none

    print "THE DICTIONARY FOR IGP ITEMS :::",dictionary_list_igp_id 

    for key1, value1 in dictionary_list_igp_id.items():
        print dictionary_list_igp_id[key1]
        print key1
    
    list_for_IGP_keys_ID = dictionary_list_igp_id.keys()
    list_for_IGP_values_ID = dictionary_list_igp_id.values()

    print "The list of valyes for the key IDs :", list_for_IGP_values_ID

    
    
    print("After filter", list_for_IGP_values_without_none)

    

    print("The list BEFORE FOR", list_for_IGP_values_without_none)
    print("The list BEFORE FOR", type(list_for_IGP_values_without_none))

    print ("The length of the IGP values", len(list_for_IGP_values_without_none)) 
    
    dictionary_igp_beam_list = {}
    dictionary_igp_beam_id_list = {}
    for count in range(0, len(list_for_IGP_values_without_none), 1):
        if len(list_for_IGP_values_without_none) != 1 :
            print "Every element", list_for_IGP_values_without_none[count]
            element_in_list = list_for_IGP_values_without_none[count] 
            print "element", type(element_in_list)
            if str(beam_id_retrurned) in element_in_list:
                element_in_list.remove(str(beam_id_retrurned))
            
            a = np.array(list_for_IGP_values_without_none)
            list_for_IGP_values_without_none = a.ravel() 
            print "value of b", list_for_IGP_values_without_none
            list_for_IGP_values_without_none = list(list_for_IGP_values_without_none)
        
        else:
            a = np.array(list_for_IGP_values_without_none)
            list_for_IGP_values_without_none = a.ravel() 
            print "value of b", list_for_IGP_values_without_none
            list_for_IGP_values_without_none = list(list_for_IGP_values_without_none)

            print "value of b", type(list_for_IGP_values_without_none)
            list_for_IGP_values_without_none = set(list_for_IGP_values_without_none) 
            print "after list from numpy", type(list_for_IGP_values_without_none)
            
            list_for_IGP_values_without_none = list(list_for_IGP_values_without_none)
            
            list_for_IGP_values_without_none.remove(str(beam_id_retrurned))

        print "After operation", list_for_IGP_values_without_none
        
        for key, value in dictionary_present.items():
            dictionary_igp_beam_list[key] = list_for_IGP_values_without_none
            if key == "IGP_None":
                del dictionary_igp_beam_list[key]
        
        for key_igp, value_igp in dictionary_list_igp_id.items():
            print "keys", dictionary_list_igp_id.keys()
            print "values", dictionary_list_igp_id.values()
            dictionary_igp_beam_id_list[key_igp] = list_for_IGP_values_without_none
            #if key_igp == "IGP_None":
            #    del dictionary_igp_beam_id_list[key_igp]
    

    print "The new dictionary", dictionary_igp_beam_list

    print "The dictionary which contains the IGP Obj ID", dictionary_igp_beam_id_list

    patch_obj = Patch_Post_Apply_Changes_IGP(ip, username, password)

    
    try:
        patch_func_response_obj_id, patch_func_response_obj_parent_id = patch_obj.patch_query_to_IGP_for_beam_id_deletion(dictionary_igp_beam_list)
    except TypeError:
        print("The Beam has already been deleted and reattached to the new IGP, PLEASE CHECK THE IGP FROM PULSE UI....")
        sys.exit(-1)

    if patch_func_response_obj_id ==  None or patch_func_response_obj_parent_id ==None:
        print("The return value for the function is None since the Beam is already deleted....")
        sys.exit(0)

    print patch_func_response_obj_id

    patch_post_apply  = patch_obj.apply_changes_IGP(patch_func_response_obj_id, dictionary_igp_beam_id_list)

    igp_name_return_value, obj_id_IGP_created = patch_obj.post_IGP_value(dictionary_igp_beam_id_list, patch_func_response_obj_id, patch_func_response_obj_parent_id, beam_id_value)

    #obj_id, obj_parent_id = get_inroutegroup_profile_details(ip, username, password, igp_name_return_value)

    print "The returned value from the p[ost IGP function", igp_name_return_value

    obj_id, beam_id = patch_obj.patch_query_to_IGP_for_beam_addition(igp_name_return_value, beam_id_value)

    dictionary_beam_id_returned = patch_obj.get_inet_info_from_beam()
    
    print("the dictionary for the returned VALUE", dictionary_beam_id_returned)
    ##print ("returned_obj_type", return_post)
    
    return (obj_id, beam_id, dictionary_igp_beam_id_list, dictionary_igp_beam_list, key_max_id, dictionary_list_igp_id, obj_id_IGP_created) 

class PP_Validation_Check_for_Freq:
    def __init__(self, ip, username, password):
        self.ip = ip
        self.username = username
        self.password = password
    
    #def pp_iac_from_command_prompt(self):
    #    ssh_object = downstream.ssh_session(self.ip, self.username, self.password)
     #   #downstream.
     #   return ssh_object
    
    def ssh_console(self):
        ssh_console_obj = Ssh_Console(self.ip, self.username, self.password)
        std_output_for_command, std_error_command = ssh_console_obj.send_string_and_wait("ls -lrt", exit_on_error= False)

        std_output_for_command = std_output_for_command.encode("utf-8")

        std_output_for_command_after_split = std_output_for_command.split("\r\n")

        print ("The std output after the split to an array", std_output_for_command_after_split)

        print("The first element", std_output_for_command_after_split[0])
        print("The last element", std_output_for_command_after_split[-1])

        #del std_output_for_command_after_split[0]
        #del std_output_for_command_after_split[-1]
        #del std_output_for_command_after_split[-1]

        print ("List is as follows", std_output_for_command_after_split) 
        stdout_command = ssh_console_obj.telnet_run_command("13256", "cluster status")
       
        stdout_str_telnet_output = ssh_console_obj.send_the_command_in_telnet("xoff")
        
        print ("The stdout telnet", stdout_str_telnet_output)
        
        stdout_str_telnet_output_actual = ssh_console_obj.send_the_command_in_telnet("cluster status")

        stdout_str_telnet_output_actual_json = stdout_str_telnet_output_actual[stdout_str_telnet_output_actual.find("{"): stdout_str_telnet_output_actual.rfind("}") + 1]
        
        stdout_str_telnet_output_actual_json1 = json.loads(stdout_str_telnet_output_actual_json)
        stdout_str_telnet_output_actual_json12 = json.dumps(stdout_str_telnet_output_actual_json1)

        print "JSON IS>>>", stdout_str_telnet_output_actual_json12  
        
        ip_list = []
        for item in stdout_str_telnet_output_actual_json1["NODES"]:
            ip_list.append(str(item["ip"].split(";")[1]))
            #ip_address = stdout_str_telnet_output_actual_json12["NODES"]["ip"]

        print("The IP Addresss is ", ip_list)

        ip_list = "".join(ip_list)

        print "The IP ADDRESS>>>>", ip_list

            #print "The IP adress of the DC is", test["ip"]
        
        print("The std out command for read", stdout_str_telnet_output_actual_json)
        #print ("The lines seperated after the command has been split", std_output_for_command_after_split)
        #std_output_for_command_after_split_deletion_of_last_str = std_output_for_command_after_split.encode("utf-8")

        return (stdout_str_telnet_output_actual_json, ip_list)
    
    #def telnet_module(self, ssh):
    
    def DC_node_IP_address(self, DC_node):
        ssh_console_DC_PP = Ssh_Console(DC_node, self.username, self.password)

        std_output_for_command, std_error_command = ssh_console_DC_PP.send_string_and_wait("ifconfig -a", exit_on_error= False)

        print ("The output for the DC Node is", std_output_for_command)

        ssh_console_DC_PP.telnet_run_command("13257", "nodes table")

        ssh_console_DC_PP.send_the_command_in_telnet("xoff")
        output_for_nodes_table_ssh_console = ssh_console_DC_PP.send_the_command_in_telnet("nodes table | grep pp_iac")

        output_for_nodes_table_ssh_console_str = output_for_nodes_table_ssh_console.encode("utf-8")

        output_for_nodes_table_pp_iac = output_for_nodes_table_ssh_console[output_for_nodes_table_ssh_console.find("\r\n") : output_for_nodes_table_ssh_console.rfind("\r\n[icrm]") + 1] 

        output_for_nodes_table_pp_iac_decode = output_for_nodes_table_pp_iac.encode("utf-8")
        
        output_for_nodes_table_pp_iac_decode_split = output_for_nodes_table_pp_iac_decode.split("\r\n")
        
        del output_for_nodes_table_pp_iac_decode_split[0]
        del output_for_nodes_table_pp_iac_decode_split[-1]
        print("The output for nodes table ", output_for_nodes_table_ssh_console_str)

        print("The pp_iac process table", output_for_nodes_table_pp_iac_decode)
        print("The pp_iac process table split", output_for_nodes_table_pp_iac_decode_split)
        
        split_list = [] 
        for test_item in output_for_nodes_table_pp_iac_decode_split:
            output_for_nodes_table_pp_iac_decode_split_after_split = test_item.split("|")
            split_list.append(output_for_nodes_table_pp_iac_decode_split_after_split)
        print ("the list value is given by", split_list)
        
        ip_list = []
        inet_id_list =[]
        

        print " The lebnhgthj odf the split l;ist", len(split_list)
        
        for j in range(0, len(split_list), 1):
            print("The split list is given by: ", split_list[j][1])
            inet_id_list.append(split_list[j][1])

        for k in range(0, len(split_list), 1):
            print("The split list for dictionary : ", split_list[k][0])
            ip_list.append(split_list[k][0])
        
        print ("The IP Address list is given ", ip_list)
        print ("The INET list is", inet_id_list)

        
        dictionary_ip_inet = zip(ip_list, inet_id_list)
        dictionary_ip_inet = dict(dictionary_ip_inet)

        print "The dictionary is given by :", dictionary_ip_inet 

        ##for test in range(1, len(output_for_nodes_table_pp_iac_decode_split_after_split), 5):
        #    output_for_nodes_table_pp_iac_decode_split_after_split[test]

    def stats_validation_pp(self, DC_node, inet_id):
        ssh_console_DC_PP = Ssh_Console(DC_node, self.username, self.password)

        std_output_for_command, std_error_command = ssh_console_DC_PP.send_string_and_wait("ifconfig -a", exit_on_error= False)

        std_output_for_command_inet, std_error_command_inet = ssh_console_DC_PP.send_string_and_wait("ps -ef | grep pp_iac | grep %s" %(inet_id), exit_on_error= False)

        print ("The std_output for the command for the inet ID", std_error_command_inet)
        try:
            std_output_for_command_inet_array = std_error_command_inet[0].encode("utf-8").split("-")
        except:
            std_output_for_command_inet_array = std_output_for_command_inet.split("-")
            for item in std_output_for_command_inet_array:
                if "cp" in item:
                    cp_value = item
                    print "THE CP VALUE", cp_value

                    console_port = cp_value[3:]
                    print "THE CONSOLE PORT>>>", console_port
             
        print "THE CONSOLE PORT", std_output_for_command_inet_array

        for cp in std_output_for_command_inet_array:
            if "cp" in cp:
                cp_value = cp
                print "THE CP VALUE", cp_value

                console_port = cp_value[3:]
                print "THE CONSOLE PORT>>>", console_port

                ssh_console_DC_PP.telnet_run_command(console_port, "nodes table")

            ssh_console_DC_PP.send_the_command_in_telnet("xoff")

            output_for_iac_command_sitm_symbol = ssh_console_DC_PP.send_the_command_in_telnet("iac 1 sitm tct | grep Symbol") 
            output_for_iac_command_sitm_modulation = ssh_console_DC_PP.send_the_command_in_telnet("iac 1 sitm tct | grep Modulation") 
            output_for_iac_command_sitm_fec = ssh_console_DC_PP.send_the_command_in_telnet("iac 1 sitm tct | grep FEC") 

            output_for_iac_command_sitm_encoded_symbol =  output_for_iac_command_sitm_symbol.encode("utf-8")
            output_for_iac_command_sitm_encoded_modulation =  output_for_iac_command_sitm_modulation.encode("utf-8")
            output_for_iac_command_sitm_encoded_fec =  output_for_iac_command_sitm_fec.encode("utf-8")

#        print ("THE OUTPUT>>>", type(output_for_iac_command_sitm_ssh_console_encoded))
#        print ("THE OUTPUT>>>", output_for_iac_command_sitm_ssh_console_encoded)

        print ("THE MOD VALUE>>", output_for_iac_command_sitm_encoded_modulation)
        output_for_iac_command_sitm_encoded_modulation_values = output_for_iac_command_sitm_encoded_modulation.splitlines()
        print ("THE MOD VALUE AFTER SPLIT", output_for_iac_command_sitm_encoded_modulation_values)
        for j in range(0, 2, 1):
            del output_for_iac_command_sitm_encoded_modulation_values[-1]

        sitm_mod_value = output_for_iac_command_sitm_encoded_modulation_values [1].strip("\t")
        sitm_mod_value_space = re.findall(r'\S+', sitm_mod_value)
        print "THE MODULATION", sitm_mod_value_space
        sitm_mod_value_space_element = sitm_mod_value_space[1][:4]
        sitm_mod_value_space_element_str = "".join(sitm_mod_value_space_element)
        print "THE VALUE", sitm_mod_value_space_element_str

        print ("THE SYMBOL VALUE>>", output_for_iac_command_sitm_encoded_symbol)
        output_for_iac_command_sitm_encoded_symbol_values = output_for_iac_command_sitm_encoded_symbol.splitlines()
        print ("THE SYMBOL VALUE AFTER SPLIT", output_for_iac_command_sitm_encoded_symbol_values)
        for j in range(0, 2, 1):
            del output_for_iac_command_sitm_encoded_symbol_values[-1]
        
        sitm_sym_value_list = []
        sitm_sym_value = output_for_iac_command_sitm_encoded_symbol_values [1]
        sitm_sym_value_strip = sitm_sym_value[5:-1]
        sitm_sym_value_list.append(sitm_sym_value_strip)
        print "Symbol Rate", sitm_sym_value_list

        sitm_sym_value_acq_on = output_for_iac_command_sitm_encoded_symbol_values [2]
        #print sitm_sym_value_acq_on
        if "Symbol rate" in sitm_sym_value_acq_on: 
            sitm_sym_value_acq_on_strip = sitm_sym_value_acq_on[5:-1] 
            sitm_sym_value_list.append(sitm_sym_value_acq_on_strip)
            print "MOD VALUES FOR INITIAL AND WITH ACQ ON>>>", sitm_sym_value_list
            sitm_sym_mod_values = "".join(sitm_sym_value_list[-2])[13:-1]
            print "THE FIRST VALUE IN LIST", sitm_sym_mod_values
              
        else:
            print "MOD VALUES FOR INITIAL AND WITH ACQ ON>>>", sitm_sym_value_list
            sitm_sym_mod_values = "".join(sitm_sym_value_list)[13:-1]
            print "THE FIRST VALUE IN LIST", sitm_sym_mod_values

        print ("THE FEC VALUE>>", output_for_iac_command_sitm_encoded_fec)
        output_for_iac_command_sitm_encoded_fec_values = output_for_iac_command_sitm_encoded_fec.splitlines()
        print("THE FEC VALUE AFTER SPLIT", output_for_iac_command_sitm_encoded_fec_values)
        for j in range(0, 2, 1):
            del output_for_iac_command_sitm_encoded_fec_values[-1]

        sitm_fec_value = output_for_iac_command_sitm_encoded_fec_values [2].strip("\t")[6:-2]
        print "SITM VALUE", sitm_fec_value

        return (sitm_fec_value, sitm_sym_mod_values, sitm_mod_value_space_element_str, ssh_console_DC_PP)

def modcod_fec_rate_func(modcod, fec_rate_bpsk_list, fec_rate_8psk_list, fec_rate_qpsk_list):
    switcher_dictionary = {
                            "BPSK": BPSK, 
                            "8PSK": EPSK,
                            "QPSK": QPSK
                        }
    print switcher_dictionary.get(modcod, lambda: "Invalid Modcod.....")    

    def BPSK(fec_rate_bpsk_list):
        bpsk_combined_list = []
        for test in fec_rate_bpsk_list:
            bpsk_combined_list.append("BPSK"+ test)
        print bpsk_combined_list


        return bpsk_combined_list
    
    def EPSK(fec_rate_8psk_list):
        epsk_combined_list = []
        for test in fec_rate_epsk_list:
            epsk_combined_list.append("8PSK"+ test)
        
        print epsk_combined_list 

        return epsk_combined_list
    
    def QPSK(fec_rate_qpsk_list):
        qpsk_combined_list = []
        for test in fec_rate_qpsk_list:
            qpsk_combined_list.append("8PSK"+ test)
        
        print qpsk_combined_list 

        return qpsk_combined_list

def change_reference_symbol_rate_cn(nms_ip, nms_username, nms_password, key_max_id, symbol, modcod):
    API = "api/1.0/config/element/channel/obj_id=%s" %(key_max_id)
    headers =  {"Content-Type":"application/json"}
    api_uri_reference = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=nms_ip, nms_api=API)

    if modcod == "BPSK":
        print("IN REFERENCE BPSK>>>")
        body = """{"rtnrefsymrate": "%s",
                    "rtnrefcn" : "1"}""" %(symbol)
    
    elif modcod == "QPSK":
        print("IN REFERENCE QPSK>>>")
        body = """{"rtnrefsymrate": "%s",
                    "rtnrefcn" : "7"}""" %(symbol)
    
    elif modcod == "8PSK":
        print("IN REFERENCE 8PSK>>>")
        body = """{"rtnrefsymrate": "%s",
                    "rtnrefcn" : "14"}""" %(symbol)
        
    reference_info_request_channel_info = requests.patch(api_uri_reference, data=body, auth=HTTPBasicAuth(nms_username, nms_password), headers=headers)
    print "THE RESPONSE FOR PATCH OF REFERENCE>>>", reference_info_request_channel_info
    
    api_test_query_apply_changes_reference = "api/1.0/config/element/%s/apply_changes" %(key_max_id)
    api_uri_apply_changes = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=nms_ip, nms_api=api_test_query_apply_changes_reference)

    body_apply_changes = """{"id" : "%s"}""" %(key_max_id)
    print "APPLY CHANGES RESPONSE", api_uri_apply_changes
    get_info_request_channel_info = requests.post(api_uri_apply_changes, data=body, auth=HTTPBasicAuth(nms_username, nms_password), headers=headers) 

    print "The JSON response for POST", get_info_request_channel_info.json()
     
    

def modcod_change_fec_iteration(fec_iteration, nms_ip, nms_username,nms_password, obj_id, upstream_carrier_id, modcod, dictionary_d, beam_id, pp_validation_check, ssh_console_obj_stdout_IP, symbol_rate, modulation_set, fec_rate, pp_ip, pp_user, pp_password, console_port_str, ssh_console_DC_PP1, term_did_list, pp_did_list, term_ip_list, pp_ip_list, work_b, carrier, gw, u_freq, dictionary_test, dictionary_list_combined, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols):
    for fec in fec_iteration:
        #print "<<<TERMTYPE>>>", term_type
        #print "<<<TERMTYPE LIST>>>", term_termtype_list
        return_response, modcod_patched, fec_rate_patched = change_maxmodcod(nms_ip, nms_username, nms_password, obj_id, upstream_carrier_id, modcod, fec)
        apply_changes(nms_ip, nms_username, nms_password, obj_id)

        print "Sleep, to let the configuration propogate......"
        time.sleep(60)
        for key, value in dictionary_d.items():
            if key == beam_id:
                sitm_fec_value, sitm_sym_value_list, sitm_mod_value, console_obj_PP = pp_validation_check.stats_validation_pp(ssh_console_obj_stdout_IP, dictionary_d[key])
                print
                time.sleep(60)
                print("Sleep to let the configurations to propogate....")
                sys.stdout.flush()
                print("=====================================")
                print "SYMBOL RATE FROM NMS ", symbol_rate 
                print "MODULATION FROM NMS ", modcod_patched 
                print "FEC RATE FROM NMS ", fec_rate_patched 
                print "SYMBOL RATE FROM PP ", sitm_sym_value_list 
                print "MODULATION RATE FROM PP ", sitm_mod_value 
                print "FEC RATE RATE FROM PP ", sitm_fec_value
                print("=====================================")
                time.sleep(2) 

                if sitm_fec_value != fec_rate_patched and modcod_patched !=sitm_mod_value and sitm_sym_value_list != symbol_rate:
                    print("THE VALIDATION HAS FAILED BETWEEN NMS AND PP DUE TO THE MISMATCH OF THE VALUES.....")
                    print("THE SCRIPT IS EXITING...")
                    sys.exit(-1)

                #Reset PP stats
                pp_ips_list = []
                pp_ips_list.append(ssh_console_obj_stdout_IP)
                for ip in pp_ips_list:
                    ssh_console_obj_return =  reset_pp_stats(ip, pp_user, pp_password, console_port_str)
                
                    


                # Run commands on PP
                for ip in pp_ip_list:
                    print("EXCEL SHEET HIT....")
                    print("Before running the PP Commands")
                    pp_commands_execution(ip, pp_user, pp_password, console_port_str, work_b, carrier, gw, u_freq, ssh_console_DC_PP1, dictionary_test, ssh_console_obj_return, dictionary_list_combined, modcod, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols,flag=True)

                # Run commands on carrierRemote
                #for did in pp_did_list:
                #    get_terminal_stats(term_ip_list[str(term_did_list[did])], "root", "iDirect")
                #    if DEBUG:
                #        print "DID: " + str(did)
                #        print "Terminal ID: " + str(term_did_list[did])
                #        print "Terminal IP: " + str(term_ip_list[str(term_did_list[did])])
 
def get_terminal_did_list(ip, username, password, tpa_port):
    command = "rmt list"
    ssh_console_sep = Ssh_Console(ip, username, password)
    ssh_console_sep.telnet_run_command(tpa_port, "xoff")
    ssh_console_sep.send_the_command_in_telnet("xoff")
    
    cmd_ouput_body = ssh_console_sep.send_the_command_in_telnet(command)
    #console_obj_PP.telnet_run_command(tpa_port, command)
    #raw_input("PRESS ENTER TO CONTINUE AFTER RMT STATS...")
    #time.sleep(4) 
    print("<<<THE COMMAND OUTPUT BODY>>>", cmd_ouput_body)
    
    cmd_ouput_body_split = cmd_ouput_body.splitlines()
    print("The command output>>>>", cmd_ouput_body_split)

    term_list = []
    
    for line in cmd_ouput_body_split: 
        if('CX' in line):
            term_list.append(line.split('(')[1].split(')')[0])
    
    print "The term list", term_list
    
    return (term_list, ssh_console_sep) 

def reset_pp_stats(ip, username, password, tpa_port):
    terminal_list, ssh_console_obj_return = get_terminal_did_list(ip, username, password, tpa_port)
    for term in terminal_list:
        ssh_console_obj_return.send_the_command_in_telnet("rmt" + term)
        cmd_output_body_for_reset = ssh_console_obj_return.send_the_command_in_telnet("rh uh stats reset")

        print ("The STATS RESET FOR THE REMOTE>>>>", cmd_output_body_for_reset)
    return ssh_console_obj_return
            

row1 = 1
#en_list_length=0 
def excel_sheet_generation(dictionary_set, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, modcod, enumerate_list, work_book_name, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols, direction="upstream"):
    global row1
    global en_list_length 
    #row = 1
    
    try:
        work_b = openpyxl.Workbook()
        work_s = work_b.active
        #work_s1 = work_b.create_sheet()
        
#        if column
        for i in zip(value):
            #work_s['A%s' %(i)] = cmd_input_for_excel 
            work_s.append(i)
         #   work_s1.append(j)
            #work_s['Command Output'] = cmd_output
            test_work = work_b.save("xlwt c:\\Users\\akarthik\\Downloads\\Upstream_Excel_File.xls")
    except: 
        #os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.chart import (LineChart, ScatterChart, Reference, Series)
        from openpyxl.chart.axis import DateAxis
        from copy import deepcopy

    finally:
        def plotting_charts(series_obj, Number_counter):
            #series = Series(data, xvalues, title_from_data=True)
            s1 = series_obj.series[0]
            #if s2 == "":
            s1.marker.symbol = "triangle"
            s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
            s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

            s1.graphicalProperties.line.noFill = True

            
            s2 = series_obj.series[1]
            #if s1 == "":
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FF9900" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FF9900" # Marker outline
            #s2.graphicalProperties.line.width = 100050 # width in EMUs

            """ 
            s2 = c1.series[2]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c1.series[3]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FF6A00" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FF6A00" # Marker outline

            s2 = c1.series[4]
            s2.graphicalProperties.line.solidFill = "FFB200"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c1.series[5]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "FFE100" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "FFE100" # Marker outline

            s2 = c1.series[6]
            s2.graphicalProperties.line.solidFill = "E9FF00"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c1.series[7]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "8CFF00" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "8CFF00" # Marker outline
            """

            s2 = series_obj.series[2]
            s2.graphicalProperties.line.solidFill = "00FF61"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = series_obj.series[3]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "00FFA5" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "00FFA5" # Marker outline
            
            """
            s2 = c1.series[10]
            s2.graphicalProperties.line.solidFill = "00FFF2"
            s2.graphicalProperties.line.dashStyle = "sysDot"

            s2 = c1.series[11]
            s2.marker.symbol = "triangle"
            s2.marker.graphicalProperties.solidFill = "00D0FF" # Marker filling
            s2.marker.graphicalProperties.line.solidFill = "00D0FF" # Marker outline
            """

            s2 = series_obj.series[4]
            s2.graphicalProperties.line.solidFill = "003FFF"
            s2.graphicalProperties.line.dashStyle = "sysDot"
            
                    # Style the lines
            Number_counter+=1 
            
            return series_obj
        import copy
        #print("The value of row", row)
        item_redundancy_elim = []
        #print ("THE COUNTER TO CHECK THE LENGTH OF ENUMERATE LIST>>>>>>", en_list_length)
        print ("THE LENGTH OF THE ENUMERATE LIST>>>>>>", enumerate_list)

        if len(enumerate_list) < 2:
            print("------------------------------------------------")
            print("THIS IS THE FIRST ITERATION>>>>>>>>>>>>")
            print("The set of keys are given by", dictionary_set.keys())
            print("The set of values are given by", dictionary_set.values())
            for item_redundancy in dictionary_set.values():
                item_only_element = set(item_redundancy)
                print("The set of values are given by", item_only_element)
                item_redundancy_elim.append(item_only_element)

            print("<<<<<THE REDUNANDCY LIST>>>>", item_only_element) 
            print("<<<<<THE MODCOD LIST LENGTH IS GIVEN>>>>", modcod_list_length)
            active_object_state = openpyxl.load_workbook(work_book_name)
            
            work_s = active_object_state.active
            
            #work_s2 = active_object_state.get_sheet_by_name("Chart")

            #work_s1 = work_b.create_sheet(
            cell1 = work_s.cell(row=1, column=1)
        # wb = copy(rb) # a writable copy (I can't read values out of this, only write to it)
        # w_sheet = wb.get_sheet(0)
            #i = 0
            #j = 0 
            
            dictionary_list_values = [] 

            for i in range(0, len(item_redundancy_elim), 1):
                dictionary_values_set = ''.join(item_redundancy_elim[i])
                dictionary_list_values.append(dictionary_values_set)
            
            print ("The VALUE AFTER JOIN>>>>>>", dictionary_list_values)
        # print("COUNTER", counter)

            for n, i in enumerate(dictionary_list_values):
                if i == "":
                    dictionary_list_values[n] = "_"
            for item in range(0, len(dictionary_list_values), 1):
                if dictionary_list_values[item] != "_":
                    dictionary_list_values[item] = int(dictionary_list_values[item])
                    print(type(dictionary_list_values[item]))            
            
            print ("THE VALUE AFTER THE UNDERSCORE", dictionary_list_values) 

            if not dictionary_list_combined:
                print(type(dictionary_list_combined))
                print("The contents of the str object", dictionary_list_combined) 
                dictionary_list_combined.append(dictionary_set.keys())
                dictionary_list_combined.append(dictionary_list_values)

            else:
                dictionary_list_combined.append(dictionary_list_values) 


            print ("The dictionary list of keys and values", dictionary_list_combined)
            
            if cell1.value == None:
                #@static_var("row1", 1)
                #row1 =1
                i=0
                for key in range(1, 14):
                    print("I am inside the for loop for the first row execution", row1)
                    #work_s['A%s' %(i)] = cmd_input_for_excel 
                    cell_ref = work_s.cell(row = row1, column = key)
                    cell_ref.value = dictionary_set.keys()[i]
                    if key == 13:
                        cell_ref.value = dictionary_set.keys()[12] 
                    i+=1

            if not cell1.value == None:
                print("I am here when cell value is not none")
                """
                for n, i in enumerate(dictionary_list_values):
                if i == "":
                    dictionary_list_values[n] = "_"
                for item in range(0, len(dictionary_list_values)-1, 1):
                if dictionary_list_values[item] != "_":
                    dictionary_list_values[item] = int(dictionary_list_values[item])
                    print(type(dictionary_list_values[item]))            
            
                print (dictionary_list_values) 
                """
                
                print("j when 0 : ", dictionary_list_values[0])
                print("j when checked for last element : ", dictionary_list_values[-1])
                j=0
            #for row_counter in range(0, len(dictionary_list_values)-1, 1):
                
                if dictionary_list_values.count("_") == 13:
                    del dictionary_list_values
        
                else:
                    print("IN ELSE BEFORE EXECUTING THE FOR LOOP>>>>", row1)
                    print("<<THE DICTIONARY LIST VALUES>>", dictionary_list_values)
                    row1+=1
                    for col in range(1, 14):
                        print("ROWWW IN FOR", row1)
                        cell_ref_col = work_s.cell(row = row1, column = col)
                        cell_ref_col.value = dictionary_list_values[j]
                        if col == 13:
                            cell_ref_col.value = dictionary_list_values[12]
                        j+=1
                #row1+=1
                    print("row value after the for loop for col is :", row1)
            categories = Reference(work_s, min_col=1, min_row=1, max_col=13) 
            if modcod == "BPSK":
                c1 = LineChart()
                c1.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c1.style = 13
                c1.y_axis.title = 'Upstream Bursts'
                c1.x_axis.title = 'Sampled Interval'
                c1.x_axis.scaling.min = 0
                #c1.y_axis.scaling.min = 0
                c1.x_axis.scaling.max = len(fec_rate_bpsk_list)*2+1
                #c1.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                #for item_col in work_s.iter_cols():
                #    for item in item_col:
                data = Reference(work_s, min_col=1, min_row=1, max_col=13, max_row=len(fec_rate_bpsk_list)*2+1)
                c1.add_data(data, titles_from_data=True)
                #c1.set_categories(categories)
                #c1.legend.delete = True
                obj_state = plotting_charts(c1, Number_counter)
                work_s.add_chart(obj_state, "A48")

            elif modcod == "8PSK":
                c2 = LineChart()
                c2.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c2.style = 13
                c2.y_axis.title = 'Upstream Bursts'
                c2.x_axis.title = 'Sampled Interval'
                c2.x_axis.scaling.min = len(fec_rate_bpsk_list)*2+2
                #c1.y_axis.scaling.min = 0
                c2.x_axis.scaling.max = len(fec_rate_8psk_list)*2+3
                #c2.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                data1 = Reference(work_s, min_col=1, min_row=1, max_col=13, max_row=len(fec_rate_8psk_list)*2+3)
                #c2.set_categories(categories)
                c2.add_data(data1, titles_from_data=True)
                #c2.legend.delete = True
                for item in dictionary_set.keys():
                    print("TYPE", type(item))
                     
                #openpyxl.chart.legend.Legend(legendPos='r', legendEntry= dictionary_set.keys(), layout=None, overlay=None, spPr=None, txPr=None, extLst=None)
                #c2.legend.legendEntry = (dictionary_set.keys(), expected_type=Rich)
                obj_state=plotting_charts(c2, Number_counter)
                work_s.add_chart(obj_state, "K48")

            elif modcod == "QPSK":
                c3 = LineChart()
                c3.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c3.style = 13
                c3.y_axis.title = 'Upstream Bursts'
                c3.x_axis.title = 'Sampled Interval'
                c3.x_axis.scaling.min = len(fec_rate_8psk_list)*2+3
                #c1.y_axis.scaling.min = 0
                c3.x_axis.scaling.max = len(fec_final_rate)*2
                #c3.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                    #for row in dictionary_list_combined:
                    #    work_s.append(row)
                data2 = Reference(work_s, min_col=1, min_row=1, max_col=13, max_row=len(fec_rate_qpsk_list)*2+4)
                c3.set_categories(categories)
                c3.add_data(data2, titles_from_data=True)
                c3.legend.delete = True
                obj_state = plotting_charts(c3, Number_counter)
                work_s.add_chart(obj_state, "H70")
            
            work_s.merge_cells("A31:E31")
            if len(symbols) == 1:
                work_s["O2"] = "THE SYMBOL IS %s" %(symbols[0])     
                if modcod == "BPSK":
                    work_s["O3"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O4"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O5"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 2:
                work_s["O6"] = "THE SYMBOL IS %s" %(symbols[1])     
                if modcod == "BPSK":
                    work_s["O7"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O8"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O9"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 3:
                work_s["O10"] = "THE SYMBOL IS %s" %(symbols[2])  
                if modcod == "BPSK":
                    work_s["O11"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O12"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O13"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)        
            elif len(symbols) == 4:
                work_s["O14"] = "THE SYMBOL IS %s" %(symbols[3])     
                if modcod == "BPSK":
                    work_s["O15"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O16"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O17"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 5:
                work_s["O18"] = "THE SYMBOL IS %s" %(symbols[4])     
                if modcod == "BPSK":
                    work_s["O19"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O20"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O21"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 6:
                work_s["O22"] = "THE SYMBOL IS %s" %(symbols[5])     
                if modcod == "BPSK":
                    work_s["O23"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O24"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O25"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 7:
                work_s["O26"] = "THE SYMBOL IS %s" %(symbols[6])     
                if modcod == "BPSK":
                    work_s["O27"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O28"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O29"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 8:
                work_s["O30"] = "THE SYMBOL IS %s" %(symbols[7])     
                if modcod == "BPSK":
                    work_s["O31"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O32"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O33"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 9:
                work_s["O34"] = "THE SYMBOL IS %s" %(symbols[8])     
                if modcod == "BPSK":
                    work_s["O35"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O36"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O37"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 10:
                work_s["O38"] = "THE SYMBOL IS %s" %(symbols[9])     
                if modcod == "BPSK":
                    work_s["O39"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s["O40"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s["O41"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
                 
            work_s["O41"] = "We are using a single carrier with User Frequency=%s and Gateway Frequency= %s" %(u_freq, gw)     
        
            print("The row is given by", row1)
            test_work = active_object_state.save(os.path.expanduser("~\\Upstream_Excel_File.xlsx")) 

            if row1 ==  len(fec_final_rate):
                del work_s



        
        #en_list_length+=1
        
        else: 

            print("_______________________________---------------")
            print("THIS IS THE SECOND ITERATUIION>>>>>>>>>>>>>>>>>>>")
            
            print("The set of keys are given by", dictionary_set.keys())
            print("The set of values are given by", dictionary_set.values())
            for item_redundancy in dictionary_set.values():
                item_only_element = set(item_redundancy)
                print("The set of values are given by", item_only_element)
                item_redundancy_elim.append(item_only_element)

            print("<<<<<THE REDUNANDCY LIST>>>>", item_only_element) 
            print("<<<<<THE MODCOD LIST LENGTH IS GIVEN>>>>", modcod_list_length) 
            work_s2 = active_object_state.get_sheet_by_name(sheet_name)
            
            #work_s2 = active_object_state.get_sheet_by_name("Chart")

            #work_s1 = work_b.create_sheet(
            cell1 = work_s2.cell(row=1, column=1)
        # wb = copy(rb) # a writable copy (I can't read values out of this, only write to it)
        # w_sheet = wb.get_sheet(0)
            #i = 0
            #j = 0 
            
            dictionary_list_values = [] 

            for i in range(0, len(item_redundancy_elim), 1):
                dictionary_values_set = ''.join(item_redundancy_elim[i])
                dictionary_list_values.append(dictionary_values_set)
            
            print ("The VALUE AFTER JOIN>>>>>>", dictionary_list_values)
        # print("COUNTER", counter)

            for n, i in enumerate(dictionary_list_values):
                if i == "":
                    dictionary_list_values[n] = "_"
            for item in range(0, len(dictionary_list_values), 1):
                if dictionary_list_values[item] != "_":
                    dictionary_list_values[item] = int(dictionary_list_values[item])
                    print(type(dictionary_list_values[item]))            
            
            print ("THE VALUE AFTER THE UNDERSCORE", dictionary_list_values) 

            if not dictionary_list_combined:
                print(type(dictionary_list_combined))
                print("The contents of the str object", dictionary_list_combined) 
                dictionary_list_combined.append(dictionary_set.keys())
                dictionary_list_combined.append(dictionary_list_values)

            else:
                dictionary_list_combined.append(dictionary_list_values) 


            print ("The dictionary list of keys and values", dictionary_list_combined)
            
            if cell1.value == None:
                #@static_var("row1", 1)
                row1 =1
                i=0
                for key in range(1, 14):
                    print("I am inside the for loop for the first row execution", row1)
                    #work_s['A%s' %(i)] = cmd_input_for_excel 
                    cell_ref = work_s2.cell(row = row1, column = key)
                    cell_ref.value = dictionary_set.keys()[i]
                    print("THIS IS THE SECOND 2nd FIRST ROW EXECUTION>>>")
                    if key == 13:
                        cell_ref.value = dictionary_set.keys()[12] 
                    i+=1

            if not cell1.value == None:
                print("I am here when cell value is not none")
                """
                for n, i in enumerate(dictionary_list_values):
                if i == "":
                    dictionary_list_values[n] = "_"
                for item in range(0, len(dictionary_list_values)-1, 1):
                if dictionary_list_values[item] != "_":
                    dictionary_list_values[item] = int(dictionary_list_values[item])
                    print(type(dictionary_list_values[item]))            
            
                print (dictionary_list_values) 
                """
                
                print("j when 0 : ", dictionary_list_values[0])
                print("j when checked for last element : ", dictionary_list_values[-1])
                j=0
            #for row_counter in range(0, len(dictionary_list_values)-1, 1):
                
                if dictionary_list_values.count("_") == 13:
                    del dictionary_list_values
        
                else:
                    print("IN ELSE BEFORE EXECUTING THE FOR LOOP>>>>", row1)
                    print("<<THE DICTIONARY LIST VALUES>>", dictionary_list_values)
                    row1+=1
                    for col in range(1, 14):
                        print("ROWWW IN FOR", row1)
                        cell_ref_col = work_s2.cell(row = row1, column = col)
                        cell_ref_col.value = dictionary_list_values[j]
                        if col == 13:
                            cell_ref_col.value = dictionary_list_values[12]
                        j+=1
                #row1+=1
                    print("row value after the for loop for col is :", row1)
                    
                #   work_s1.append(j)
                    #work_s['Command Output'] = cmd_output

            
           
            if modcod == "BPSK":
                c1 = LineChart()
                c1.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c1.style = 13
                c1.y_axis.title = 'Upstream Bursts'
                c1.x_axis.title = 'Sampled Interval'
                c1.x_axis.scaling.min = 0
                #c1.y_axis.scaling.min = 0
                c1.x_axis.scaling.max = modcod_list_length
                c1.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                data = Reference(work_s2, min_col=1, min_row=1, max_col=13, max_row=4)
                
                c1.add_data(data, titles_from_data=True)
                obj_state = plotting_charts(c1, Number_counter)
                work_s2.add_chart(obj_state, "A48")

            elif modcod == "8PSK":
                c2 = LineChart()
                c2.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c2.style = 13
                c2.y_axis.title = 'Upstream Bursts'
                c2.x_axis.title = 'Sampled Interval'
                c2.x_axis.scaling.min = 0
                #c1.y_axis.scaling.min = 0
                c2.x_axis.scaling.max = modcod_list_length
                c2.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                data1 = Reference(work_s2, min_col=1, min_row=1, max_col=13, max_row=8)

                c2.add_data(data1, titles_from_data=True)
                obj_state=plotting_charts(c2, Number_counter)
                work_s2.add_chart(obj_state, "K48")

            elif modcod == "QPSK":
                c3 = LineChart()
                c3.title = "Gateway Frequency = %s,User Frequency = %s, Modcod=%s" %(gw, u_freq, modcod)
                c3.style = 13
                c3.y_axis.title = 'Upstream Bursts'
                c3.x_axis.title = 'Sampled Interval'
                c3.x_axis.scaling.min = 0
                #c1.y_axis.scaling.min = 0
                c3.x_axis.scaling.max = modcod_list_length
                c3.x_axis.unit = 5
                #c1.y_axis.scaling.max = 70000
                    #for row in dictionary_list_combined:
                    #    work_s.append(row)
                data2 = Reference(work_s2, min_col=1, min_row=1, max_col=13, max_row=12)
                c3.add_data(data2, titles_from_data=True)
                obj_state = plotting_charts(c3, Number_counter)
                work_s2.add_chart(obj_state, "H70")

            
            work_s2.merge_cells("A31:E31")
            #work_s2["A32"] = "We are using a single carrier with MODCOD Sheet2=%s" %(modcod)     
            """
            if modcod == "BPSK":
                work_s2["O10"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)+2)     
            elif modcod == "8PSK":
                work_s2["O11"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)+2)     
            elif modcod == "QPSK":
                work_s2["O12"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)+2) 
            """
            if len(symbols) == 1:
                work_s2["O2"] = "THE SYMBOL IS %s" %(symbols[0])     
                if modcod == "BPSK":
                    work_s2["O3"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O4"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O5"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 2:
                work_s2["O6"] = "THE SYMBOL IS %s" %(symbols[1])     
                if modcod == "BPSK":
                    work_s2["O7"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O8"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O9"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 3:
                work_s2["O10"] = "THE SYMBOL IS %s" %(symbols[2])  
                if modcod == "BPSK":
                    work_s2["O11"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O12"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O13"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)        
            elif len(symbols) == 4:
                work_s2["O14"] = "THE SYMBOL IS %s" %(symbols[3])     
                if modcod == "BPSK":
                    work_s2["O15"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O16"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O17"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 5:
                work_s2["O18"] = "THE SYMBOL IS %s" %(symbols[4])     
                if modcod == "BPSK":
                    work_s2["O19"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O20"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O21"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 6:
                work_s2["O22"] = "THE SYMBOL IS %s" %(symbols[5])     
                if modcod == "BPSK":
                    work_s2["O23"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O24"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O25"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 7:
                work_s2["O26"] = "THE SYMBOL IS %s" %(symbols[6])     
                if modcod == "BPSK":
                    work_s2["O27"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O28"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O29"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 8:
                work_s2["O30"] = "THE SYMBOL IS %s" %(symbols[7])     
                if modcod == "BPSK":
                    work_s2["O31"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O32"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O33"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 9:
                work_s2["O34"] = "THE SYMBOL IS %s" %(symbols[8])     
                if modcod == "BPSK":
                    work_s2["O35"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O36"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O37"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            elif len(symbols) == 10:
                work_s2["O38"] = "THE SYMBOL IS %s" %(symbols[9])     
                if modcod == "BPSK":
                    work_s2["O39"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_bpsk_list)*2)     
                elif modcod == "8PSK":
                    work_s2["O40"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_8psk_list)*2)     
                elif modcod == "QPSK":
                    work_s2["O41"] = "THE MODCOD IS=%s and corresponds to %s rows" %(modcod, len(fec_rate_qpsk_list)*2)     
            
            work_s2["O42"] = "We are using a single Upstream carrier with User Frequency=%s and Gateway Frequency= %s" %(u_freq, gw)     
        
            print("The row is given by", row1)
            test_work = active_object_state.save(os.path.expanduser("~\\Upstream_Excel_File.xlsx")) 

            if work_s2 == len(fec_final_rate):
                del work_s2

            return row1

                
        # s = "s" 
        # for k in range(3, 34):
        #     var = s +"%s" %(k)
        #     var = c1.series[k]
        #     var.graphicalProperties.line.solidFill = "00AAAA"
        #     var.graphicalProperties.line.dashStyle = "sysDot"  
                
                        
                #work_s2 = copy.deepcopy(work_s)
         
    #        elif row1 == 40:
    #           work_s2.add_chart(c1, "A2") 

            
                 
        """This is the code section for the line chart"""
        
    
         

        

    
        # Style the lines
    
    


def pp_commands_execution(ip, pp_user, pp_password, console_port_tpa, work_b, carrier, gw, u_freq, ssh_console_DC_PP1, dictionary_test, ssh_console_obj_return, dictionary_list_combined,modcod, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols,flag=False):

    terminal_list = get_terminal_did_list(ip, pp_user, pp_password, console_port_tpa)

    cmd_output_array = []
    list_seperation_for_elements = []
    list_to_have_seperated_values = []
    list_bytes = []
    list_values = []
    list_bytes_test = []
    list_values_test = []

    dictionary_test_final  = {}

    list_bytes_test_new = []
     
    list_values_test_new = []

    dictionary_test["rx_traffic_bursts"] = [] 
    dictionary_test["rx_acquisition_bursts"] = []
    dictionary_test["invalid_inroute_burst"] = []
    dictionary_test["num_acq_good"] = []
    dictionary_test["num_acq_crc"] = []
    dictionary_test["num_acq_ghost"] = []
    dictionary_test["num_acq_stolen"] = []
    dictionary_test["num_acq_missing"] = []
    dictionary_test["num_traffic_good"] = []
    dictionary_test["num_traffic_crc"] = []
    dictionary_test["num_traffic_ghost"] = []
    dictionary_test["num_traffic_stolen"] = []
    dictionary_test["num_traffic_missing"] = []

    Number_counter = 1     
    for term in terminal_list:
        print("<<<INSIDE TERMINAL LIST FOR LOOP>>")
        #cmd_output, cmd_err = ssh_console_DC_PP1.telnet_run_command(ip, pp_user, pp_password, console_port_tpa, "modcod", term)
        # print find_command_output_body(cmd_output, "modcod") + "\n"
        #print cmd_output + "\n"
        # write_access_log(ip, "modcod", cmd_ouput)

        cmd_output = ssh_console_obj_return.send_the_command_in_telnet("dvbs2rmt snr")
        # write_access_log(ip, "dvbs2rmt snr", cmd_ouput)

        cmd_output = ssh_console_obj_return.send_the_command_in_telnet("rh uh stats")
        print "The command output for the dvbs2rmt stats \n", cmd_output
        print type(cmd_output)
        sys.stdout.flush()

        cmd_output_section = cmd_output[cmd_output.find("rx_traffic_bursts = "): cmd_output.rfind("admin@telnet:127.0.0.1") + 1]
        print "Command output section \n", cmd_output_section
        
        
        print("cmd after flush", cmd_output_section)

       
        cmd_output_array.append(cmd_output_section)
        print "The command output array is ", cmd_output_array

        if not cmd_output_array[-1]:
            cmd_output_array.pop(-1)

        #cmd_output_array = list(cmd_output_array)
        
        #print "list is given by", cmd_output_array

        cmd_output_array_Upstream_Excel_File = cmd_output_array[0].split("\r\n")

        print "The cmd_output after split", cmd_output_array_Upstream_Excel_File

        cmd_output_array_Upstream_Excel_File.pop(-1)
        if not cmd_output_array_Upstream_Excel_File[-1]:
            cmd_output_array_Upstream_Excel_File.pop(-1)
        
        print "The array after pop", cmd_output_array_Upstream_Excel_File
        
        for elements_sep in range(0, len(cmd_output_array_Upstream_Excel_File), 1):
            cmd_output_array_Upstream_Excel_File_after_split = cmd_output_array_Upstream_Excel_File[elements_sep].split(" = ")
            print "output array", cmd_output_array_Upstream_Excel_File_after_split
            list_to_have_seperated_values.append(cmd_output_array_Upstream_Excel_File_after_split)
        
        print "The list is:", list_to_have_seperated_values
        print "The length list is:", len(list_to_have_seperated_values)



        print ("The list of key value pairs", list_to_have_seperated_values)
        print ("The list of key value pairs 1st element", list_to_have_seperated_values[0][0])
        print ("The list of key value pairs 2nd element", list_to_have_seperated_values[1][0])
        print ("The list of key value pairs 3rd element", list_to_have_seperated_values[2][0])

        
        for i in range(0, len(list_to_have_seperated_values), 1):
            dictionary_test_final[list_to_have_seperated_values[i][0]] = list_to_have_seperated_values[i][1]
        
        print "The dictionary which contains the keys and values", dictionary_test
        print "The dictionary test final which contains the keys and values", dictionary_test_final

        for key in list(dictionary_test_final.keys()):
            if dictionary_test_final[key] == "0":
                #dictionary_test.pop(value)
                del dictionary_test_final[key]

        print "The dictionary after the key value has been popped", dictionary_test
        print " I am here after the dict values have been popped"
        #print "The list of bytes present", list_bytes
        #print "The list of values present", list_values
        """
        for i in range(0, len(list_values), 1):
            if list_values[i] == "0":
                list_values.remove(i)
        """     
        #for count in dictionary_test:
        #    list_bytes_test = list(dictionary_test.keys())
        #    list_values_test = list(dictionary_test.values())
    
        if "rx_traffic_bursts" in dictionary_test_final.keys():
            dictionary_test["rx_traffic_bursts"].append(dictionary_test_final["rx_traffic_bursts"])
            set(dictionary_test["rx_traffic_bursts"])
            print("The dictionary_test1", dictionary_test)
            #work_sheet_object_excel1 = func_contain_excel_generation(dictionary_test["ByteTotal_Mode_qpsk_1_4"])
        
        if "rx_acquisition_bursts" in dictionary_test_final.keys():
            dictionary_test["rx_acquisition_bursts"].append(dictionary_test_final["rx_acquisition_bursts"])
            set(dictionary_test["rx_acquisition_bursts"])
            
            print("The dictionary_test2", dictionary_test)
            #work_sheet_object_excel2 = func_contain_excel_generation(dictionary_test["ByteTotal_Mode_qpsk_1_3"])

        if "invalid_inroute_burst" in dictionary_test_final.keys():
            
            dictionary_test["invalid_inroute_burst"].append(dictionary_test_final["invalid_inroute_burst"])
            set(dictionary_test["invalid_inroute_burst"])
            
            print("The dictionary_test3", dictionary_test)
        
        if "num_acq_good" in dictionary_test_final.keys():
            dictionary_test["num_acq_good"].append(dictionary_test_final["num_acq_good"])
            set(dictionary_test["num_acq_good"])

            print("The dictionary_test4", dictionary_test)

        if "num_acq_crc" in dictionary_test_final.keys():
            dictionary_test["num_acq_crc"].append(dictionary_test_final["num_acq_crc"])
            set(dictionary_test["num_acq_crc"])

            print("The dictionary_test5", dictionary_test)
        
        if "num_acq_ghost" in dictionary_test_final.keys():
            dictionary_test["num_acq_ghost"].append(dictionary_test_final["num_acq_ghost"])
            set(dictionary_test["num_acq_ghost"])

            print("The dictionary_test6", dictionary_test)

        if "num_acq_stolen" in dictionary_test_final.keys():
            dictionary_test["num_acq_stolen"].append(dictionary_test_final["num_acq_stolen"])
            set(dictionary_test["num_acq_stolen"])

            print("The dictionary_test7", dictionary_test)

        if "num_acq_missing" in dictionary_test_final.keys():
            dictionary_test["num_acq_missing"].append(dictionary_test_final["num_acq_missing"])
            set(dictionary_test["num_acq_missing"])

            print("The dictionary_test8", dictionary_test)

        if "num_traffic_good" in dictionary_test_final.keys():
            
            dictionary_test["num_traffic_good"].append(dictionary_test_final["num_traffic_good"])
            set(dictionary_test["num_traffic_good"])

        if "num_traffic_crc" in dictionary_test_final.keys():
            
            dictionary_test["num_traffic_crc"].append(dictionary_test_final["num_traffic_crc"])
            set(dictionary_test["num_traffic_crc"])

        if "num_traffic_ghost" in dictionary_test_final.keys():

            dictionary_test["num_traffic_ghost"].append(dictionary_test_final["num_traffic_ghost"])
            set(dictionary_test["num_traffic_ghost"])

        if "num_traffic_stolen" in dictionary_test_final.keys():

            dictionary_test["num_traffic_stolen"].append(dictionary_test_final["num_traffic_stolen"])
            set(dictionary_test["num_traffic_stolen"])

        if "num_traffic_missing" in dictionary_test_final.keys():
            
            dictionary_test["num_traffic_missing"].append(dictionary_test_final["num_traffic_missing"])
            set(dictionary_test["num_traffic_missing"])


        list_bytes_test_new = dictionary_test.keys()
        list_values_test_new = dictionary_test.values()
            
        
            #list_values_test.append(list_values)

        print dictionary_test.keys()
        print dictionary_test.values() 
        print "Bytes list sent", list_bytes_test_new
        print "Bytes values sent", list_values_test_new
        print "The keys in the list are given by", list_bytes
        print "The list After the 0 values are removed", list_values

        """
        work_book_name = "c:\\Users\\akarthik\\Downloads\\Upstream_Excel_File.xlsx"
        if os.path.exists(work_book_name):
            print "The workbook is given by"
            active_object_state = openpyxl.load_workbook(work_book_name)
           # for key,value in kwargs.iteritems():
            excel_sheet_generation(dictionary_test, active_object_state, work_book_name)
        """

        print("the dictionary is given as follows", dictionary_test)

        downstream_modified_gw_user_pair.func_contain_excel_generation(dictionary_test, dictionary_list_combined, Number_counter, carrier, gw, u_freq, modcod_list_length, modcod, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols, direction = "upstream") 

        


            #if " = " in cmd_output_array_Upstream_Excel_File[elements_sep]:
            #    cmd_output_array_Upstream_Excel_File[elements_sep].strip(" = ")
            #    list_seperation_for_elements.append(cmd_output_array_Upstream_Excel_File[elements_sep])

        #print "The list after striping the required pattern", list_seperation_for_elements 



        """ 
        for count in range(cmd_output_array_Upstream_Excel_File):
            if "\n" in cmd_output_array_Upstream_Excel_File[count]:
                cmd_output_array_Upstream_Excel_File[count].strip("\n")
        """
        print cmd_output_array_Upstream_Excel_File

        print type(list_bytes)
        print type(list_values)

        #print ("The TEST_OBJ RETURN VALUE", test_obj)
        # write_access_log(ip, "dvbs2rmt stats", cmd_ouput)
    
    return terminal_list, dictionary_test

def get_terminal_list(ip, username, password):
    term_did_list = {}
    API = "api/1.0/config/element/satelliterouter"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
    try:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_did_list[term['obj_attributes']['did']] = term['obj_id'] 
    finally:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_did_list[term['obj_attributes']['did']] = term['obj_id']
        print ("The dictionary for the did list", term_did_list) 

    return term_did_list

def ping_terminals_from_nms(ip, username, password, terminal_ip):
    #cmd_ping = ['ping', '-c', '10', terminal_ip, '>', '/dev/null', '&']
    cmd_ping = "ping -c 4 %s" %(terminal_ip)
    ssh_console_obj_terminals = Ssh_Console(ip, "idirect", "iDirect")
    
    std_out_for_ping, std_error_for_ping = ssh_console_obj_terminals.send_string_and_wait(cmd_ping, exit_on_error=False) 
    if "PING" in std_out_for_ping:
        std_out_for_ping = ""
    print("THE STD OUT FOR PING>>>", std_out_for_ping)
    print("THE STD ERR FOR PING>>>", std_error_for_ping)
    
    return std_out_for_ping, std_error_for_ping
    
    
    
    



def get_terminal_info(ip, username, password):
    term_ip_list = {}
    term_termtype_list = []
    API = "api/1.0/config/element/terminal"
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)
     
    try:     
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        for term in terminal_info.json()['data']:
            term_ip_list[term['obj_attributes']['coremodule_id']] = term['obj_attributes']['mgmtipaddress']
            term_termtype_list.append(term['obj_attributes']['terminaltype_id'])
    #        break
    except:
        #os.system("pip install requests, json")
        import requests, json
    #    brea
    finally:
        terminal_info = requests.get(api_uri, data=None, headers=headers, params=None, auth=HTTPBasicAuth(username, password))
        print "I AM IN TERMINAL INFO...", terminal_info
        for term in terminal_info.json()['data']:
            term_ip_list[term['obj_attributes']['coremodule_id']] = term['obj_attributes']['mgmtipaddress']
            term_termtype_list.append(term['obj_attributes']['terminaltype_id'])

    return term_ip_list, term_termtype_list

"""
def get_pp_info(ip, username, password):
    cmd_ouput = telnet_run_command(ip, username, password, "13256", "cluster status")
    
    write_access_log(ip, "cluster status", cmd_ouput)
    
    test = process_cluster_output(cmd_ouput, "cluster status")
    json_cmd = json.loads(test)

    ip_list = []
    for item in json_cmd["NODES"]:
        ip_list.append(str(item["ip"].split(";")[1]))
    
    print "The IP LIST is : ", ip_list
    
    return ip_list
"""

def get_and_post_details_NMS(ip, username, password, beam_id):
    API = "api/1.0/config/element/channel/beam_id=%s" %(beam_id)
    headers =  {"Content-Type":"application/json"}
    api_uri = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=ip, nms_api=API)

class Revert_Config_Changes:
    def __init__(self, ip, username, password):
        self.ip = ip
        self.username = username
        self.password = password

    def delete_igp_created(self, igp_obj_id):
        API = "api/1.0/config/element/inroutegroupprofile/%s?limit=0" %(igp_obj_id)
        headers =  {"Content-Type":"application/json"}
        api_uri_igp_delete = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=API)
        
        print("THE URI FOR DELETION>>>>>>", api_uri_igp_delete)
        terminal_info = requests.delete(api_uri_igp_delete, auth=HTTPBasicAuth(self.username, self.password), headers=headers)
        print("THE DELETION ACTION IS AS FOLOBJECT IDLOWS FOR THE IGP CREATED>>>>", terminal_info)

    def apply_changes_deleted_igp(self, igp_obj_id):
        body = """{"id": "%s"}"""%(igp_obj_id)
        headers =  {"Content-Type":"application/json"}
        
        print "THE BODY", body
        api_test_query_apply_changes = "api/1.0/config/element/%s/apply_changes" %(igp_obj_id)
        api_uri_apply = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_apply_changes)

        print "APPLY CHANGES RESPONSE FOR DELETION////", api_uri_apply 
        get_info_request_channel_info = requests.post(api_uri_apply, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers) 
        print("THE POST RESPONSE FOR THE APPLY CHANGES>>>", get_info_request_channel_info)

    def revert_back_the_beam_for_old_IGP(self, dictionary_list_igp_id, beam_id):
        
        body = """{"beamid_list": [%s]}""" %(beam_id)
        print ("THE BODY TO BE POSTED TO IGP", body) 
        headers =  {"Content-Type":"application/json"}
         
        for igp_id in dictionary_list_igp_id.keys():
            api_query_igpid = "api/1.0/config/element/inroutegroupprofile/obj_id=%s" %(igp_id)
            api_uri_revert_changes = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_query_igpid) 
           
            print ("THE URI FOR CHANGES IS GIVEN BY>>>>>", api_uri_revert_changes)
            get_info_request_channel_info = requests.patch(api_uri_revert_changes, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers) 

            print ("THE RESPONSE FOR CHANGES>>>>", get_info_request_channel_info)

        channel_id_keys_list = dictionary_list_igp_id.keys()

        return channel_id_keys_list 
    
    def apply_changes_for_reattaching_beam(self, channel_id_keys_list):
        headers =  {"Content-Type":"application/json"}

        for item_keys in channel_id_keys_list: 
            body = """{"id": "%s"}"""%(item_keys)
            print "THE BODY", body
            api_test_query_apply_changes_reattaching = "api/1.0/config/element/%s/apply_changes" %(item_keys)
            api_uri_apply_changes = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=self.ip, nms_api=api_test_query_apply_changes_reattaching) 
           
            print ("THE URI FOR APPLY CHANGES IS GIVEN BY>>>>>", api_uri_apply_changes)
            get_info_request_channel_info = requests.post(api_uri_apply_changes, data=body, auth=HTTPBasicAuth(self.username, self.password), headers=headers)

            print ("THE RESPONSE FOR CHANGES>>>>", get_info_request_channel_info)

def time_calculation(timer):
    time_set = int(timer) / 12
    time_set_sec = time_set / time_set
    print("THE TYPE OF TIME SET", type(time_set_sec))
    print("THE TYPE OF TIME SET111", (time_set_sec))
    raw_input("STOP...")
    
    return time_set_sec 

def main(nms_Ip, nms_user, nms_pass, pp_Ip, pp_User, pp_pass, symbol_list, time):
#def main(nms_Ip, nms_user, nms_pass, pp_Ip, pp_User, pp_pass, symbol_list):
    try:
        filename = "c:\\Users\\akarthik\\Downloads\\config_upstream.json" 
        
        nms_ip, nms_username, nms_password, pp_ip, pp_user, pp_password, symbols_custom, sleep_timer = func_include_filename(nms_Ip, nms_user, nms_pass, pp_Ip, pp_User, pp_pass, symbol_list, time) 
        print("THE RETURNED VALUE FOR TIME VALUE", sleep_timer)
        raw_input("Wait to check for the value of the timer..")
        time_set = time_calculation(sleep_timer)
        print(" The value of the time set as per the calculation", time_set) 
        #nms_ip, nms_username, nms_password, pp_ip, pp_user, pp_password, symbols_custom = func_include_filename(filename) 
        #get_network_name_and_id(nms_ip, nms_username, nms_password)
        igp_name = "Test_IGP_Test1"
        object_id_returned, object_parent_id_returned = get_inroutegroup_profile_details(nms_ip, nms_username, nms_password, igp_name)
        
        pp_validation_check = PP_Validation_Check_for_Freq(pp_ip, pp_user, pp_password) 
        #pp_validation_check.pp_iac_from_command_prompt()
        ssh_console_obj_stdout = pp_validation_check.ssh_console()[0]
        
        ssh_console_obj_stdout_IP = pp_validation_check.ssh_console()[1]

        pp_validation_check.DC_node_IP_address(ssh_console_obj_stdout_IP)

        term_did_list = get_terminal_list(nms_ip, nms_username, nms_password)

        term_ip_list, term_termtype_list = get_terminal_info(nms_ip, nms_username, nms_password)

        """Section to get the console port of pp_tpa process"""
        ssh_console_DC_PP1 = Ssh_Console(ssh_console_obj_stdout_IP, pp_user, pp_password)

        std_output_for_command, std_error_command = ssh_console_DC_PP1.send_string_and_wait("ifconfig -a", exit_on_error= False)

        std_output_for_command_pp_tpa_out, std_error_command_inet = ssh_console_DC_PP1.send_string_and_wait("ps -ef | grep pp_tpa", exit_on_error= False)      

        print ("The std_output_error for the command for the inet ID", std_error_command_inet)
        print ("The std_output for the command for the inet ID", std_output_for_command_pp_tpa_out)

        if "~]#" in std_error_command_inet[-1]:
            del std_error_command_inet[-1]

        #for item in std_error_command_inet:
        #    if "pp_tpachk" in std_error_command_inet:
        

        
        for item1 in std_error_command_inet:
            
            if "pp_tpachk -pid" in item1 or "pp_tpachk -ipc" in item1 or "tpa_checkpoint" in item1:
                print ("I AM IN CHECK POINT>>>")
                std_error_command_inet.remove(item1)

        for item2 in std_error_command_inet:
            if "grep pp_tpa" in item2:
                print ("I AM IN CHECKPOINT 2")
                std_error_command_inet.remove(item2) 

        print ("THE LIST IS GIVEN BY>>>>>", std_error_command_inet)

        print ("THE LASTE ELEMENT IN THE ARRAY>>>>>", std_error_command_inet[-1])

        std_output_for_command_inet_array = std_error_command_inet[-1].encode("utf-8").split("-")

        print "THE CONSOLE PORT", std_output_for_command_inet_array

        console_port_array = []
        for cp_tpa in std_output_for_command_inet_array:
            if "cp" in cp_tpa:
                cp_value_tpa = cp_tpa
                print "THE CP VALUE TPA", cp_value_tpa

                console_port_tpa = cp_value_tpa[3:-1]
                print "THE CONSOLE PORT TPA>>>", console_port_tpa
                console_port_array.append(console_port_tpa)

        console_port_str = "".join(console_port_array)

        print("THE CONSOLE PORT STR>>>>", console_port_str)

        #raw_input("Enter to continue...")
        """This ends the module for getting the console port"""

        pp_ip_list = []

        pp_ip_list.append(ssh_console_obj_stdout_IP)

        print ("<<THE PP IP LIST>>>", pp_ip_list)

        pp_did_list = []
        for ip in pp_ip_list:
            pp_did_list = pp_did_list + get_terminal_did_list(ip, pp_user, pp_password, console_port_str)[0]

        print ("The DID LIST", pp_did_list)

        print("parent_id", type(object_id_returned))
        
        if len(symbols_custom) != 10:
            symbols = symbols_custom
            print("<<<<<THE SYMBOL LIST IS >>>>>>", symbols)
        
        else:
            symbols = [500]  
            print("<<<<<THE SYMBOL LIST IS >>>>>>", symbols)
        
        modcod_list = ["BPSK", "8PSK", "QPSK"]
        #modcod_list = ["BPSK"]

        #fec_rate_bpsk_list = ["1/2", "2/3"]
        fec_rate_bpsk_list = ["1/2"]

        #fec_rate_8psk_list = ["3/4", "6/7", "2/3", "4/5"]
        fec_rate_8psk_list = ["3/4"]

        #fec_rate_qpsk_list = ["3/4", "6/7", "1/2", "2/3"]
        fec_rate_qpsk_list = ["3/4"]

        fec_final_rate = fec_rate_bpsk_list + fec_rate_8psk_list + fec_rate_qpsk_list  
        
        #gw_freq_list = [1.5, 1.55, 1.6, 1.65, 1.7, 1.75, 1.8, 1.85, 1.9, 1.95, 1.10]
        gw_freq_list = [1.5, 1.55]
        
        #user_freq_list = [19.80, 19.85, 19.90, 19.95, 20.0, 20.05, 20.10, 20.15, 20.20, 20.25]
        user_freq_list = [19.80, 19.85]

        gw_user_pair = zip(gw_freq_list, user_freq_list)
        
        #Body insertion for the request posted to the NMS GUI
        #return_type = 1
        #body1, body2 = request_body_insertion(object_id_returned, return_type)
        
        #Object created to create the instance of the post request class
        
        #post_obj = Post_Request_Reverse_Carrier(nms_ip, nms_username, nms_password, object_id_returned, object_parent_id_returned)
        
        #return_post = post_obj.request_for_upstream(body1, "api/1.0/config/element/upstreamcarrier")[1]
        #return_post = str(return_post)
        ##print ("returned_obj_type", return_post)

        #body1, body2 = request_body_insertion(object_id_returned, return_post) 

        #return_obj_type_inroutecomp = post_obj.request_for_composition(body2, "api/1.0/config/element/inroutegroupcomposition")

        ##print ("The return value for the inroutegroup", return_obj_type_inroutecomp)
        


        #Object created to delete the instance of the delete request class  
        
        ##delete_obj = Delete_Request_Reverse_Carrier(nms_ip, nms_username, nms_password, object_id_returned, object_parent_id_returned)
        ##delete_obj.delete_inroutegroup_composition("api/1.0/config/element/inroutegroupcomposition/?limit=0&obj_id=%s"%(return_obj_type_inroutecomp))
        
        ##delete_obj.delete_upstreamcarrier("api/1.0/config/element/upstreamcarrier/?limit=0&obj_id=%s" %(return_post))
        #delete_obj.delete_upstreamcarrier("api/1.0/config/element/inroutegroupcomposition/?limit=0&obj_id=%s"%(return_obj_type_inroutecomp))
        ##print(object_id_returned)
        
        
        api_test_query_channel = "api/1.0/config/element/channel"
        
        obj_id, beam_id, dictionary1, dictionary2, key_max_id, dictionary_list_igp_id, igp_obj_id_created  = get_max_bw_channel_and_beam_id(nms_ip, nms_username, nms_password, api_test_query_channel,direction= "upstream")
        
        dictionary_test = OrderedDict()
        dictionary_list_values = []
        dictionary_list_combined = []
        down_list = []

        down_list.append(key_max_id)
        print ("THE CHANNEL ID FOR THE MAXIMUM BW IS>>>", down_list)
        #Body insertion for the request posted to the NMS GUI
        return_type = 1
        body1, body2 = request_body_insertion(obj_id, return_type)
        
        #Object created to create the instance of the post request class
        
        post_obj = Post_Request_Reverse_Carrier(nms_ip, nms_username, nms_password, obj_id)
        
        return_post = post_obj.request_for_upstream(body1, "api/1.0/config/element/upstreamcarrier")[1]
        return_post = str(return_post)
        print ("returned_obj_type", return_post)

        body1, body2 = request_body_insertion(obj_id, return_post) 

        return_obj_type_inroutecomp = post_obj.request_for_composition(body2, "api/1.0/config/element/inroutegroupcomposition")

        print ("The return value for the inroutegrou", return_obj_type_inroutecomp)

        headers = {"Content-Type":"application/json"}
        
        
        body = """{"id": "%s"}"""%(obj_id)
        
        
        
        api_test_query_apply_changes_uri = "api/1.0/config/element/%s/apply_changes" %(obj_id)
        api_uri_apply_changes = 'http://{api_endpoint}/{nms_api}'.format(api_endpoint=nms_ip, nms_api=api_test_query_apply_changes_uri)

        print "APPLY CHANGES RESPONSE", api_uri_apply_changes
        get_info_request_channel_info = requests.post(api_uri_apply_changes, data=body, auth=HTTPBasicAuth(nms_username, nms_password), headers=headers) 

        print "The JSON response for POST", get_info_request_channel_info.json()
        
        

        

        nms_obj = Patch_Post_Apply_Changes_IGP(nms_ip, nms_username, nms_password)

        list_inet_beam_id, list_beam_list_inet_id, dictionary_d = nms_obj.get_inet_info_from_beam()

        print("<<THE DICTIONARY LIST VALUE>>>", list_inet_beam_id)

        print("<<< THE DICTIONARY LIST INET >> ", list_beam_list_inet_id)

        print("<<<THE DICTIONARY AFTER COMBINATION>>", dictionary_d)
        
        '''Generate the wokbook object for the class'''
        import openpyxl
        work_b = openpyxl.Workbook()
        work_book_name = os.path.expanduser("~\\Upstream_Excel_File.xlsx") 
        work_b.save(work_book_name) 

        

        #gw_frequency_ds, user_frequency_ds = get_user_gw_frequency_of_downstream_for_skipping(nms_ip, nms_username, nms_password, key_max_id)
        #gw_freq_ds_keys_list = "".join(map(str, gw_frequency_ds))
        #usr_freq_ds_values_list = "".join(map(str, user_frequency_ds))

        channel_value, fw_gw_user_tuple_list, fw_bw_obj_id, fw_gw_freq, fw_user_freq = downstream_modified_gw_user_pair.get_channel_info(nms_ip, nms_username, nms_password)

        print ("THE FW GW FREQUENCY>>>", fw_gw_freq)
        fw_gw_freq_str = fw_gw_freq.encode("utf-8")
        print ("THE FW GW FREQUENCY str>>>", fw_gw_freq_str)
        fw_gw_freq_int = "".join(fw_gw_freq_str)
        print ("THE FW GW FREQUENCY str>>>", fw_gw_freq_int)
        
        print("THE FW USER FREQUENCY>>>", fw_user_freq)
        fw_user_freq_str = fw_user_freq.encode("utf-8")
        print("THE FW USER FREQUENCY str>>>", fw_user_freq_str)
        fw_user_freq_int = "".join(fw_user_freq_str) 
        print("THE FW USER FREQUENCY str>>>", fw_user_freq_int)

        print ("THE RW GW-USER PAIR>>>", gw_user_pair)
        
        fw_user_freq_list = []
        fw_gw_freq_list = [] 
        
        fw_user_freq_list.append(fw_user_freq)
        fw_gw_freq_list.append(fw_gw_freq)

        tuple_list = (fw_gw_freq_int, fw_user_freq_int)
        gw_user_tuple = tuple(tuple_list)

        print("<<<<THE DOWNSTREAM GW USER TUPLE>>>", gw_user_tuple)
        enumerate_list = []
        modcod_list_length = 10
        
        #Ping terminals
        for did in pp_did_list:
            terminal_list = term_did_list[did]
            print("The terminal list is GIVEN BY>>>", terminal_list)
            terminal_ip = term_ip_list[str(term_did_list[did])]
            ping_output_for_terminals, ping_error_terminals = ping_terminals_from_nms(nms_ip, nms_username, nms_password, terminal_ip) 
            print ("THE PING SUCCEEDED....WE ARE PROCEEDING WITH THE TEST....")
            print ("---------------------------------------------------------")
            if ((not ping_error_terminals and ping_output_for_terminals == "") or ("ttl=" in ping_error_terminals[0] and "ttl=" in ping_error_terminals[1]) or "ttl=" in ping_error_terminals[-5]):
                print ("THE PING SUCCEEDED....WE ARE PROCEEDING WITH THE TEST....")
                print ("---------------------------------------------------------")

            #if "ttl=" not in ping_output_for_terminals:
            else:
                print("THE SAT0 IP OF TERMINAL IS NOT PINGABLE....EXITING....")
                sys.exit(-1)

        for carrier in down_list:
            print("------------------------------------")
            print("<<<OBJ ID IS GIVEN BY>>>>", carrier)
                # for channel in c_list:
                #     print("<<<<<<<<<<<THE CARRIER FREQUNCY>>>>>>")
                #     print("-------------------------------------")
                #     print("THE CARRIER FREQUENCY IS GIVEN BY.....", channel)
            i=1
            for j, (gw, u_freq) in enumerate(gw_user_pair):
                if (str(gw), str(u_freq)) == (fw_gw_freq_int, fw_user_freq_int):
                    print("I AM IN EQUALALITY CONDITION CHECK>>>>")
                    for i, (gw, u_freq) in enumerate(gw_user_pair):
                        print("I AM IN FOR LOOP FOR EQUALITY....")
                        gw_user_pair = gw_user_pair[i+1]
                        gw, u_freq = gw_user_pair
                        print("THE USER FREQ>>>", u_freq)
                        print("THE GW FREQ>>>", gw)
                        break 
                enumerate_list.append(j)
                print ("THE LENGTH OF THE ENUMERATE LIST>>>>", len(enumerate_list))
                
                if i == i+1:
                    modcod_list_length+=10
                print("<<<GW FREQUENCY IS>>", gw)
                change_gw_frequency(nms_ip, nms_username, nms_password, key_max_id, str(gw))
                apply_changes(nms_ip, nms_username, nms_password, key_max_id)
                print("<<<<USER_FREQUENCY IS>>>>", u_freq)
                change_user_frequency(nms_ip, nms_username, nms_password, key_max_id, str(u_freq))
                apply_changes(nms_ip, nms_username, nms_password, key_max_id) 
                active_object_state = openpyxl.load_workbook(work_book_name)
                sheet_name = "Chart%s,%s"%(gw, u_freq)
                work_s2 = active_object_state.create_sheet(sheet_name)  
                for symbol in symbols:
                    print("The sleep time", time_set)
                    #while (time_set):
                    print("<<<<<<<<SYMBOL LIST HIT>>>>>>>>>", symbol)
                    obj_id_upstream = change_symbol_rate(nms_ip, nms_username, nms_password, obj_id, symbol)
                    apply_changes(nms_ip, nms_username, nms_password, obj_id) 
            # Change maxmodcod
                    for modcod in modcod_list:
                        print("<<<MOD COD >>>", modcod)
                        print("<<<<<?THE OBJECT ID OF THE IGP IS>>>>>>?", obj_id)
                        if modcod == "BPSK":
                            print("I AM IN BPSK")
                            symbol_rate, modulation_set, fec_rate, upstream_carrier_id = nms_obj.get_IGP_details_created_from_post(obj_id)
                            change_reference_symbol_rate_cn(nms_ip, nms_username, nms_password, key_max_id, symbol, modcod) 
                            modcod_change_fec_iteration(fec_rate_bpsk_list, nms_ip, nms_username,nms_password, obj_id, upstream_carrier_id, modcod, dictionary_d, beam_id, pp_validation_check, ssh_console_obj_stdout_IP, symbol_rate, modulation_set, fec_rate, pp_ip, pp_user, pp_password, console_port_str, ssh_console_DC_PP1, term_did_list, pp_did_list, term_ip_list, pp_ip_list, work_b, carrier, gw, u_freq, dictionary_test, dictionary_list_combined, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list ,fec_rate_qpsk_list, symbols)
                        elif modcod == "8PSK":
                            print("I AM IN 8PSK")
                            symbol_rate, modulation_set, fec_rate, upstream_carrier_id = nms_obj.get_IGP_details_created_from_post(obj_id)
                            change_reference_symbol_rate_cn(nms_ip, nms_username, nms_password, key_max_id, symbol, modcod) 
                            modcod_change_fec_iteration(fec_rate_8psk_list, nms_ip, nms_username,nms_password, obj_id, upstream_carrier_id, modcod, dictionary_d, beam_id, pp_validation_check, ssh_console_obj_stdout_IP, symbol_rate, modulation_set, fec_rate, pp_ip, pp_user, pp_password, console_port_str, ssh_console_DC_PP1, term_did_list, pp_did_list, term_ip_list, pp_ip_list, work_b, carrier, gw, u_freq, dictionary_test, dictionary_list_combined, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols)
                        elif modcod == "QPSK":
                            print("I AM IN QPSK")
                            symbol_rate, modulation_set, fec_rate, upstream_carrier_id = nms_obj.get_IGP_details_created_from_post(obj_id)
                            change_reference_symbol_rate_cn(nms_ip, nms_username, nms_password, key_max_id, symbol, modcod) 
                            modcod_change_fec_iteration(fec_rate_qpsk_list, nms_ip, nms_username,nms_password, obj_id, upstream_carrier_id, modcod, dictionary_d, beam_id, pp_validation_check, ssh_console_obj_stdout_IP, symbol_rate, modulation_set, fec_rate, pp_ip, pp_user, pp_password, console_port_str, ssh_console_DC_PP1, term_did_list, pp_did_list, term_ip_list, pp_ip_list, work_b, carrier, gw, u_freq, dictionary_test, dictionary_list_combined, modcod_list_length, enumerate_list, work_s2, active_object_state, sheet_name, fec_final_rate, fec_rate_bpsk_list,fec_rate_8psk_list,fec_rate_qpsk_list, symbols)
                    print("The time set>>", time_set)
                        #thread 
                    i+=1
                        


        revert_config_obj = Revert_Config_Changes(nms_ip, nms_username, nms_password)
        print ("THE OBJ ID FOR THE IGP BEFORE THE RESTORE IS", obj_id)
        revert_config_obj.delete_igp_created(obj_id)
        revert_config_obj.apply_changes_deleted_igp(obj_id)
        channel_list_returned = revert_config_obj.revert_back_the_beam_for_old_IGP(dictionary_list_igp_id, beam_id)
        revert_config_obj.apply_changes_for_reattaching_beam(channel_list_returned)
        print("PLEASE CLOSE THE CONSOLE TO EXIT TO THE CMD PROMPT.....")
        
        """
            for item in list_beam_list_inet_id:
                item_str = "".join(item)
                item_stripped = item_str[9:]
                print "The item stripped..", item_stripped
        """
    
    except KeyboardInterrupt:
        print("IN KEYBOARD EXCEPTION>>>>")
        revert_config_obj = Revert_Config_Changes(nms_ip, nms_username, nms_password)
        revert_config_obj.delete_igp_created(obj_id)
        revert_config_obj.apply_changes_deleted_igp(obj_id)
        channel_list_returned = revert_config_obj.revert_back_the_beam_for_old_IGP(dictionary_list_igp_id, beam_id)
        revert_config_obj.apply_changes_for_reattaching_beam(channel_list_returned)
        print("PLEASE CLOSE THE CONSOLE TO EXIT TO THE CMD PROMPT.....")
    except ValueError:
        print("IN EXCEPTION>>")
        print("PLEASE CLOSE THE CONSOLE TO EXIT TO THE CMD PROMPT.....")
        sys.exit(-1)
    #except:
     #   revert_config_obj = Revert_Config_Changes(nms_ip, nms_username, nms_password)
      #  revert_config_obj.delete_igp_created(obj_id)
       # revert_config_obj.apply_changes_deleted_igp(obj_id)
        #channel_list_returned = revert_config_obj.revert_back_the_beam_for_old_IGP(dictionary_list_igp_id, beam_id)
        #revert_config_obj.apply_changes_for_reattaching_beam(channel_list_returned)
    
    #finally:
    #    revert_config_obj = Revert_Config_Changes(nms_ip, nms_username, nms_password)
    #    revert_config_obj.delete_igp_created(igp_obj_id_created)
    #    revert_config_obj.apply_changes_deleted_igp(igp_obj_id_created)
    #    channel_list_returned = revert_config_obj.revert_back_the_beam_for_old_IGP(dictionary_list_igp_id, beam_id)
    #    revert_config_obj.apply_changes_for_reattaching_beam(channel_list_returned)


    

    #pp_validation_check.

    #print("The std out string is given by", ssh_console_obj_stdout_IP)

if __name__ == "__main__":
    filename = "C:\\Users\\akarthik\\Downloads\\config_upstream.json"
    main(nms_Ip, nms_user, nms_pass, pp_Ip, pp_User, pp_pass, symbol_list)
    func_include_filename(nms_ip, nms_user, nms_pass, pp_ip, pp_user, pp_pass, symbol_list)
    """
    
    data = json.load(open(filename))
    nms_ip=data['nms_ip']
    nms_username= data['nms_username']
    nms_password= data['nms_password']
    igp_name=data['igp_name']
    api_test_query = "api/1.0/config/element/channel" 
    get_max_bw_channel_and_beam_id(nms_ip, nms_username, nms_password, api_test_query, direction="upstream") 
    """